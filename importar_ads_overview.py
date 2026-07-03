#!/usr/bin/env python3
"""
Rhode — Importa o CUSTO REAL do GMV Max do export "Campaign overview data" (Ads Manager).

O export "creative data for product campaigns" (importar_gmvmax.py) mostra custo=0 nas
campanhas Vendas Líquidas. Já o "Campaign overview data" traz a coluna "Custo do GMV
máximo" = gasto real por dia (todas as campanhas). Esse é o custo de mídia pago na conta
de ads (fora do settlement do seller) → precisa entrar no lucro.

Granularidade: 1 linha por dia (id = data). Idempotente (upsert merge-duplicates).
Uso:  python3 importar_ads_overview.py                      # lê dados/campanhas/tiktok/
      python3 importar_ads_overview.py <arquivo.xlsx> ...   # arquivos específicos
"""
import os, sys, json, glob, urllib.request, urllib.error
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
SB_URL = os.environ["SUPABASE_URL"]; SB_KEY = os.environ["SUPABASE_SERVICE_KEY"]
SBH = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type": "application/json"}
DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dados", "campanhas", "tiktok")


def fnum(x):
    try: return float(str(x).replace(",", "."))
    except (ValueError, TypeError): return 0.0


def upsert(rows, chunk=500):
    if not rows: return
    url = f"{SB_URL}/rest/v1/ads_custo?on_conflict=id"
    for i in range(0, len(rows), chunk):
        req = urllib.request.Request(url, data=json.dumps(rows[i:i+chunk]).encode(),
            headers={**SBH, "Prefer": "resolution=merge-duplicates,return=minimal"}, method="POST")
        try:
            urllib.request.urlopen(req, timeout=120)
        except urllib.error.HTTPError as e:
            print(f"  [ERRO upsert] {e.code}: {e.read()[:300]}"); raise
    print(f"  ✓ ads_custo: {len(rows)} dias upsertados")


def parse(fn):
    wb = openpyxl.load_workbook(fn); ws = wb.active
    hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(hdr)}
    col_dia = idx.get("Por dia", 0)
    col_custo = next((i for h, i in idx.items() if h and "Custo do GMV" in str(h)), None)
    col_rec = next((i for h, i in idx.items() if h and "Receita" in str(h)), None)
    col_ped = next((i for h, i in idx.items() if h and "Pedidos de SKU" in str(h)), None)
    col_roi = idx.get("ROI")
    if col_custo is None:
        print(f"  ⚠ {os.path.basename(fn)}: sem coluna de custo — pulando"); return []
    rows = []
    for ri in range(2, ws.max_row + 1):
        vals = [ws.cell(row=ri, column=c).value for c in range(1, ws.max_column + 1)]
        d = vals[col_dia]
        if not d or "Total" in str(d): continue
        try:
            dt = d.date() if hasattr(d, "date") else datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
        except (ValueError, TypeError): continue
        rows.append({
            "id": dt.isoformat(), "data": dt.isoformat(), "periodo": dt.strftime("%Y-%m"),
            "custo": round(fnum(vals[col_custo]), 2),
            "receita": round(fnum(vals[col_rec]) if col_rec is not None else 0, 2),
            "pedidos": int(fnum(vals[col_ped]) if col_ped is not None else 0),
            "roi": round(fnum(vals[col_roi]) if col_roi is not None else 0, 2),
        })
    return rows


def main():
    files = sys.argv[1:] or sorted(glob.glob(os.path.join(DIR, "Campaign overview data*.xlsx")))
    if not files:
        print("Nenhum 'Campaign overview data*.xlsx' encontrado em dados/campanhas/tiktok/"); return
    all_rows = {}
    for fn in files:
        rows = parse(fn)
        for r in rows: all_rows[r["id"]] = r  # dedup por dia (último arquivo vence)
        tot = sum(r["custo"] for r in rows)
        print(f"  {os.path.basename(fn)[:50]}: {len(rows)} dias · custo R$ {tot:,.2f}")
    rows = list(all_rows.values())
    print(f"\n  → {len(rows)} dias · custo total R$ {sum(r['custo'] for r in rows):,.2f}")
    upsert(rows)
    print("  ✓ concluído.")


if __name__ == "__main__":
    main()
