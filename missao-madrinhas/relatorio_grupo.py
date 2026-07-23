#!/usr/bin/env python3
"""
Missão das Madrinhas — RELATÓRIO DO GRUPO (snapshot + entradas/saídas + conciliação).

O que faz:
  1. Lê os telefones que estão NO GRUPO agora (Z-API).
  2. Salva um snapshot. Comparando com o snapshot anterior → QUEM ENTROU e QUEM SAIU.
  3. (Opcional) Cruza com as respostas do Form B → válidas por madrinha,
     quem registrou e não entrou, e quem está no grupo sem ter registrado.
  4. Gera uma planilha .xlsx com tudo separado em abas.

USO:
  # 1) achar o ID do grupo (só na 1ª vez)
  python3 missao-madrinhas/relatorio_grupo.py --listar-grupos

  # 2) snapshot + relatório (rode quando quiser; quanto mais rodar, melhor o histórico)
  python3 missao-madrinhas/relatorio_grupo.py --grupo <ID_DO_GRUPO>

  # 3) com conciliação por madrinha (exporte as respostas do Form B em CSV antes)
  python3 missao-madrinhas/relatorio_grupo.py --grupo <ID> --respostas respostas.csv

Credenciais: ZAPI_INSTANCE / ZAPI_TOKEN / ZAPI_CLIENT_TOKEN no .env (não vão pro código).
"""
import os, re, sys, csv, json, glob, argparse, datetime
from collections import defaultdict
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SNAP_DIR = os.path.join(ROOT, 'missao-madrinhas', 'snapshots')

# ───────────────────────────── infra ─────────────────────────────
def load_env(path):
    env = {}
    if os.path.exists(path):
        for line in open(path, encoding='utf-8'):
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip().strip('"').strip("'")
    return env

ENV = {**load_env(os.path.join(ROOT, '.env')), **os.environ}
INST, TOK, CT = ENV.get('ZAPI_INSTANCE'), ENV.get('ZAPI_TOKEN'), ENV.get('ZAPI_CLIENT_TOKEN')
BASE = f'https://api.z-api.io/instances/{INST}/token/{TOK}'
HDR = {'Client-Token': CT or ''}

def norm(p):
    """Telefone BR -> chave canônica (DDD + 8 últimos dígitos). Tolera +55 e o 9º dígito."""
    d = re.sub(r'\D', '', str(p or ''))
    if d.startswith('55') and len(d) >= 12:
        d = d[2:]
    if len(d) < 10:
        return None
    return d[:2] + d[-8:]

def fmt_tel(p):
    d = re.sub(r'\D', '', str(p or ''))
    if d.startswith('55') and len(d) >= 12:
        d = d[2:]
    if len(d) == 11:
        return f'({d[:2]}) {d[2:7]}-{d[7:]}'
    if len(d) == 10:
        return f'({d[:2]}) {d[2:6]}-{d[6:]}'
    return str(p)

# ───────────────────────────── Z-API ─────────────────────────────
def listar_grupos():
    r = requests.get(f'{BASE}/chats', headers=HDR, timeout=40); r.raise_for_status()
    print('ID do grupo'.ljust(32), 'Nome')
    print('-' * 64)
    n = 0
    for c in r.json():
        pid = str(c.get('phone', ''))
        if c.get('isGroup') or pid.endswith('-group') or pid.endswith('@g.us'):
            print(pid.ljust(32), c.get('name') or c.get('subject') or '')
            n += 1
    print(f'\n{n} grupos. Use o ID da coluna esquerda em --grupo.')

def participantes_do_grupo(group_id):
    r = requests.get(f'{BASE}/group-metadata/{group_id}', headers=HDR, timeout=40); r.raise_for_status()
    data = r.json()
    parts = data.get('participants') or data.get('groupParticipants') or []
    out = []
    for p in parts:
        ph = (p.get('phone') or p.get('id')) if isinstance(p, dict) else p
        admin = bool(p.get('isAdmin') or p.get('isSuperAdmin')) if isinstance(p, dict) else False
        if ph:
            out.append({'raw': str(ph), 'key': norm(ph), 'admin': admin})
    return [x for x in out if x['key']], data.get('subject', '(grupo)')

# ─────────────────────────── snapshots ───────────────────────────
def snapshot_anterior():
    os.makedirs(SNAP_DIR, exist_ok=True)
    arqs = sorted(glob.glob(os.path.join(SNAP_DIR, 'grupo_*.json')))
    if not arqs:
        return None
    return json.load(open(arqs[-1], encoding='utf-8'))

def salvar_snapshot(group_id, subject, parts, agora):
    os.makedirs(SNAP_DIR, exist_ok=True)
    caminho = os.path.join(SNAP_DIR, f'grupo_{agora:%Y%m%d_%H%M%S}.json')
    json.dump({'ts': agora.isoformat(), 'grupo': group_id, 'subject': subject,
               'participantes': [{'raw': p['raw'], 'key': p['key']} for p in parts]},
              open(caminho, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    return caminho

# ──────────────────────────── respostas ──────────────────────────
def ler_respostas(path):
    rows = list(csv.reader(open(path, encoding='utf-8-sig')))
    if not rows:
        sys.exit('CSV de respostas vazio.')
    head = [h.strip().lower() for h in rows[0]]
    iM = next((i for i, h in enumerate(head) if 'convidou' in h or 'madrinha' in h), None)
    iP = next((i for i, h in enumerate(head) if 'whatsapp' in h or 'telefone' in h or 'phone' in h), None)
    iN = next((i for i, h in enumerate(head) if 'nome' in h), None)
    if iM is None or iP is None:
        sys.exit('O CSV precisa das colunas "Quem te convidou" e "Seu WhatsApp".')
    out = []
    for r in rows[1:]:
        if len(r) <= max(i for i in [iM, iP, iN or 0]):
            continue
        madr = (r[iM] or '').strip()
        tel = (r[iP] or '').strip()
        if madr and norm(tel):
            out.append({'madrinha': madr, 'tel': tel, 'key': norm(tel),
                        'nome': (r[iN] or '').strip() if iN is not None else ''})
    return out

# ───────────────────────────── planilha ──────────────────────────
RED, INDIGO, YEL = 'FE2C55', '1B3A5C', 'F2D413'
thin = Side(style='thin', color='D9D9D9')
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def _head(ws, row, cols, fill=INDIGO):
    for i, c in enumerate(cols, 1):
        cel = ws.cell(row=row, column=i, value=c)
        cel.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
        cel.fill = PatternFill('solid', fgColor=fill)
        cel.alignment = Alignment(horizontal='left', vertical='center')
        cel.border = BORD

def _rows(ws, start, dados):
    for r, linha in enumerate(dados, start):
        for c, val in enumerate(linha, 1):
            cel = ws.cell(row=r, column=c, value=val)
            cel.font = Font(name='Arial', size=10)
            cel.border = BORD

def montar_xlsx(caminho, ctx):
    wb = Workbook()

    # ---- Resumo
    ws = wb.active; ws.title = 'Resumo'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 42; ws.column_dimensions['B'].width = 16
    ws['A1'] = 'RELATÓRIO DO GRUPO — Missão das Madrinhas'
    ws['A1'].font = Font(name='Arial', bold=True, size=15, color=RED)
    ws['A2'] = f"{ctx['subject']} · gerado em {ctx['agora']:%d/%m/%Y %H:%M}"
    ws['A2'].font = Font(name='Arial', italic=True, size=9, color='666666')
    ws['A3'] = ctx['ref_txt']
    ws['A3'].font = Font(name='Arial', italic=True, size=9, color='666666')

    kpis = [
        ('Pessoas no grupo agora', ctx['total_grupo']),
        ('Entraram desde o snapshot anterior', '—' if ctx.get('baseline') else ctx['n_entradas']),
        ('Saíram desde o snapshot anterior', '—' if ctx.get('baseline') else ctx['n_saidas']),
    ]
    if ctx['tem_resp']:
        kpis += [
            ('Registros no Form B (telefones únicos)', ctx['n_reg']),
            ('VÁLIDAS (registrou e está no grupo)', ctx['n_validas']),
            ('Registrou mas NÃO está no grupo', ctx['n_fora']),
            ('No grupo SEM registro no formulário', ctx['n_sem_reg']),
        ]
    r = 5
    for k, v in kpis:
        ws.cell(row=r, column=1, value=k).font = Font(name='Arial', bold=True, size=10)
        c = ws.cell(row=r, column=2, value=v)
        c.font = Font(name='Arial', bold=True, size=14)
        c.alignment = Alignment(horizontal='center')
        c.fill = PatternFill('solid', fgColor='F5F5F5')
        c.border = BORD
        r += 1

    # ---- Participantes
    ws = wb.create_sheet('Participantes no grupo')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCD', (20, 26, 22, 18)): ws.column_dimensions[col].width = w
    _head(ws, 1, ['Telefone', 'Nome (do formulário)', 'Veio pela madrinha', 'Registrou?'])
    _rows(ws, 2, ctx['tab_participantes']); ws.freeze_panes = 'A2'

    # ---- Entradas / Saídas
    for titulo, dados, cor in (('Entradas', ctx['tab_entradas'], '1F9D57'),
                               ('Saidas', ctx['tab_saidas'], 'B00000')):
        ws = wb.create_sheet(titulo)
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABC', (20, 26, 22)): ws.column_dimensions[col].width = w
        _head(ws, 1, ['Telefone', 'Nome (do formulário)', 'Madrinha'], fill=cor)
        _rows(ws, 2, dados or [['— nenhuma —', '', '']]); ws.freeze_panes = 'A2'

    # ---- Conciliação por madrinha
    if ctx['tem_resp']:
        ws = wb.create_sheet('Conciliação por madrinha')
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABCD', (30, 14, 14, 14)): ws.column_dimensions[col].width = w
        _head(ws, 1, ['Madrinha', 'Registradas', 'VÁLIDAS (no grupo)', 'Fora do grupo'])
        _rows(ws, 2, ctx['tab_conc']); ws.freeze_panes = 'A2'

        ws = wb.create_sheet('Registrou e não entrou')
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABC', (20, 26, 30)): ws.column_dimensions[col].width = w
        _head(ws, 1, ['Telefone', 'Nome', 'Madrinha'], fill='B00000')
        _rows(ws, 2, ctx['tab_fora'] or [['— nenhuma —', '', '']]); ws.freeze_panes = 'A2'

    # ---- Histórico
    ws = wb.create_sheet('Histórico')
    ws.sheet_view.showGridLines = False
    for col, w in zip('ABCD', (22, 16, 14, 14)): ws.column_dimensions[col].width = w
    _head(ws, 1, ['Data/hora', 'No grupo', 'Entraram', 'Saíram'])
    _rows(ws, 2, ctx['tab_hist']); ws.freeze_panes = 'A2'

    wb.save(caminho)

# ─────────────────────────────── main ────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--grupo'); ap.add_argument('--respostas')
    ap.add_argument('--listar-grupos', action='store_true')
    ap.add_argument('--saida')
    a = ap.parse_args()

    if not (INST and TOK and CT):
        sys.exit('Faltam ZAPI_INSTANCE / ZAPI_TOKEN / ZAPI_CLIENT_TOKEN no .env')
    if a.listar_grupos:
        return listar_grupos()
    if not a.grupo:
        sys.exit('uso: --grupo <ID> [--respostas respostas.csv]   (ou --listar-grupos)')

    agora = datetime.datetime.now()
    parts, subject = participantes_do_grupo(a.grupo)
    atual = {p['key']: p['raw'] for p in parts}

    ant = snapshot_anterior()
    baseline = ant is None
    if ant:
        antes = {p['key']: p['raw'] for p in ant['participantes']}
        ref_txt = f"Comparado com o snapshot de {ant['ts'][:16].replace('T', ' ')}"
        entradas = [k for k in atual if k not in antes]
        saidas = [k for k in antes if k not in atual]
    else:
        antes, entradas, saidas = {}, [], []
        ref_txt = ('Primeira execução — este é o snapshot BASE (foto inicial do grupo). '
                   'Entradas/saídas passam a ser medidas a partir da próxima execução.')

    # conciliação
    resp = ler_respostas(a.respostas) if a.respostas else []
    tem_resp = bool(resp)
    por_key = {}
    for x in resp:
        por_key.setdefault(x['key'], x)   # 1º registro daquele telefone
    reg_keys = set(por_key)

    validas = reg_keys & set(atual)
    fora = reg_keys - set(atual)
    sem_reg = set(atual) - reg_keys

    conc = defaultdict(lambda: {'reg': set(), 'val': set()})
    for k, x in por_key.items():
        conc[x['madrinha']]['reg'].add(k)
        if k in atual:
            conc[x['madrinha']]['val'].add(k)
    tab_conc = sorted(([m, len(v['reg']), len(v['val']), len(v['reg']) - len(v['val'])]
                       for m, v in conc.items()), key=lambda r: -r[2])

    def info(k, fallback_raw=''):
        x = por_key.get(k)
        return (fmt_tel(atual.get(k) or antes.get(k) or fallback_raw),
                x['nome'] if x else '', x['madrinha'] if x else '')

    tab_participantes = [[fmt_tel(raw),
                          por_key.get(k, {}).get('nome', ''),
                          por_key.get(k, {}).get('madrinha', ''),
                          'sim' if k in reg_keys else 'não']
                         for k, raw in sorted(atual.items())]
    tab_entradas = [list(info(k)) for k in entradas]
    tab_saidas = [list(info(k, antes.get(k, ''))) for k in saidas]
    tab_fora = [[fmt_tel(por_key[k]['tel']), por_key[k]['nome'], por_key[k]['madrinha']] for k in sorted(fora)]

    # histórico (todos os snapshots + este)
    tab_hist = []
    for arq in sorted(glob.glob(os.path.join(SNAP_DIR, 'grupo_*.json'))):
        try:
            s = json.load(open(arq, encoding='utf-8'))
            tab_hist.append([s['ts'][:16].replace('T', ' '), len(s['participantes']), '', ''])
        except Exception:
            pass
    tab_hist.append([f'{agora:%Y-%m-%d %H:%M}', len(atual), len(entradas), len(saidas)])

    salvar_snapshot(a.grupo, subject, parts, agora)

    saida = a.saida or os.path.join(ROOT, 'missao-madrinhas', f'Relatorio Grupo Madrinhas_{agora:%Y-%m-%d}.xlsx')
    montar_xlsx(saida, {
        'subject': subject, 'agora': agora, 'ref_txt': ref_txt,
        'total_grupo': len(atual), 'n_entradas': len(entradas), 'n_saidas': len(saidas), 'baseline': baseline,
        'tem_resp': tem_resp, 'n_reg': len(reg_keys), 'n_validas': len(validas),
        'n_fora': len(fora), 'n_sem_reg': len(sem_reg),
        'tab_participantes': tab_participantes, 'tab_entradas': tab_entradas,
        'tab_saidas': tab_saidas, 'tab_conc': tab_conc, 'tab_fora': tab_fora, 'tab_hist': tab_hist,
    })

    print(f'\nGrupo: {subject}')
    print(f'  no grupo agora ......... {len(atual)}')
    print('  entraram ............... ' + ('— (baseline)' if baseline else str(len(entradas))))
    print('  saíram ................. ' + ('— (baseline)' if baseline else str(len(saidas))))
    if tem_resp:
        print(f'  registros (Form B) ..... {len(reg_keys)}')
        print(f'  VÁLIDAS (reg + grupo) .. {len(validas)}')
        print(f'  registrou e não entrou . {len(fora)}')
        print(f'  no grupo sem registro .. {len(sem_reg)}')
        if tab_conc:
            print('\n  ranking por válidas:')
            for m, reg, val, f in tab_conc[:10]:
                print(f'    {val:>3} válidas ({reg} reg) — {m}')
    print(f'\n📄 Planilha: {saida}')
    print(f'   {ref_txt}')

if __name__ == '__main__':
    main()
