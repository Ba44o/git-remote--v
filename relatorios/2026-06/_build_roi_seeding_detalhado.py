"""
Build: ROI Seeding detalhado (Rhode Jeans × TikTok Shop) — FONTE REAL DA API
============================================================================
Fonte de verdade das peças = endpoint de free sample do TikTok Shop:
  POST /affiliate_seller/202409/sample_applications/search
  → warehouse/raw_sample_applications.csv (1 linha por pedido de amostra:
     creator × SKU × status).

Peça "realmente enviada" = status em {COMPLETED, SHIPPED, CONTENT_PENDING}.
GMV atribuído (por COORTE) = TODO o GMV de afiliada da creator (raw_creator_product,
fonte Affiliate Orders) na janela [1º convite do mês, próximo convite dela). Cada
creator entra no mês em que foi convidada (não no 1º convite all-time), então a linha
de cada mês reflete o retorno do seeding daquele mês. ROI/ROAS em fórmula viva.

Saída: relatorios/2026-06/ROI Seeding 2025-2026.xlsx
"""
from pathlib import Path
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[2]
WH = BASE / "warehouse"
OUT = BASE / "relatorios/2026-06/ROI Seeding 2025-2026.xlsx"

SENT = {"COMPLETED", "SHIPPED", "CONTENT_PENDING"}  # peça de fato despachada

# ---- dados -----------------------------------------------------------------
HOJE = date.today().strftime("%Y-%m-%d")          # ATTR_FIM: fim da janela de atribuição

led = pd.read_csv(WH / "raw_sample_applications.csv", dtype={"product_id": str, "sku_id": str})
led["ck"] = led.creator.str.lower().str.strip()
led["sent"] = led.status.isin(SENT)
led["cd"] = led.convite_data.astype(str)
led["has_date"] = led.convite_data.notna() & (led.cd.str.len() >= 10)
led["cm"] = led.cd.str[:7].where(led.has_date, "—")   # mês do convite (— = sem data)

cp = pd.read_csv(WH / "raw_creator_product.csv", dtype={"product_id": str})
cp["ck"] = cp.creator.str.lower().str.strip()
CP_INI, CP_FIM = cp.data.min(), cp.data.max()      # janela real de vendas coletada

# ── ATRIBUIÇÃO POR COORTE (creator × mês do convite) ────────────────────────
# Cada creator entra no MÊS EM QUE FOI CONVIDADA (não no 1º convite all-time).
# GMV = TODO o GMV de afiliada dela (todos os produtos, fonte Affiliate Orders)
# na janela [1º convite do mês, próximo convite dela) — capada no convite seguinte
# para cada venda cair na coorte de seeding mais recente (soma ADITIVA, sem overlap;
# o TOTAL = cada creator do 1º convite dela até hoje). A última coorte vai até HOJE.
# Mesma lógica de atribuição do relatório mensal detalhado (seeding_report.py).
def _rate(s):
    m = s.mode()
    return m.iloc[0] if len(m) else None

recs = []
for ck, sub in led.groupby("ck"):
    creator = sub.creator.iloc[0]; nickname = sub.nickname.iloc[0]
    rate = _rate(sub.commission_rate)
    cpck = cp[cp.ck == ck]
    dated = sub[sub.has_date]
    starts = sorted(dated.groupby("cm").cd.min().to_dict().items())   # [(mes, d_ini)] cronológico
    sd = [d[:10] for _, d in starts]
    for i, (m, dini) in enumerate(starts):
        d0 = dini[:10]
        smm = dated[dated.cm == m]
        env = int(smm.sent.sum()); ped_am = len(smm)
        if i + 1 < len(sd):                       # janela capada no próximo convite da creator
            win = cpck[(cpck.data >= d0) & (cpck.data < sd[i + 1])]
        else:                                     # última coorte: aberta até HOJE
            win = cpck[cpck.data >= d0]
        recs.append(dict(
            ck=ck, creator=creator, nickname=nickname, rate=rate, mes=m,
            conv_ini=d0, conv_fim=smm.cd.max()[:10],
            pedidas=ped_am, enviadas=env, canceladas=ped_am - env,
            skus=int(smm.product_id.nunique()),
            gmv=round(float(win.gmv.sum()), 2), com=round(float(win.comissao.sum()), 2),
            ped=int(win.pedidos.sum()),
        ))
    und = sub[~sub.has_date]                       # convites sem data: contam peças, mas GMV=0 (não dá pra atribuir)
    if len(und):
        env = int(und.sent.sum()); ped_am = len(und)
        recs.append(dict(
            ck=ck, creator=creator, nickname=nickname, rate=rate, mes="—",
            conv_ini="", conv_fim="",
            pedidas=ped_am, enviadas=env, canceladas=ped_am - env,
            skus=int(und.product_id.nunique()), gmv=0.0, com=0.0, ped=0,
        ))

g = pd.DataFrame(recs)
g["rate"] = pd.to_numeric(g["rate"], errors="coerce").fillna(0.0)  # 0.08 → fração
# FAIXA ACORDADA: taxa nominal só em 8% ou 12% (<=11% → 8% ; >=11,5% → 12%)
g["faixa"] = g.rate.apply(lambda x: 0.08 if x < 0.115 else 0.12)
g["status_envio"] = g.enviadas.apply(lambda x: "Enviado" if x > 0 else "Nada enviado")
# ordena por DATA do 1º convite da coorte (mais antigo primeiro); sem data vai pro fim
g["_sortkey"] = g.conv_ini.replace("", "9999-99-99")
g = g.sort_values(["_sortkey", "gmv"], ascending=[True, False]).drop(columns="_sortkey").reset_index(drop=True)

# ---- estilos ---------------------------------------------------------------
NAVY = "1F2937"; ACCENT = "4F46E5"; LIGHT = "EEF2FF"; GREY = "F3F4F6"
GREEN = "ECFDF5"; AMBER = "FEF3C7"
hfont = Font(bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor=NAVY)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
money = 'R$ #,##0.00'; pct = '0.0%'

wb = openpyxl.Workbook()

# ============ ABA RESUMO ====================================================
ws = wb.active
ws.title = "Resumo"
ws.sheet_view.showGridLines = False
ws["B2"] = "ROI Seeding 2025/2026"
ws["B2"].font = Font(bold=True, size=16, color=NAVY)
ws["B3"] = ("Free sample REAL da API (sample_applications) · peça enviada = COMPLETED/SHIPPED/CONTENT_PENDING · "
            "GMV = afiliação da creator (Affiliate Orders) na janela do convite · coorte por MÊS do convite")
ws["B3"].font = Font(size=10, color="6B7280")

ws["B5"] = "Premissas (edite os valores em amarelo — tudo recalcula)"
ws["B5"].font = Font(bold=True, size=11, color=ACCENT)
prem = [
    ("Custo unitário da peça-amostra (COGS)", 40.0, money, "custo de produção de 1 peça"),
    ("Frete por envio (por creator)", 25.0, money, "logística de 1 remessa"),
    ("Margem de contribuição sobre GMV", 0.55, pct, "quanto do GMV vira lucro antes de comissão/seeding"),
]
for i, (lbl, val, fmt, note) in enumerate(prem):
    r = 6 + i
    ws[f"B{r}"] = lbl; ws[f"B{r}"].font = Font(size=10)
    c = ws[f"C{r}"]; c.value = val; c.number_format = fmt
    c.fill = PatternFill("solid", fgColor=AMBER); c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center"); c.border = border
    ws[f"D{r}"] = "← " + note; ws[f"D{r}"].font = Font(size=9, color="9CA3AF", italic=True)
PC = {"custo_unit": "Resumo!$C$6", "frete": "Resumo!$C$7", "margem": "Resumo!$C$8"}

ws["B10"] = "Resultado consolidado"
ws["B10"].font = Font(bold=True, size=11, color=ACCENT)
kpis = [
    "Creators que receberam amostra", "Peças enviadas (total)",
    "Peças pedidas (inclui canceladas)", "Peças canceladas (não enviadas)",
    "GMV atribuído (afiliação pós-convite)", "Comissão paga sobre seeding",
    "Custo total de seeding (peças+frete)", "Margem de contribuição gerada",
    "Lucro líquido do seeding", "ROAS (GMV ÷ custo seeding)",
    "ROI (lucro ÷ custo seeding)", "Custo por peça enviada",
]
kpi_row0 = 11
for i, lbl in enumerate(kpis):
    r = kpi_row0 + i
    ws[f"B{r}"] = lbl; ws[f"B{r}"].font = Font(size=10)
    ws[f"B{r}"].fill = PatternFill("solid", fgColor=GREY); ws[f"B{r}"].border = border
    cc = ws[f"C{r}"]; cc.border = border
    cc.alignment = Alignment(horizontal="right"); cc.font = Font(bold=True, size=10)
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 40

# ============ ABA DETALHE ===================================================
ws2 = wb.create_sheet("ROI por creator")
ws2.sheet_view.showGridLines = False
# (key, título, largura) — mapa nome→coluna torna as fórmulas robustas a inserções
COLS = [
    ("num", "#", 4), ("creator", "Creator", 22), ("handle", "@ TikTok", 20),
    ("status_envio", "Status envio", 13), ("mes", "Mês convite", 11),
    ("conv_ini", "1º convite", 12), ("conv_fim", "Último convite", 13),
    ("pedidas", "Peças pedidas", 10), ("enviadas", "Peças enviadas", 11),
    ("canceladas", "Peças canceladas", 11), ("skus", "SKUs amostrados", 11),
    ("rate", "Comissão %", 10), ("custo_am", "Custo amostra (R$)", 13),
    ("frete", "Frete (R$)", 10), ("custo_seed", "Custo seeding (R$)", 13),
    ("pedidos", "Pedidos atrib.", 10), ("gmv", "GMV atribuído (R$)", 16),
    ("com_paga", "Comissão paga (R$)", 14), ("margem", "Margem contrib. (R$)", 15),
    ("lucro", "Lucro líquido (R$)", 14), ("roas", "ROAS", 9), ("roi", "ROI", 9),
]
IDX = {k: i for i, (k, _, _) in enumerate(COLS, 1)}
def L(k): return get_column_letter(IDX[k])         # letra da coluna pelo nome
NC = len(COLS)
for j, (k, name, w) in enumerate(COLS, 1):
    c = ws2.cell(row=1, column=j, value=name)
    c.font = hfont; c.fill = hfill; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.row_dimensions[1].height = 30
ws2.freeze_panes = "D2"

center = ("num", "status_envio", "mes", "conv_ini", "conv_fim", "pedidas", "enviadas", "canceladas", "skus", "pedidos")
for i, row in g.iterrows():
    r = 2 + i
    vals = {
        "num": i + 1, "creator": row.nickname or row.creator, "handle": "@" + str(row.creator),
        "status_envio": row.status_envio, "mes": row.mes,
        "conv_ini": row.conv_ini or "—", "conv_fim": row.conv_fim or "—",
        "pedidas": int(row.pedidas), "enviadas": int(row.enviadas),
        "canceladas": int(row.canceladas), "skus": int(row.skus), "rate": float(row.faixa),
        "custo_am": f"={L('enviadas')}{r}*{PC['custo_unit']}",
        "frete": f"=IF({L('enviadas')}{r}>0,{PC['frete']},0)",
        "custo_seed": f"={L('custo_am')}{r}+{L('frete')}{r}",
        "pedidos": row.ped, "gmv": row.gmv, "com_paga": row.com,
        "margem": f"={L('gmv')}{r}*{PC['margem']}",
        "lucro": f"={L('margem')}{r}-{L('com_paga')}{r}-{L('custo_seed')}{r}",
        "roas": f"=IF({L('custo_seed')}{r}>0,{L('gmv')}{r}/{L('custo_seed')}{r},0)",
        "roi": f"=IF({L('custo_seed')}{r}>0,{L('lucro')}{r}/{L('custo_seed')}{r},0)",
    }
    for k, v in vals.items():
        c = ws2.cell(row=r, column=IDX[k], value=v)
        c.border = border; c.font = Font(size=9)
        if k == "rate": c.number_format = pct
        elif k == "roas": c.number_format = '#,##0.0"x"'
        elif k == "roi": c.number_format = pct
        elif k in ("custo_am", "frete", "custo_seed", "gmv", "com_paga", "margem", "lucro"):
            c.number_format = money
        if k in center: c.alignment = Alignment(horizontal="center")
        if k == "status_envio":
            c.font = Font(size=9, bold=True, color="047857" if v == "Enviado" else "B45309")
    if i % 2:
        for j in range(1, NC + 1):
            cell = ws2.cell(row=r, column=j)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
    # destaque leve da linha "Nada enviado"
    if row.status_envio == "Nada enviado":
        for j in range(1, NC + 1):
            ws2.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FEF2F2")

n = len(g); last = 1 + n; tr = last + 1
ws2.cell(row=tr, column=IDX["status_envio"], value="TOTAL (filtrado)").font = Font(bold=True)
sum_money = ("custo_seed", "gmv", "com_paga", "margem", "lucro")
sum_int = ("pedidas", "enviadas", "canceladas", "pedidos")
# SUBTOTAL(109,...) = soma só das linhas VISÍVEIS → o total responde ao autofiltro
for k in sum_money + sum_int:
    cell = ws2.cell(row=tr, column=IDX[k], value=f"=SUBTOTAL(109,{L(k)}2:{L(k)}{last})")
    cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor=LIGHT)
    cell.number_format = "0" if k in sum_int else money
ws2.cell(row=tr, column=IDX["roas"], value=f"=IF({L('custo_seed')}{tr}>0,{L('gmv')}{tr}/{L('custo_seed')}{tr},0)").number_format = '#,##0.0"x"'
ws2.cell(row=tr, column=IDX["roi"], value=f"=IF({L('custo_seed')}{tr}>0,{L('lucro')}{tr}/{L('custo_seed')}{tr},0)").number_format = pct
for k in ("roas", "roi"):
    ws2.cell(row=tr, column=IDX[k]).font = Font(bold=True)
    ws2.cell(row=tr, column=IDX[k]).fill = PatternFill("solid", fgColor=LIGHT)
# AUTOFILTRO no cabeçalho (cobre só as linhas de dados, não o TOTAL)
ws2.auto_filter.ref = f"A1:{get_column_letter(NC)}{last}"

# ---- KPIs no Resumo (agregados do TOTAL do programa — estáveis, independem do filtro) ----
D = "'ROI por creator'"
def RNG(k): return f"{D}!{L(k)}2:{L(k)}{last}"
S = lambda k: f"SUM({RNG(k)})"
n_cre_env = int(g.loc[g.enviadas > 0, "ck"].nunique())   # creators DISTINTAS atendidas (grão = coorte)
kf = [
    (n_cre_env, None),   # estático: nº de creators únicas (uma creator pode ter coortes em vários meses)
    (f"={S('enviadas')}", "0"),
    (f"={S('pedidas')}", "0"),
    (f"={S('canceladas')}", "0"),
    (f"={S('gmv')}", money),
    (f"={S('com_paga')}", money),
    (f"={S('custo_seed')}", money),
    (f"={S('margem')}", money),
    (f"={S('lucro')}", money),
    (f"=IF({S('custo_seed')}>0,{S('gmv')}/{S('custo_seed')},0)", '#,##0.0"x"'),
    (f"=IF({S('custo_seed')}>0,{S('lucro')}/{S('custo_seed')},0)", pct),
    (f"=IF({S('enviadas')}>0,{S('custo_seed')}/{S('enviadas')},0)", money),
]
for i, (f, fmt) in enumerate(kf):
    cc = ws["C" + str(kpi_row0 + i)]; cc.value = f
    if fmt: cc.number_format = fmt
for off in (9, 10):  # ROAS / ROI destaque
    ws[f"C{kpi_row0+off}"].fill = PatternFill("solid", fgColor=GREEN)
    ws[f"C{kpi_row0+off}"].font = Font(bold=True, size=11, color="047857")

# ============ ABA RESUMO POR MÊS ============================================
wsm = wb.create_sheet("Resumo por mês")
wsm.sheet_view.showGridLines = False
meses = sorted(g.mes.unique().tolist(), key=lambda x: (x == "—", x))
mcols = [("Mês", 11), ("Creators (convites)", 12), ("Peças enviadas", 12),
         ("Peças canceladas", 13), ("Custo seeding (R$)", 14), ("GMV atribuído (R$)", 16),
         ("Comissão paga (R$)", 14), ("Lucro líquido (R$)", 14), ("ROAS", 9), ("ROI", 9)]
for j, (name, w) in enumerate(mcols, 1):
    c = wsm.cell(row=1, column=j, value=name)
    c.font = hfont; c.fill = hfill; c.border = border
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wsm.column_dimensions[get_column_letter(j)].width = w
wsm.row_dimensions[1].height = 28
MES = L("mes")  # letra da coluna 'Mês convite' na aba detalhe
def SI(col, mr):  # SUMIF por mês → fica em sincronia com as premissas/detalhe
    return f"SUMIF({D}!{MES}2:{MES}{last},$A{mr},{D}!{L(col)}2:{L(col)}{last})"
for i, m in enumerate(meses):
    r = 2 + i
    wsm.cell(row=r, column=1, value=m)
    wsm.cell(row=r, column=2, value=f"=COUNTIF({D}!{MES}2:{MES}{last},$A{r})")
    wsm.cell(row=r, column=3, value=f"={SI('enviadas', r)}")
    wsm.cell(row=r, column=4, value=f"={SI('canceladas', r)}")
    wsm.cell(row=r, column=5, value=f"={SI('custo_seed', r)}").number_format = money
    wsm.cell(row=r, column=6, value=f"={SI('gmv', r)}").number_format = money
    wsm.cell(row=r, column=7, value=f"={SI('com_paga', r)}").number_format = money
    wsm.cell(row=r, column=8, value=f"={SI('lucro', r)}").number_format = money
    wsm.cell(row=r, column=9, value=f"=IF(E{r}>0,F{r}/E{r},0)").number_format = '#,##0.0"x"'
    wsm.cell(row=r, column=10, value=f"=IF(E{r}>0,H{r}/E{r},0)").number_format = pct
    for j in range(1, 11):
        cell = wsm.cell(row=r, column=j); cell.border = border; cell.font = Font(size=10)
        if j in (1, 2, 3, 4): cell.alignment = Alignment(horizontal="center")
    if i % 2:
        for j in range(1, 11):
            cell = wsm.cell(row=r, column=j)
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor="FAFAFA")
tm = 2 + len(meses)
wsm.cell(row=tm, column=1, value="TOTAL").font = Font(bold=True)
for col, letter in [(2, "B"), (3, "C"), (4, "D"), (5, "E"), (6, "F"), (7, "G"), (8, "H")]:
    cell = wsm.cell(row=tm, column=col, value=f"=SUM({letter}2:{letter}{tm-1})")
    cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor=LIGHT)
    if col >= 5: cell.number_format = money
wsm.cell(row=tm, column=9, value=f"=IF(E{tm}>0,F{tm}/E{tm},0)").number_format = '#,##0.0"x"'
wsm.cell(row=tm, column=10, value=f"=IF(E{tm}>0,H{tm}/E{tm},0)").number_format = pct
for col in (9, 10):
    wsm.cell(row=tm, column=col).font = Font(bold=True)
    wsm.cell(row=tm, column=col).fill = PatternFill("solid", fgColor=LIGHT)
wsm.freeze_panes = "A2"

# ============ ABA METODOLOGIA ===============================================
ws3 = wb.create_sheet("Metodologia")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["B"].width = 105
meto = [
    ("Metodologia & fontes", 14, True, NAVY),
    ("", 10, False, "000000"),
    ("Fonte das peças (100% API TikTok Shop)", 11, True, ACCENT),
    ("• Endpoint: POST /affiliate_seller/202409/sample_applications/search — pedidos de free sample do TikTok.", 10, False, "374151"),
    ("• 1 linha por creator × SKU pedido. Persistido em warehouse/raw_sample_applications.csv. total_count=337.", 10, False, "374151"),
    ("• 'Peça enviada' = status em COMPLETED, SHIPPED ou CONTENT_PENDING (peça de fato despachada).", 10, False, "374151"),
    ("• Status excluídos: *_CANCELLED (cancelada antes do envio) e AWAITING_SHIPMENT (aprovada, ainda não despachada).", 10, False, "B45309"),
    ("• 'Peças enviadas' por creator = nº de pedidos de amostra com status enviado. É o que vira CUSTO.", 10, False, "374151"),
    ("• '1º/Último convite' = create_time do pedido de amostra (via API de Order, GET /order/202309/orders). '—' = pedido fora da retenção da API.", 10, False, "374151"),
    ("• 'Status envio': Enviado (≥1 peça despachada) vs Nada enviado (pediu mas tudo cancelado/pendente). Use o AUTOFILTRO do cabeçalho; o TOTAL (SUBTOTAL) responde ao filtro.", 10, False, "374151"),
    ("", 10, False, "000000"),
    ("Comissão", 11, True, ACCENT),
    ("• 'Comissão %' = FAIXA acordada (8% ou 12%), consolidada das taxas reais da API (8/9/10% → 8%; 12% → 12%).", 10, False, "374151"),
    ("• 'Comissão paga (R$)' = comissão REAL das vendas (raw_creator_product), não taxa × GMV.", 10, False, "374151"),
    ("", 10, False, "000000"),
    ("Atribuição de GMV — por COORTE de convite (mês)", 11, True, ACCENT),
    ("• Cada creator entra no MÊS EM QUE FOI CONVIDADA (coorte), não no 1º convite all-time. Uma creator reconvidada aparece em cada mês.", 10, False, "374151"),
    ("• GMV/comissão/pedidos = TODO o GMV de afiliada da creator (todos os produtos, fonte Affiliate Orders), não só o produto amostrado.", 10, False, "374151"),
    (f"• Janela por coorte = [1º convite do mês, próximo convite dela) — capada no convite seguinte, então cada venda cai na coorte de seeding mais recente. Última coorte vai até hoje ({HOJE}).", 10, False, "374151"),
    ("• Soma ADITIVA: os meses somam ao TOTAL sem dupla contagem (= cada creator do 1º convite dela até hoje). Mesma lógica do relatório mensal detalhado.", 10, False, "374151"),
    (f"• Vendas coletadas (raw_creator_product, Affiliate Orders): {CP_INI}→{CP_FIM}. Coortes anteriores a {CP_INI} só capturam vendas a partir dessa data.", 10, False, "B45309"),
    ("", 10, False, "000000"),
    ("Custo e retorno (premissas editáveis no Resumo)", 11, True, ACCENT),
    ("• Custo seeding = (peças enviadas × custo unitário) + frete por envio.", 10, False, "374151"),
    ("• Margem contrib. = GMV × margem. Lucro líquido = margem − comissão paga − custo seeding.", 10, False, "374151"),
    ("• ROAS = GMV ÷ custo seeding.   ROI = lucro líquido ÷ custo seeding.", 10, False, "374151"),
    ("", 10, False, "000000"),
    ("Ressalvas", 11, True, "B45309"),
    ("• Custo de peça/frete são premissas — troque pelos valores reais de COGS/logística da Rhode.", 10, False, "374151"),
    ("• GMV é bruto (não desconta devoluções).", 10, False, "374151"),
    ("• Cada pedido de amostra = 1 unidade de 1 SKU (a API não expõe múltiplas unidades por pedido).", 10, False, "374151"),
]
for i, (txt, sz, bold, color) in enumerate(meto, 1):
    c = ws3.cell(row=i, column=2, value=txt)
    c.font = Font(size=sz, bold=bold, color=color)
    c.alignment = Alignment(wrap_text=True, vertical="top")

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print("OK:", OUT)
print(f"creators={n} | peças enviadas={int(g.enviadas.sum())} | canceladas={int(g.canceladas.sum())} "
      f"| GMV=R$ {g.gmv.sum():,.2f} | com=R$ {g.com.sum():,.2f}")
