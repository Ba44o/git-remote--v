"""
Rhode Jeans — Coletor de Dados TikTok Shop via API
---------------------------------------------------
Busca dados diários de loja e lives, salva em Excel
para o gerar_dashboard.py processar.

Uso: python3 coletar_dados.py [--dias 90]
"""

import os, json, hmac, hashlib, time, sys, glob
import requests
import pandas as pd
from datetime import datetime, timedelta, date
from dotenv import load_dotenv

load_dotenv()
try:
    import token_store                       # access_token da tabela api_tokens (fallback .env)
except Exception:
    token_store = None

APP_KEY      = os.getenv("TIKTOK_APP_KEY")
APP_SECRET   = os.getenv("TIKTOK_APP_SECRET")
# Token da tabela (fonte de verdade na nuvem) → fallback .env (local).
ACCESS_TOKEN = (token_store.get_access_token() if token_store else None) or os.getenv("TIKTOK_ACCESS_TOKEN")
SHOP_CIPHER  = os.getenv("TIKTOK_SHOP_CIPHER")
BASE_URL     = "https://api.tiktok-shops.com"
PASTA        = "dados/marketplace/tiktokshop"

# ── ASSINATURA ──────────────────────────────────────────────────
def assinar(path, params, body_str=""):
    exclude = {"sign", "access_token", "x-tts-access-token"}
    pares = "".join(f"{k}{v}" for k, v in sorted(params.items())
                    if k not in exclude and not isinstance(v, (list, dict)))
    msg = APP_SECRET + path + pares + body_str + APP_SECRET
    return hmac.new(APP_SECRET.encode(), msg.encode(), hashlib.sha256).hexdigest()

def chamar(method, path, params=None, body=None):
    p = params or {}
    p.update({
        "app_key":      APP_KEY,
        "timestamp":    int(time.time()),
        "sign_method":  "HmacSHA256",
        "shop_cipher":  SHOP_CIPHER,
    })
    body_str = json.dumps(body, separators=(",", ":")) if body and method != "GET" else ""
    p["sign"] = assinar(path, p, body_str)
    url = BASE_URL + path
    headers = {
        "Content-Type":       "application/json",
        "x-tts-access-token": ACCESS_TOKEN,
    }
    for tentativa in range(3):
        try:
            if method == "GET":
                r = requests.get(url, params=p, headers=headers, timeout=60)
            else:
                r = requests.post(url, params=p, data=body_str, headers=headers, timeout=60)
            try:
                return r.json()
            except Exception:
                return {"code": -1, "message": f"HTTP {r.status_code}: resposta não-JSON"}
        except requests.exceptions.Timeout:
            if tentativa < 2:
                print(f"  ⏳ Timeout, tentando novamente ({tentativa+2}/3)...")
                time.sleep(5)
            else:
                return {"code": -1, "message": "Timeout após 3 tentativas"}
        except Exception as e:
            return {"code": -1, "message": str(e)}

# ── DATAS ────────────────────────────────────────────────────────
def ts(d): return int(datetime.strptime(str(d), "%Y-%m-%d").timestamp())
def hoje_str(): return date.today().strftime("%Y-%m-%d")
def dias_atras(n): return (date.today() - timedelta(days=n)).strftime("%Y-%m-%d")

# ── PEDIDOS (Overview diário) ────────────────────────────────────
def buscar_pedidos(inicio, fim):
    """Busca pedidos e agrega por dia → formato Overview"""
    print(f"  Buscando pedidos de {inicio} a {fim}...")
    pedidos = []
    cursor = ""
    pagina = 0
    while pagina < 200:
        pagina += 1
        # page_size vai como query param; filtros vão no body
        q_extra = {"page_size": 100, "sort_field": "create_time", "sort_order": "ASC"}
        if cursor:
            q_extra["page_token"] = cursor
        body = {
            "create_time_ge": ts(inicio),
            "create_time_lt":  ts(fim) + 86399,
        }

        data = chamar("POST", "/order/202309/orders/search", params=q_extra, body=body)
        if data.get("code") != 0:
            print(f"  ⚠  Pedidos: {data.get('message')} (code {data.get('code')})")
            break

        orders = data.get("data", {}).get("orders", [])
        if not orders:
            break
        pedidos.extend(orders)
        print(f"    página {pagina}: +{len(orders)} pedidos ({len(pedidos)} total)")
        cursor = data.get("data", {}).get("next_page_token", "")
        if not cursor:
            break
        time.sleep(0.3)

    print(f"  → {len(pedidos)} pedidos encontrados")
    return pedidos

def processar_pedidos(pedidos):
    """Extrai dados enriquecidos dos pedidos: diário, produtos, pagamentos, status"""
    if not pedidos:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    rows, rows_prod = [], []
    for o in pedidos:
        dt       = datetime.fromtimestamp(o.get("create_time", 0)).strftime("%Y-%m-%d")
        status   = o.get("status", "")
        payment  = o.get("payment", {})
        total    = float(payment.get("total_amount", 0))
        reimb    = float(payment.get("seller_discount", 0))
        cpf      = o.get("cpf", "")
        metodo   = o.get("payment_method_name", "Outros") or "Outros"
        items    = o.get("line_items", [])
        cancelado = status == "CANCELLED"

        # ── Localização: estado (L1) ──────────────────────────────────
        estado = ""
        for d in o.get("recipient_address", {}).get("district_info", []):
            if d.get("address_level") == "L1":
                estado = d.get("address_name", "")
                break

        # ── Cancelamento ─────────────────────────────────────────────
        cancel_reason = o.get("cancel_reason", "") if cancelado else ""
        cancel_init   = o.get("cancellation_initiator", "") if cancelado else ""

        rows.append({"Data": dt, "Status": status, "GMV": total, "Reimb": reimb,
                     "Itens": len(items), "CPF": cpf, "Pagamento": metodo,
                     "Cancelado": cancelado,
                     "Estado": estado, "Cancel_Reason": cancel_reason,
                     "Cancel_Init": cancel_init})

        for item in items:
            rows_prod.append({
                "Data":     dt,
                "Produto":  item.get("product_name", "")[:80],
                "SKU":      item.get("sku_name", "")[:40],
                "Preco":    float(item.get("sale_price", 0)),
                "Desconto": float(item.get("seller_discount", 0)),
                "Cancelado": cancelado,
            })

    df = pd.DataFrame(rows)
    df["Data"] = pd.to_datetime(df["Data"])

    # ── Diário ──────────────────────────────────────────────────────
    def agg_day(g):
        v = g[~g["Cancelado"]]   # pedidos confirmados
        c = g[g["Cancelado"]]    # pedidos cancelados
        total_ped = len(v) + len(c)
        return pd.Series({
            "GMV":        g["GMV"].sum(),          # bruto: todos os pedidos
            "GMV_CF":     v["GMV"].sum(),           # líquido: só confirmados
            "Pedidos":    len(v),                   # confirmados
            "Cancelados": len(c),                   # cancelados
            "Itens":      v["Itens"].sum(),
            "Clientes":   v["CPF"].nunique(),
            "Taxa_Cancel": round(len(c) / total_ped * 100, 1) if total_ped > 0 else 0.0,
        })
    diario = df.groupby("Data").apply(agg_day, include_groups=False).reset_index()
    diario["Ticket"] = diario.apply(
        lambda r: r["GMV_CF"] / r["Pedidos"] if r["Pedidos"] > 0 else 0, axis=1)

    # ── Produtos ─────────────────────────────────────────────────────
    if rows_prod:
        dp = pd.DataFrame(rows_prod)
        prods = (dp[~dp["Cancelado"]]
                   .groupby("Produto")
                   .agg(GMV=("Preco","sum"), Pedidos=("Preco","count"), Desconto=("Desconto","sum"))
                   .reset_index()
                   .assign(GMV_Liq=lambda x: x["GMV"]-x["Desconto"])
                   .sort_values("GMV", ascending=False)
                   .head(50)
                   .reset_index(drop=True))
    else:
        prods = pd.DataFrame()

    # ── Pagamentos ──────────────────────────────────────────────────
    v = df[~df["Cancelado"]]
    if not v.empty:
        pag = v.groupby("Pagamento").agg(GMV=("GMV","sum"), Pedidos=("GMV","count")).reset_index()
        tot = pag["GMV"].sum()
        pag["Pct"] = (pag["GMV"]/tot*100).round(1) if tot > 0 else 0
        pag = pag.sort_values("GMV", ascending=False).reset_index(drop=True)
    else:
        pag = pd.DataFrame()

    # ── Status ───────────────────────────────────────────────────────
    st = df.groupby("Status").agg(Count=("GMV","count"), GMV=("GMV","sum")).reset_index()
    tot_c = st["Count"].sum()
    st["Pct"] = (st["Count"]/tot_c*100).round(1) if tot_c > 0 else 0
    st = st.sort_values("Count", ascending=False).reset_index(drop=True)

    # ── Estados ──────────────────────────────────────────────────────
    v = df[~df["Cancelado"]]
    estados_mask = df["Estado"].str.strip() != ""
    df_est = df[estados_mask & ~df["Cancelado"]]
    if not df_est.empty:
        est = (df_est.groupby("Estado")
               .agg(GMV=("GMV","sum"), Pedidos=("GMV","count"), Clientes=("CPF","nunique"))
               .reset_index()
               .sort_values("GMV", ascending=False)
               .reset_index(drop=True))
        tot_e = est["GMV"].sum()
        est["Pct"] = (est["GMV"] / tot_e * 100).round(1) if tot_e > 0 else 0
    else:
        est = pd.DataFrame()

    # ── Cancelamentos ────────────────────────────────────────────────
    df_can = df[df["Cancelado"]]
    if not df_can.empty:
        can = (df_can.groupby(["Cancel_Reason","Cancel_Init"])
               .agg(Count=("GMV","count"), GMV=("GMV","sum"))
               .reset_index()
               .sort_values("Count", ascending=False)
               .reset_index(drop=True))
        tot_can = can["Count"].sum()
        can["Pct"] = (can["Count"] / tot_can * 100).round(1) if tot_can > 0 else 0
    else:
        can = pd.DataFrame()

    return diario, prods, pag, st, est, can

# ── ANALYTICS DE LOJA ────────────────────────────────────────────
def buscar_analytics(inicio, fim, granularidade="1D"):
    """Shop Performance consolidado da loja: GMV, compradores, ticket médio,
    cancelamentos e breakdown por fonte (LIVE / VIDEO / PRODUCT_CARD).
    Endpoint verificado vivo em 2026-06 (escopo data.shop_analytics.public.read).
    Ver ROADMAP decisão #12. Retorna lista de intervalos: 1 por dia se
    granularidade='1D', 1 agregado se 'ALL'. NÃO está wirado no main() ainda —
    ativação da Performance Diária é o item 4.1 do ROADMAP (adiado por decisão)."""
    print(f"  Buscando analytics de {inicio} a {fim} (gran={granularidade})...")
    # end_date_lt é EXCLUSIVO → soma 1 dia pra cobrir 'fim' inclusive
    fim_excl = (datetime.strptime(fim, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    params = {
        "start_date_ge": inicio,
        "end_date_lt":   fim_excl,
        "granularity":   granularidade,   # "1D" (por dia) ou "ALL" (agregado)
    }
    MAX_TENT = 4
    for tentativa in range(MAX_TENT):     # 36009007 = timeout de servidor, transiente
        resp = chamar("GET", "/analytics/202405/shop/performance", params=params)
        if resp.get("code") != 36009007:
            break
        if tentativa < MAX_TENT - 1:
            print(f"  ⏳ Analytics timeout do servidor, tentando de novo ({tentativa+2}/{MAX_TENT})...")
            time.sleep(2)
    if resp.get("code") == 0:
        data = resp.get("data", {})
        intervalos = data.get("performance", {}).get("intervals", [])
        # Lag de ~2 dias: a API só consolida dias até latest_available_date.
        # Descarta dias recentes ainda não finalizados pra não ingerir parcial.
        latest = data.get("latest_available_date")
        if latest:
            antes = len(intervalos)
            intervalos = [i for i in intervalos if i.get("start_date", "") <= latest]
            if len(intervalos) < antes:
                print(f"  → {antes - len(intervalos)} dia(s) não finalizado(s) ignorado(s) (dados até {latest})")
        print(f"  → Analytics OK, {len(intervalos)} dia(s) finalizado(s)")
        return intervalos
    print(f"  ⚠  Analytics: {resp.get('message')} (code {resp.get('code')})")
    return []

# ── FINANCE ───────────────────────────────────────────────────────
def buscar_finance(inicio, fim):
    """Statements financeiros (settlements) no período. Pagina via page_token,
    ordenado por statement_time DESC. Cada statement traz receita, taxa TikTok,
    frete, ajuste e settlement_amount (o que a loja realmente recebe).
    Retorna lista de statements dentro de [inicio, fim]."""
    print(f"  Buscando finance de {inicio} a {fim}...")
    ge, lt = ts(inicio), ts(fim) + 86399
    statements, cursor, pagina = [], "", 0
    while pagina < 200:
        pagina += 1
        params = {"page_size": 50, "sort_field": "statement_time", "sort_order": "DESC"}
        if cursor:
            params["page_token"] = cursor
        resp = chamar("GET", "/finance/202309/statements", params=params)
        if resp.get("code") != 0:
            print(f"  ⚠  Finance: {resp.get('message')} (code {resp.get('code')})")
            break
        data  = resp.get("data", {})
        batch = data.get("statements", [])
        for s in batch:
            t = int(s.get("statement_time", 0) or 0)
            if ge <= t <= lt:
                statements.append(s)
        cursor = data.get("next_page_token", "")
        # DESC: se o batch já passou do início da janela, ou acabou a paginação, para
        if not cursor or (batch and int(batch[-1].get("statement_time", 0) or 0) < ge):
            break
    print(f"  → {len(statements)} statement(s) no período")
    return statements

# ── DEVOLUÇÕES / REEMBOLSOS ───────────────────────────────────────
def buscar_devolucoes(inicio, fim):
    """Devoluções/reembolsos no período. Server filtra por create_time; pagina
    via page_token. Cada return traz refund_total, motivo, status, tipo e os
    itens devolvidos (product_name, seller_sku, refund por item)."""
    print(f"  Buscando devoluções de {inicio} a {fim}...")
    body = {"create_time_ge": ts(inicio), "create_time_lt": ts(fim) + 86399}
    out, cursor, pagina = [], "", 0
    while pagina < 300:
        pagina += 1
        params = {"page_size": 50, "sort_field": "create_time", "sort_order": "DESC"}
        if cursor:
            params["page_token"] = cursor
        resp = chamar("POST", "/return_refund/202309/returns/search", params=params, body=body)
        if resp.get("code") != 0:
            print(f"  ⚠  Devoluções: {resp.get('message')} (code {resp.get('code')})")
            break
        data  = resp.get("data", {})
        batch = data.get("return_orders", [])
        out.extend(batch)
        cursor = data.get("next_page_token", "")
        if not cursor or not batch:
            break
    print(f"  → {len(out)} devolução(ões) no período")
    return out

# ── PEDIDOS DE AFILIADO (creator attribution) ─────────────────────
def buscar_affiliate_orders(inicio, fim):
    """Pedidos atribuídos a afiliadas no período. Cada SKU traz creator_username,
    comissão estimada, base de comissão (≈ GMV) e content_type (SHOP/VIDEO/LIVE).
    Server filtra por create_time; pagina via page_token (page_size 100)."""
    print(f"  Buscando pedidos de afiliado de {inicio} a {fim}...")
    body = {"create_time_ge": ts(inicio), "create_time_lt": ts(fim) + 86399}
    out, cursor, pagina = [], "", 0
    while pagina < 2000:
        pagina += 1
        params = {"page_size": 100}
        if cursor:
            params["page_token"] = cursor
        # retry em timeout de servidor (36009007) — antes quebrava e TRUNCAVA a coleta
        resp = None
        for tent in range(5):
            resp = chamar("POST", "/affiliate_seller/202410/orders/search", params=params, body=body)
            if resp.get("code") != 36009007:
                break
            print(f"    ⏳ timeout pág {pagina}, retry {tent+2}/5...")
            time.sleep(2)
        if resp.get("code") != 0:
            # NÃO truncar em silêncio: aborta pra não gerar dado parcial inconsistente
            raise RuntimeError(f"affiliate_orders falhou na página {pagina} (após {len(out)} pedidos): "
                               f"{resp.get('message')} (code {resp.get('code')})")
        data  = resp.get("data", {})
        batch = data.get("orders", [])
        out.extend(batch)
        cursor = data.get("next_page_token", "")
        if pagina % 25 == 0:
            print(f"    …{len(out)} pedidos")
        if not cursor or not batch:
            break
    print(f"  → {len(out)} pedidos de afiliado")
    return out

# ── TARGET COLLABORATIONS (convites de seeding) ───────────────────
def buscar_target_collaborations(statuses=("VALID", "ONGOING", "EXPIRING", "COMPLETED", "CANCELING"), max_detalhe=400):
    """Convites diretos a creators (target collaborations). Search por status +
    detalhe por convite (creators + product_ids + comissão). Cada convite ≈ 1 creator
    com 1+ produtos e flag de amostra grátis. Retorna lista enriquecida."""
    SEARCH = "/affiliate_seller/202409/target_collaborations/search"
    colabs = {}
    for st in statuses:
        cursor, pagina = "", 0
        while pagina < 200:
            pagina += 1
            body = {"collaboration_status": st, "page_size": 50}
            if cursor:
                body["page_token"] = cursor
            resp = chamar("POST", SEARCH, params={"page_size": 50}, body=body)
            if resp.get("code") != 0:
                if pagina == 1:
                    print(f"  ⚠  TargetCollab {st}: {resp.get('message','')[:50]}")
                break
            data = resp.get("data", {}) or {}
            for c in data.get("target_collaborations", []):
                c["_status"] = st
                colabs.setdefault(c["id"], c)
            cursor = data.get("next_page_token", "")
            if not cursor:
                break
        print(f"  → {st}: acumulado {len(colabs)} convites")

    # detalhe por convite (creators + products)
    enriched, ids = [], list(colabs.keys())[:max_detalhe]
    if len(colabs) > max_detalhe:
        print(f"  ⚠  {len(colabs)} convites; detalhando os {max_detalhe} primeiros")
    for i, cid in enumerate(ids, 1):
        d = chamar("GET", f"/affiliate_seller/202409/target_collaborations/{cid}")
        tc = (d.get("data") or {}).get("target_collaboration") or {}
        base = colabs[cid]
        enriched.append({
            "id": cid, "status": base.get("_status", ""), "name": base.get("name", ""),
            "start_time": base.get("start_time"), "end_time": base.get("end_time"),
            "has_free_sample": (tc.get("free_sample_rule") or base.get("free_sample_rule") or {}).get("has_free_sample", False),
            "creators": tc.get("creators", []),
            "products": tc.get("products", []),
        })
        if i % 50 == 0:
            print(f"    …detalhe {i}/{len(ids)}")
    print(f"  → {len(enriched)} convites detalhados")
    return enriched

# ── SAMPLE APPLICATIONS (free sample REAL — peças por creator) ─────
def buscar_sample_applications(enrich_dates=True):
    """Pedidos de free sample do TikTok Shop (fonte de verdade das PEÇAS de seeding).
    POST /affiliate_seller/202409/sample_applications/search (body vazio = todos).
    Grão: 1 linha por creator × SKU pedido. Peça enviada = status COMPLETED/SHIPPED/
    CONTENT_PENDING. Enriquece com a DATA do convite = create_time do pedido de amostra
    (GET /order/202309/orders), em lotes com fallback individual."""
    PATH = "/affiliate_seller/202409/sample_applications/search"
    apps, cursor, pagina = [], "", 0
    while pagina < 200:
        pagina += 1
        params = {"page_size": 50}
        if cursor:
            params["page_token"] = cursor
        resp = chamar("POST", PATH, params=params, body={})
        if resp.get("code") != 0:
            if pagina == 1:
                print(f"  ⚠  SampleApplications: {resp.get('message','')[:60]}")
            break
        data = resp.get("data", {}) or {}
        for a in data.get("sample_applications", []):
            cr = a.get("creator") or {}
            pr = a.get("product") or {}
            apps.append({
                "id":              str(a.get("id", "")),
                "creator":         cr.get("username", ""),
                "nickname":        cr.get("nickname", ""),
                "follower":        cr.get("follower_count", 0),
                "creator_gmv":     (cr.get("gmv") or {}).get("amount", ""),
                "product_id":      str(pr.get("id", "")),
                "sku_id":          str(pr.get("sku_id", "")),
                "sku_name":        pr.get("sku_name", ""),
                "title":           pr.get("title", ""),
                "status":          a.get("status", ""),
                "order_id":        str(a.get("order_id", "")),
                "commission_rate": a.get("commission_rate", ""),
                "convite_data":    "",
            })
        cursor = data.get("next_page_token", "")
        if not cursor:
            break
    print(f"  → {len(apps)} pedidos de amostra")

    if enrich_dates and apps:
        oids = sorted({a["order_id"] for a in apps if a["order_id"]})
        m = {}
        def _fetch(batch):
            r = chamar("GET", "/order/202309/orders", params={"ids": ",".join(batch)})
            if r.get("code") != 0:
                return False
            for o in (r.get("data") or {}).get("orders", []):
                ct = o.get("create_time")
                if ct:
                    m[str(o.get("id"))] = datetime.fromtimestamp(int(ct)).strftime("%Y-%m-%d")
            return True
        B = 20
        for i in range(0, len(oids), B):
            batch = oids[i:i + B]
            if not _fetch(batch):          # lote falhou → tenta id a id (limite do parâmetro)
                for oid in batch:
                    _fetch([oid])
        for a in apps:
            a["convite_data"] = m.get(a["order_id"], "")
        print(f"  → datas de convite recuperadas: {sum(1 for a in apps if a['convite_data'])}/{len(apps)}")
    return apps

# ── CATÁLOGO (product_id → título) ────────────────────────────────
def buscar_produtos():
    """Catálogo: product_id → título, status, seller_sku principal. Pagina products/search."""
    out, cursor, pagina = {}, "", 0
    while pagina < 50:
        pagina += 1
        params = {"page_size": 100}
        if cursor:
            params["page_token"] = cursor
        resp = chamar("POST", "/product/202309/products/search", params=params, body={"status": "ALL"})
        if resp.get("code") != 0:
            print(f"  ⚠  Produtos: {resp.get('message','')[:50]}")
            break
        data = resp.get("data", {}) or {}
        for p in data.get("products", []):
            skus = p.get("skus", []) or []
            out[p["id"]] = {
                "title": p.get("title", ""),
                "status": p.get("status", ""),
                "seller_sku": (skus[0].get("seller_sku", "") if skus else ""),
            }
        cursor = data.get("next_page_token", "")
        if not cursor:
            break
    print(f"  → {len(out)} produtos no catálogo")
    return out

# ── LIVES ─────────────────────────────────────────────────────────
def buscar_lives(inicio, fim):
    """Busca lista de lives e métricas de cada uma"""
    print(f"  Buscando lives de {inicio} a {fim}...")
    lives = []

    # Tenta endpoint de live list
    endpoints = [
        "/live/202309/lives/search",
        "/creator/202309/live/list",
        "/live/202309/session/list",
    ]
    data_lives = None
    for ep in endpoints:
        body = {
            "start_time": ts(inicio),
            "end_time":   ts(fim) + 86399,
            "page_size":  50,
        }
        resp = chamar("POST", ep, body=body)
        if resp.get("code") == 0:
            data_lives = resp.get("data", {}).get("live_list") or resp.get("data", {}).get("lives") or []
            print(f"  → Endpoint {ep} OK, {len(data_lives)} lives")
            break
        else:
            print(f"  ⚠  {ep}: {resp.get('message')} ({resp.get('code')})")

    if not data_lives:
        print("  → Nenhum endpoint de live respondeu. Mantendo dados existentes.")
        return []

    # Para cada live, busca métricas detalhadas
    for live in data_lives:
        live_id = live.get("live_id") or live.get("id")
        if not live_id:
            continue
        metrics = buscar_metricas_live(live_id)
        if metrics:
            live.update(metrics)
        time.sleep(0.3)
        lives.append(live)

    print(f"  → {len(lives)} lives com métricas")
    return lives

def buscar_metricas_live(live_id):
    """Busca métricas detalhadas de uma live"""
    endpoints = [
        f"/live/202309/lives/{live_id}/analytics",
        f"/creator/202309/live/{live_id}/metrics",
    ]
    for ep in endpoints:
        resp = chamar("GET", ep)
        if resp.get("code") == 0:
            return resp.get("data", {})
    return {}

def montar_df_lives(lives):
    """Converte lista de lives em DataFrame no formato Creator-Live-Performance"""
    if not lives:
        return pd.DataFrame()
    rows = []
    for l in lives:
        rows.append({
            "Livestream":          l.get("title") or l.get("live_title", "Live"),
            "Start time":          datetime.fromtimestamp(l.get("start_time", 0)),
            "Duration":            l.get("duration", 0),
            "Gross revenue":       float(l.get("gmv", 0)) / 100,
            "Direct GMV":          float(l.get("direct_gmv", 0)) / 100,
            "Items sold":          l.get("items_sold", 0),
            "Customers":           l.get("customers", 0),
            "Orders paid for":     l.get("orders", 0),
            "Views":               l.get("views", 0),
            "Viewers":             l.get("viewers", 0),
            "Peak viewers":        l.get("peak_viewers", 0),
            "New followers":       l.get("new_followers", 0),
            "Avg view duration":   l.get("avg_watch_duration", 0),
            "Likes":               l.get("likes", 0),
            "Comments":            l.get("comments", 0),
            "Shares":              l.get("shares", 0),
            "Product impressions": l.get("product_impressions", 0),
            "Product clicks":      l.get("product_clicks", 0),
            "CTR":                 l.get("ctr", 0),
            "CTOR (SKU orders)":   l.get("ctor", 0),
        })
    return pd.DataFrame(rows)

# ── SALVAR EXCEL ─────────────────────────────────────────────────
def salvar_overview(diario, prods, pag, st, est, can):
    if diario.empty:
        print("  → Sem dados de loja para salvar")
        return
    hoje    = datetime.now().strftime("%Y%m%d%H%M%S")
    caminho = f"{PASTA}/Overview_My Business Performance_{hoje}.xlsx"

    from openpyxl import Workbook
    wb = Workbook()

    # ── Sheet1: formato compatível com export manual ─────────────────
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Business Performance Overview"])
    ws1.append([])
    ws1.append([])
    ws1.append([])
    ws1.append(["Data","GMV Bruto","GMV Líquido","Itens Vendidos",
                "Clientes Únicos","Views","Visitas","Pedido SKU","Pedidos","Taxa Cancelamento"])
    for _, r in diario.iterrows():
        ws1.append([
            r["Data"].strftime("%d/%m/%Y"),
            round(r["GMV"],2), round(r["GMV_CF"],2),
            int(r["Itens"]), int(r["Clientes"]), 0, 0, 0, int(r["Pedidos"]),
            f"{round(r['Taxa_Cancel'],1)}%",
        ])

    # ── Diario: dados enriquecidos por dia ───────────────────────────
    ws2 = wb.create_sheet("Diario")
    ws2.append(["Data","GMV","GMV_CF","Pedidos","Cancelados",
                "Itens","Clientes","Ticket","Taxa_Cancel"])
    for _, r in diario.iterrows():
        ws2.append([r["Data"].strftime("%Y-%m-%d"), round(r["GMV"],2),
                    round(r["GMV_CF"],2), int(r["Pedidos"]), int(r["Cancelados"]),
                    int(r["Itens"]), int(r["Clientes"]), round(r["Ticket"],2),
                    round(r["Taxa_Cancel"],1)])

    # ── Produtos: ranking de produtos ────────────────────────────────
    if not prods.empty:
        ws3 = wb.create_sheet("Produtos")
        ws3.append(["Produto","GMV","Pedidos","Desconto","GMV_Liq"])
        for _, r in prods.iterrows():
            ws3.append([r["Produto"], round(r["GMV"],2), int(r["Pedidos"]),
                        round(r["Desconto"],2), round(r["GMV_Liq"],2)])

    # ── Pagamentos: métodos de pagamento ─────────────────────────────
    if not pag.empty:
        ws4 = wb.create_sheet("Pagamentos")
        ws4.append(["Metodo","GMV","Pedidos","Pct"])
        for _, r in pag.iterrows():
            ws4.append([r["Pagamento"], round(r["GMV"],2), int(r["Pedidos"]), round(r["Pct"],1)])

    # ── Status: distribuição de status ───────────────────────────────
    if not st.empty:
        ws5 = wb.create_sheet("Status")
        ws5.append(["Status","Count","GMV","Pct"])
        for _, r in st.iterrows():
            ws5.append([r["Status"], int(r["Count"]), round(r["GMV"],2), round(r["Pct"],1)])

    # ── Estados: pedidos por estado ───────────────────────────────────
    if not est.empty:
        ws6 = wb.create_sheet("Estados")
        ws6.append(["Estado","GMV","Pedidos","Clientes","Pct"])
        for _, r in est.iterrows():
            ws6.append([r["Estado"], round(r["GMV"],2), int(r["Pedidos"]),
                        int(r["Clientes"]), round(r["Pct"],1)])

    # ── Cancelamentos: motivos e iniciadores ─────────────────────────
    if not can.empty:
        ws7 = wb.create_sheet("Cancelamentos")
        ws7.append(["Motivo","Iniciador","Qtd","GMV","Pct"])
        for _, r in can.iterrows():
            ws7.append([r["Cancel_Reason"], r["Cancel_Init"],
                        int(r["Count"]), round(r["GMV"],2), round(r["Pct"],1)])

    wb.save(caminho)
    print(f"  ✅ Overview salvo: {caminho}")

def salvar_lives(df):
    if df.empty:
        print("  → Sem dados de lives para salvar")
        return
    hoje = datetime.now().strftime("%Y%m%d%H%M%S")
    caminho = f"{PASTA}/Creator-Live-Performance_{hoje}.xlsx"

    writer = pd.ExcelWriter(caminho, engine="openpyxl")
    wb = writer.book
    ws = wb.create_sheet("Sheet1")
    ws.append(["Live Performance Report"])
    ws.append([])
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(caminho)
    print(f"  ✅ Lives salvo: {caminho}")

# ── MAIN ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    dias = 7  # padrão: 7 dias para update diário (use --dias 90 para coleta histórica)
    for i, arg in enumerate(sys.argv[1:]):
        if arg == "--dias" and i + 1 < len(sys.argv[1:]):
            dias = int(sys.argv[i + 2])

    inicio = dias_atras(dias)
    fim    = hoje_str()

    print(f"\n🔄 Rhode Jeans — Coleta de Dados TikTok Shop")
    print(f"   Período: {inicio} → {fim}")
    print("=" * 50)

    os.makedirs(PASTA, exist_ok=True)

    # Overview
    print("\n📦 Pedidos / Overview:")
    pedidos              = buscar_pedidos(inicio, fim)
    diario, prods, pag, st, est, can = processar_pedidos(pedidos)
    salvar_overview(diario, prods, pag, st, est, can)

    # Lives
    print("\n🎬 Lives:")
    lives    = buscar_lives(inicio, fim)
    df_lives = montar_df_lives(lives)
    salvar_lives(df_lives)

    print("\n" + "=" * 50)
    print("✅ Coleta concluída!")
    print("\nPróximo passo: python3 gerar_dashboard.py")
