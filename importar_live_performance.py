#!/usr/bin/env python3
"""
Rhode — Importa o export de LIVE do Seller Center (Livestream / Creator-Live-Performance).
Fonte da verdade do GMV de live (coluna "Attributed GMV"). Popula a tabela live_attr.

Lê todos os `dados/lives/exports/Creator-Live-Performance_*.xlsx` (aba performance_detail),
dedupe por room_id, upsert no Supabase. Bate ao centavo com o Seller Center.

Uso:  python3 importar_live_performance.py
      python3 importar_live_performance.py --arquivo "caminho/para/export.xlsx"
"""
import os, sys, re, glob, json, argparse, hashlib, urllib.request, urllib.error
from datetime import datetime
from dotenv import load_dotenv

BASE = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE, ".env"))
SB = os.environ["SUPABASE_URL"]; SK = os.environ["SUPABASE_SERVICE_KEY"]
SBH = {"apikey": SK, "Authorization": f"Bearer {SK}", "Content-Type": "application/json"}

import openpyxl


def money(s):
    """Aceita número, US (R$ 5,818.38) e BR (R$ 3.545,94). Detecta o decimal
       pelo último separador quando há vírgula E ponto."""
    if s is None: return 0.0
    if isinstance(s, (int, float)): return float(s)
    t = re.sub(r"[^\d,.\-]", "", str(s).replace("R$", "").replace("\xa0", "").strip())
    if not t: return 0.0
    if "," in t and "." in t:
        if t.rfind(",") > t.rfind("."): t = t.replace(".", "").replace(",", ".")  # BR: 3.545,94
        else: t = t.replace(",", "")                                              # US: 5,818.38
    elif "," in t:
        t = t.replace(",", ".") if re.search(r",\d{1,2}$", t) else t.replace(",", "")
    try: return float(t or 0)
    except: return 0.0


def inte(s):
    return int(round(money(s)))


def parse_dt(s):
    if s is None: return None
    if isinstance(s, datetime): return s
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try: return datetime.strptime(str(s).strip(), fmt)
        except: pass
    return None


def ler_arquivo(path):
    """Lê os 2 formatos do export de Livestream do Seller Center:
       A) novo (aba performance_detail): tem coluna 'Room ID' + 'Room Title'.
       B) antigo (Sheet1): coluna 'Livestream' (nome) sem Room ID.
       Ambos têm 'Attributed GMV'. Chave de dedupe = período|nome|minuto de início."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["performance_detail"] if "performance_detail" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # cabeçalho = a linha que tem "Attributed GMV"
    hdr_i = next((i for i, r in enumerate(rows) if r and any(str(c).strip() == "Attributed GMV" for c in r if c)), None)
    if hdr_i is None:
        print(f"  ⚠ {os.path.basename(path)}: sem coluna 'Attributed GMV' — pulando"); return []
    hdr = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    idx = {h: i for i, h in enumerate(hdr)}
    def gi(*names):
        for n in names:
            if n in idx: return idx[n]
        return None
    ci = {
        "room": gi("Room ID"), "titulo": gi("Room Title", "Livestream"),
        "inicio": gi("Start Time", "Start time"), "fim": gi("End Time"), "dur": gi("Duration"),
        "gmv": gi("Attributed GMV"), "itens": gi("Attributed items sold"),
        "ped": gi("Attributed orders"), "sku": gi("Attributed SKU orders"),
        "cust": gi("Customers"), "aov": gi("AOV"),
        "views": gi("Views"), "impr": gi("Impressions", "LIVE impressions"),
    }
    def cell(r, key):
        i = ci[key]; return r[i] if (i is not None and i < len(r)) else None
    out = []
    for r in rows[hdr_i + 1:]:
        if not r: continue
        name = str(cell(r, "titulo")).strip() if cell(r, "titulo") else ""
        if name.lower() in ("livestream", "room title", "total", ""):  # cabeçalho repetido / total / vazio
            if not name: continue
            continue
        dt = parse_dt(cell(r, "inicio"))
        if not dt: continue  # sem data não dá pra alocar período
        periodo = dt.strftime("%Y-%m")
        rid = str(cell(r, "room")).strip() if cell(r, "room") else ""
        if not (rid.isdigit() and len(rid) >= 10):
            continue  # sem Room ID (export antigo) → deixa pra API; casar por room_id evita duplicar
        dkey = f"{periodo}|{name.lower()}|{dt.strftime('%Y-%m-%d %H:%M')}"
        out.append({
            "_dkey": dkey, "id": rid, "room_id": rid,
            "titulo": name[:200],
            "inicio": dt.isoformat(),
            "fim": (parse_dt(cell(r, "fim")).isoformat() if parse_dt(cell(r, "fim")) else None),
            "duracao": str(cell(r, "dur") or "")[:20],
            "data": dt.strftime("%Y-%m-%d"), "periodo": periodo,
            "origem": "propria",  # export do Livestream = lives da própria loja Rhode
            "gmv": round(money(cell(r, "gmv")), 2),
            "pedidos": inte(cell(r, "ped")), "itens": inte(cell(r, "itens")),
            "sku_orders": inte(cell(r, "sku")), "customers": inte(cell(r, "cust")),
            "aov": round(money(cell(r, "aov")), 2),
            "views": inte(cell(r, "views")), "impressions": inte(cell(r, "impr")),
        })
    return out


def upsert(rows, chunk=500):
    if not rows: return
    for i in range(0, len(rows), chunk):
        body = json.dumps(rows[i:i + chunk]).encode()
        req = urllib.request.Request(f"{SB}/rest/v1/live_attr?on_conflict=id", data=body,
            headers={**SBH, "Prefer": "resolution=merge-duplicates,return=minimal"}, method="POST")
        try:
            urllib.request.urlopen(req, timeout=120)
        except urllib.error.HTTPError as e:
            print(f"  [ERRO upsert] {e.code}: {e.read()[:300]} (rodou live_attr.sql?)"); raise
    print(f"  ✓ live_attr: {len(rows)} salas gravadas")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--arquivo")
    ap.add_argument("--dry", action="store_true", help="só mostra totais por mês, não grava")
    args = ap.parse_args()
    files = [args.arquivo] if args.arquivo else sorted(glob.glob(os.path.join(BASE, "dados/lives/exports/Creator-Live-Performance_*.xlsx")))
    if not files:
        print("Nenhum export Creator-Live-Performance_*.xlsx em dados/lives/exports/"); return
    dedup = {}  # _dkey → registro (dedupe entre formatos/exports; prefere o que tem Room ID)
    for fpath in files:
        recs = ler_arquivo(fpath)
        if recs:
            pers = sorted({x["periodo"] for x in recs})
            print(f"  {os.path.basename(fpath)}: {len(recs)} salas · GMV R$ {sum(x['gmv'] for x in recs):,.2f} · {','.join(pers)}")
        for x in recs:
            prev = dedup.get(x["_dkey"])
            if prev is None:
                dedup[x["_dkey"]] = x
            else:  # colisão: prefere Room ID numérico; senão o export mais recente (files ordenados asc)
                prev_num = prev["room_id"].isdigit(); x_num = x["room_id"].isdigit()
                if x_num and not prev_num: dedup[x["_dkey"]] = x
                elif prev_num and not x_num: pass
                else: dedup[x["_dkey"]] = x
    rows = [{k: v for k, v in x.items() if k != "_dkey"} for x in dedup.values() if x["periodo"]]
    per = {}
    for x in rows:
        p = per.setdefault(x["periodo"], {"n": 0, "gmv": 0.0, "ped": 0, "views": 0})
        p["n"] += 1; p["gmv"] += x["gmv"]; p["ped"] += x["pedidos"]; p["views"] += x["views"]
    print(f"\n═════ live_attr — {len(rows)} salas · {len(per)} meses ═════")
    for p in sorted(per):
        d = per[p]
        print(f"  {p}: {d['n']:>3} salas · GMV R$ {d['gmv']:>13,.2f} · {d['ped']:>5} ped · {d['views']:>8,} views")
    if args.dry:
        print("\n  (dry-run — nada gravado)"); return
    upsert(rows)
    print("  ✓ concluído.")


if __name__ == "__main__":
    main()
