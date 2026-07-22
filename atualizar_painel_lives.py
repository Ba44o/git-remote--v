#!/usr/bin/env python3
"""
Rhode — Atualiza a tabela `lives` (painel INTERNO dash-live) com dados frescos.
  • Funil/GMV: export do Seller Center "Creator-Live-Performance" (formato novo
    performance_detail, com Room ID). Parser próprio (o etl_lives só lê o formato antigo).
  • Ads (ads_cost/ads_gmv/ads_roas): GMV Max via API (tabela live_sessao), casado por room_id.
  • FALLBACK VIA API (v3-api): live que ainda NÃO tem export entra na `lives` a partir do
    `live_attr` (Analytics /shop_lives, coletado no cron) — só o lado de VENDA (GMV/pedidos/
    itens/clientes/AOV/ads). O funil (views/CTR/CTOR/GPM/engajamento) a API NÃO expõe: fica
    vazio até o export chegar. Assim o painel nunca mais fica em branco esperando export.
ROOT de propósito (fora de agente_rhode/ → NÃO dispara etl_sync.yml). Idempotente (upsert live_key).
Uso:  python3 atualizar_painel_lives.py
"""
import os, re, glob, json, urllib.request, urllib.error, urllib.parse
from datetime import datetime, timedelta, timezone
from collections import defaultdict
import openpyxl
from dotenv import load_dotenv

BASE = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE, ".env"))
SB = os.environ["SUPABASE_URL"]; SK = os.environ["SUPABASE_SERVICE_KEY"]
SBH = {"apikey": SK, "Authorization": f"Bearer {SK}", "Content-Type": "application/json"}


def money(s):
    if s is None: return 0.0
    if isinstance(s, (int, float)): return float(s)
    t = re.sub(r"[^\d,.\-]", "", str(s).replace("R$", "").strip())
    if not t: return 0.0
    if "," in t and "." in t:
        t = t.replace(".", "").replace(",", ".") if t.rfind(",") > t.rfind(".") else t.replace(",", "")
    elif "," in t:
        t = t.replace(",", ".") if re.search(r",\d{1,2}$", t) else t.replace(",", "")
    try: return float(t)
    except: return 0.0


def num(s):  # tira % e devolve número (ex "6.62%" → 6.62)
    if s is None: return 0.0
    if isinstance(s, (int, float)): return float(s)
    try: return float(re.sub(r"[^\d.\-]", "", str(s).replace("%", "").replace(",", ".")) or 0)
    except: return 0.0


def inte(s): return int(round(money(s)))


def dur_sec(s):
    if s is None: return 0
    m = re.match(r'(?:(\d+)h)?(?:(\d+)m)?', str(s).strip())
    if m and (m.group(1) or m.group(2)):
        return int(m.group(1) or 0) * 3600 + int(m.group(2) or 0) * 60
    return inte(s)


def parse_dt(s):
    for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try: return datetime.strptime(str(s).strip(), f)
        except: pass
    return None


# ───────── casamento export × API ─────────
# O export grava Start Time em BRT-naive; a API grava UTC-aware. Além das 3h de fuso, as duas
# fontes divergem ~3s no início da mesma live — casar por timestamp exato (ou por minuto, que
# quebra na virada: 15:00:57 → 15:01) DUPLICARIA a live no painel. Casa por título + tolerância.
BRT = timezone(timedelta(hours=-3))
TOL_SEG = 150  # drift observado export×API: 4s a 60s. Lives distintas nunca ficam < 4min.


def norm_t(t): return (t or "").strip().lower()


def dt_lives(s):
    """`lives.started_at`: BRT-naive gravado com tz → ignora o offset."""
    try: return datetime.fromisoformat(str(s)).replace(tzinfo=None)
    except: return None


def dt_attr(s):
    """`live_attr.inicio`/`fim`: UTC-aware → converte pra BRT e solta o tz."""
    try: return datetime.fromisoformat(str(s)).astimezone(BRT).replace(tzinfo=None)
    except: return None


def idx_add(index, title, dt, key):
    index.append((dt, norm_t(title), key))


def idx_find(index, title, dt):
    """live_key da mesma live, ou None. Casa por TEMPO (sinal forte); o título só desempata
    quando há mais de uma candidata na janela. O título NUNCA bloqueia o match — o live_attr
    às vezes vem com titulo vazio (maio/26), e indexar por título duplicava a live no painel."""
    if dt is None: return None
    t = norm_t(title)
    cands = [(abs((dt - d2).total_seconds()), tt, k) for d2, tt, k in index
             if d2 is not None and abs((dt - d2).total_seconds()) <= TOL_SEG]
    if not cands: return None
    iguais = [c for c in cands if t and c[1] == t]
    return min(iguais or cands, key=lambda c: c[0])[2]


def sb_get(path):
    return json.load(urllib.request.urlopen(urllib.request.Request(SB + path, headers=SBH), timeout=120))


def read_export(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["performance_detail"] if "performance_detail" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows) if r and any(str(c).strip() == "Room ID" for c in r if c)), None)
    if hi is None:
        return []  # formato antigo (sem Room ID) — já está na `lives` via etl_lives
    hdr = [str(c).strip() if c is not None else "" for c in rows[hi]]
    out = []
    for r in rows[hi + 1:]:
        if not r or r[0] is None: continue
        rid = str(r[0]).strip()
        if not rid.isdigit() or len(rid) < 10: continue
        out.append(dict(zip(hdr, r)))
    return out


def build_record(d):
    dt = parse_dt(d.get("Start Time"))
    gmv = money(d.get("Attributed GMV"))
    if not dt or gmv <= 0:
        return None, None
    rid = str(d.get("Room ID")).strip(); title = str(d.get("Room Title") or "").strip()
    iso = dt.isoformat(); ds = dur_sec(d.get("Duration"))
    rec = {
        "live_key": f"{iso}|{title}", "title": title[:200], "started_at": iso,
        "date": dt.strftime("%Y-%m-%d"), "month": dt.strftime("%Y-%m"),
        "day_of_week": dt.weekday(), "hour": dt.hour,
        "duration_sec": ds, "duration_min": round(ds / 60, 1),
        "gmv_bruto": round(gmv, 2), "gmv_direct": round(gmv, 2),
        "items_sold": inte(d.get("Attributed items sold")), "attributed_sku_orders": inte(d.get("Attributed SKU orders")),
        "customers": inte(d.get("Customers")), "avg_price": round(money(d.get("AOV")), 2),
        "orders": inte(d.get("Attributed orders")), "views": inte(d.get("Views")),
        "new_followers": inte(d.get("New followers")), "likes": inte(d.get("Likes")),
        "comments": inte(d.get("Comments")), "shares": inte(d.get("Shares")),
        "product_impressions": inte(d.get("Product Impressions") or d.get("Product impressions")),
        "product_clicks": inte(d.get("Product clicks")),
        "ctr": round(num(d.get("CTR")), 4), "ctor": round(num(d.get("CTOR (SKU orders)") or d.get("CTOR")), 4),
        "live_impressions": inte(d.get("Impressions") or d.get("LIVE impressions")),
        "impressions_per_hour": inte(d.get("Impressions Per Hour")),
        "gmv_per_hour": round(money(d.get("GMV per hour")), 2),
        "show_gpm": round(money(d.get("Show GPM")), 2), "watch_gpm": round(money(d.get("Watch GPM")), 2),
        "tap_through_rate": round(num(d.get("Tap through rate")) / 100, 6),
        "live_ctr": round(num(d.get("LIVE CTR")) / 100, 6),
        "follow_rate": round(num(d.get("Follow rate")) / 100, 6), "comment_rate": round(num(d.get("Comment rate")) / 100, 6),
        "share_rate": round(num(d.get("Share rate")) / 100, 6), "like_rate": round(num(d.get("Like rate")) / 100, 6),
        "sku_order_rate": round(num(d.get("SKU order rate")) / 100, 6),
        "ads_cost": None, "ads_gmv": None, "ads_roas": None,  # preenchido depois (GMV Max); keys iguais p/ todos (PostgREST)
        "schema_version": "v2-api",
    }
    return rec, rid


def upsert(rows):
    for i in range(0, len(rows), 500):
        body = json.dumps(rows[i:i + 500]).encode()
        req = urllib.request.Request(SB + "/rest/v1/lives?on_conflict=live_key", data=body,
            headers={**SBH, "Prefer": "resolution=merge-duplicates,return=minimal"}, method="POST")
        try:
            urllib.request.urlopen(req, timeout=120)
        except urllib.error.HTTPError as e:
            print("  [ERRO upsert]", e.code, e.read()[:300]); raise


# Colunas de funil que a API NÃO tem. Vão explicitamente como NULL (nunca 0): o `avg()` do
# dash-live ignora null mas trata 0 como valor válido — gravar 0 afundaria a média de views/
# cliques do período (mesmo bug que já rolou com as "lives v2 sem viewers").
FUNIL_NULL = ("views", "viewers", "peak_viewers", "avg_view_duration_sec", "gmv_per_viewer",
              "product_impressions", "product_clicks", "live_impressions", "impressions_per_hour",
              "gmv_per_hour", "ctr", "ctor", "live_ctr", "tap_through_rate", "sku_order_rate",
              "show_gpm", "watch_gpm", "new_followers", "likes", "comments", "shares",
              "follow_rate", "comment_rate", "share_rate", "like_rate")


def build_api_record(x, ads):
    """Linha ENXUTA a partir do live_attr (API): só o lado de VENDA, que a API realmente sabe.
    O funil vai NULL — o painel então ignora essa live ao calcular CTOR/R$-viewer/views."""
    dt = dt_attr(x.get("inicio")); gmv = float(x.get("gmv") or 0)
    if not dt or gmv <= 0:  # mesma regra do export: live sem GMV não entra no painel
        return None
    title = (x.get("titulo") or "").strip(); iso = dt.isoformat()
    # live em andamento não tem `fim` → duração NULL (não 0, que afundaria a média do painel)
    fim = dt_attr(x.get("fim")); ds = max(0, int((fim - dt).total_seconds())) if fim else None
    a = ads.get(str(x.get("room_id"))) or {}
    cost, agmv = round(a.get("cost", 0.0), 2), round(a.get("gmv", 0.0), 2)
    return {
        "live_key": f"{iso}|{title}", "title": title[:200], "started_at": iso,
        "date": dt.strftime("%Y-%m-%d"), "month": dt.strftime("%Y-%m"),
        "day_of_week": dt.weekday(), "hour": dt.hour,
        "duration_sec": ds, "duration_min": round(ds / 60, 1) if ds else None,
        "gmv_bruto": round(gmv, 2), "gmv_direct": round(gmv, 2),
        "orders": int(float(x.get("pedidos") or 0)),
        "attributed_sku_orders": int(float(x.get("sku_orders") or 0)),
        "items_sold": int(float(x.get("itens") or 0)),
        "customers": int(float(x.get("customers") or 0)),
        "avg_price": round(float(x.get("aov") or 0), 2),
        "ads_cost": cost or None, "ads_gmv": agmv or None,
        "ads_roas": round(agmv / cost, 2) if cost else None,
        "schema_version": "v3-api",
        **{c: None for c in FUNIL_NULL},
    }


def delete_keys(keys):
    """Remove linhas v3-api que o export acabou de cobrir (senão a live ficaria duplicada:
    live_key da API e do export diferem pelos ~3s de drift)."""
    for i in range(0, len(keys), 50):
        q = ",".join('"' + k.replace('"', '""') + '"' for k in keys[i:i + 50])
        url = SB + "/rest/v1/lives?schema_version=eq.v3-api&live_key=in.(" + urllib.parse.quote(q) + ")"
        req = urllib.request.Request(url, headers={**SBH, "Prefer": "return=minimal"}, method="DELETE")
        try:
            urllib.request.urlopen(req, timeout=120)
        except urllib.error.HTTPError as e:
            print("  [ERRO delete]", e.code, e.read()[:200])


def main():
    # 1) exports (formato novo, com Room ID) → registros + mapa room_id→live_key
    recs = {}; room2key = {}
    for f in sorted(glob.glob(os.path.join(BASE, "dados/lives/exports/Creator-Live-Performance_*.xlsx"))):
        for d in read_export(f):
            rec, rid = build_record(d)
            if rec:
                recs[rec["live_key"]] = rec; room2key[rid] = rec["live_key"]
    # 2) ads do GMV Max (live_sessao) agregado por room_id
    ads = defaultdict(lambda: {"cost": 0.0, "gmv": 0.0})
    for x in sb_get("/rest/v1/live_sessao?select=room_id,cost,receita&limit=5000"):
        a = ads[str(x["room_id"])]; a["cost"] += float(x["cost"] or 0); a["gmv"] += float(x["receita"] or 0)
    filled = 0
    for rid, key in room2key.items():
        a = ads.get(rid)
        if a and (a["cost"] > 0 or a["gmv"] > 0):
            recs[key]["ads_cost"] = round(a["cost"], 2); recs[key]["ads_gmv"] = round(a["gmv"], 2)
            recs[key]["ads_roas"] = round(a["gmv"] / a["cost"], 2) if a["cost"] else None
            filled += 1
    rows = list(recs.values())
    if rows:
        upsert(rows)
        print(f"✓ export: {len(rows)} sessões (funil completo) · {filled} com ads (GMV Max)")
    else:
        print("· nenhum export no formato novo — painel fica só com o que vier da API")

    # 3) FALLBACK API: live que ainda não tem export entra pelo live_attr (só lado de venda).
    #    Índice do que JÁ existe (export desta rodada + o que está gravado na `lives`) pra não
    #    duplicar — casando por título + tolerância, porque os timestamps divergem ~3s.
    index = []; cobertas = {}; sv = {}
    for r in rows:
        idx_add(index, r["title"], dt_lives(r["started_at"]), r["live_key"])
        sv[r["live_key"]] = r["schema_version"]
    for x in sb_get("/rest/v1/lives?select=live_key,started_at,title,schema_version&limit=5000"):
        d = dt_lives(x["started_at"])
        idx_add(index, x["title"], d, x["live_key"])
        sv.setdefault(x["live_key"], x.get("schema_version"))
        if x.get("schema_version") == "v3-api":
            cobertas[x["live_key"]] = (d, x["title"])

    # 3a) linha v3-api que o export agora cobre → apaga (o export é a fonte rica)
    if cobertas and rows:
        exp_idx = []
        for r in rows:
            idx_add(exp_idx, r["title"], dt_lives(r["started_at"]), r["live_key"])
        obsoletas = [k for k, (d, t) in cobertas.items()
                     if idx_find(exp_idx, t, d) and idx_find(exp_idx, t, d) != k]
        if obsoletas:
            delete_keys(obsoletas)
            print(f"  ↻ {len(obsoletas)} linha(s) v3-api substituída(s) pelo export (funil completo)")
            # não precisa mexer no `index`: a linha do export cobre o mesmo título/horário,
            # então a live segue casando e não é recriada pela API.

    novos = {}; n_novas = 0
    for x in sb_get("/rest/v1/live_attr?select=room_id,inicio,fim,titulo,gmv,pedidos,sku_orders,itens,customers,aov&limit=5000"):
        dt = dt_attr(x.get("inicio"))
        key = idx_find(index, x.get("titulo"), dt)
        if key and sv.get(key) != "v3-api":
            continue  # export cobre essa live → fonte rica manda, API não encosta
        rec = build_api_record(x, ads)
        if not rec:
            continue
        if key:
            rec["live_key"] = key  # refresca a linha v3-api existente (GMV da API converge)
        elif rec["live_key"] not in novos:
            n_novas += 1
            idx_add(index, rec["title"], dt, rec["live_key"]); sv[rec["live_key"]] = "v3-api"
        novos[rec["live_key"]] = rec
    if novos:
        upsert(list(novos.values()))
        com_ads = sum(1 for r in novos.values() if r.get("ads_cost"))
        print(f"✓ API (live_attr): {len(novos)} live(s) sem export ({n_novas} nova(s)) · "
              f"{com_ads} com ads · funil NULL até o export chegar")
        for r in sorted(novos.values(), key=lambda r: r["date"]):
            print(f"    {r['date']} {r['started_at'][11:16]} · R$ {r['gmv_bruto']:>9,.2f} · {r['title'][:38]}")
    else:
        print("· API: nenhuma live fora do export (tudo coberto)")

    # 4) resumo do que ficou na tabela
    per = defaultdict(lambda: {"n": 0, "gmv": 0.0, "ads": 0, "api": 0})
    for r in sb_get("/rest/v1/lives?select=month,gmv_bruto,ads_cost,schema_version&limit=5000"):
        p = per[r["month"]]; p["n"] += 1; p["gmv"] += float(r["gmv_bruto"] or 0)
        p["ads"] += 1 if r.get("ads_cost") else 0
        p["api"] += 1 if r.get("schema_version") == "v3-api" else 0
    print("\n  mês      lives   GMV            c/ads   via API (sem funil)")
    for m in sorted(per)[-6:]:
        p = per[m]
        print(f"  {m}  {p['n']:>5}   R$ {p['gmv']:>11,.2f}  {p['ads']:>5}   {p['api']:>5}")


if __name__ == "__main__":
    main()
