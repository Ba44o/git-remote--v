"""
Rhode Jeans — Núcleo reutilizável do Relatório MENSAL de Seeding (envio de amostras)
====================================================================================
Gera o template de seeding (6 abas + .md) para QUALQUER mês, comparando sempre com o
mês imediatamente anterior. Toda a narrativa é DATA-DRIVEN (nada hardcoded do mês).

Fontes (materializadas antes de chamar build_report):
  • warehouse/raw_sample_applications.csv          → peças de amostra (sample_applications)
  • warehouse/raw_vendas_seeding_<AAAA-MM>.csv      → vendas p/ atribuição de conversão/ROI

Regras: compara vs. mês anterior; nunca inventa (faltante = "sem dado"); % 1 casa;
R$ 2 casas; ordenação decrescente; aliases de handle colapsados; ROI financeiro
(lucro÷custo) = "sem dado" enquanto não houver COGS/frete real.

Uso programático:
  from seeding_report import build_report
  build_report("2026-06", exec_date="2026-07-06")
"""
from __future__ import annotations
from pathlib import Path
from collections import Counter, defaultdict
from datetime import date, datetime
import calendar
import csv

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

BASE = Path(__file__).resolve().parent
WH = BASE / "warehouse"

SENT = {"COMPLETED", "SHIPPED", "CONTENT_PENDING"}
ALIASES = {
    "NATMARQUESSS": "NATMARQUESVI",
    "TACIANECREATOR": "TACIANETORRESS",
    "TACIRECOMENDA": "TACIANETORRESS",
}
TIER_RULES = [(500000, "Black"), (150000, "Diamond"), (80000, "Gold"),
              (50000, "Silver"), (20000, "Bronze"), (0, "Iniciante")]
TIER_ORDER = ["Black", "Diamond", "Gold", "Silver", "Bronze", "Iniciante"]
MESES_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

# estilos
NAVY = "1F2937"; ACCENT = "4F46E5"; LIGHT = "EEF2FF"; GREY = "F3F4F6"
GREEN = "ECFDF5"; RED = "FEF2F2"; AMBER = "FEF3C7"
GREEN_TXT = "047857"; RED_TXT = "B91C1C"; AMBER_TXT = "B45309"
TIER_FILL = {"Black": "111827", "Diamond": "4F46E5", "Gold": "B45309",
             "Silver": "6B7280", "Bronze": "92400E", "Iniciante": "9CA3AF"}
money = 'R$ #,##0.00'
GMV_AFIL_NOTA = (
    "GMV afiliação (R$)\n"
    "= creator_gmv devolvido pela API do TikTok Shop por creator = GMV BRUTO ACUMULADO "
    "dela como afiliada da loja (histórico / 'cacife' all-time da creator).\n\n"
    "USO: serve SÓ como proxy para estimar o Tier (por isso o tier está marcado 'estim.').\n\n"
    "NÃO CONFUNDIR:\n"
    "• NÃO é o retorno da amostra — isso é a coluna 'GMV gerado (R$)' (vendas pós-convite).\n"
    "• NÃO é o GMV oficial do programa — o oficial usa GMV LÍQUIDO acumulado "
    "(performance_periods); este é bruto/all-time, por isso aproximado."
)


# ── helpers de mês ──────────────────────────────────────────────────────────
def prev_month(mes: str) -> str:
    y, m = int(mes[:4]), int(mes[5:7])
    return f"{y - 1}-12" if m == 1 else f"{y}-{m - 1:02d}"


def mes_label(mes: str) -> str:
    return f"{MESES_PT[int(mes[5:7])]}/{mes[:4]}"


def _tier(g):
    for t, l in TIER_RULES:
        if g >= t:
            return l
    return "Iniciante"


def _ck(u):
    return ALIASES.get(u.strip().upper(), u.strip().upper())


def _d_pct(a, b):
    return (a - b) / b * 100 if b else None


def taxa_env(a):
    return a["enviadas"] / a["pedidas"] * 100 if a["pedidas"] else 0


def taxa_canc(a):
    return a["canceladas"] / a["pedidas"] * 100 if a["pedidas"] else 0


def fmt_val(v, kind):
    if kind == "int":
        return f"{v:d}" if isinstance(v, int) else f"{v:.0f}"
    if kind == "money":
        return f"R$ {v:,.2f}"
    if kind == "num1":
        return f"{v:.1f}"
    if kind in ("pctv", "pp"):
        return f"{v:.1f}%"
    return str(v)


# ═══════════════════════════════════════════════════════════════════════════
def compute_seeding(mes: str, exec_date: str | None = None, base: Path = BASE) -> dict:
    """Computa TODO o modelo do relatório de seeding do mês (sem renderizar).
    Reusado por build_report (xlsx/md) e por sync_seeding_to_sheets (Google Sheets)."""
    exec_date = exec_date or date.today().strftime("%Y-%m-%d")
    PREV = prev_month(mes)
    MESN, ANO = MESES_PT[int(mes[5:7])], mes[:4]
    PREVN, PREVANO = MESES_PT[int(PREV[5:7])], PREV[:4]
    MES_LBL, PREV_LBL = mes_label(mes), mes_label(PREV)
    ATTR_FIM = exec_date

    wh = base / "warehouse"
    rows = list(csv.DictReader(open(wh / "raw_sample_applications.csv")))

    # ── agregação por mês ───────────────────────────────────────────────────
    def build_month(mm):
        sub = [r for r in rows if r["convite_data"][:7] == mm]
        byc = defaultdict(list)
        for r in sub:
            byc[_ck(r["creator"])].append(r)
        creators = []
        for c, rs in byc.items():
            sent = [r for r in rs if r["status"] in SENT]
            canc = [r for r in rs if "CANCELLED" in r["status"]]
            gmv = max(float(r["creator_gmv"] or 0) for r in rs)
            datas = sorted(r["convite_data"] for r in rs if r["convite_data"])
            creators.append({
                "canon": c, "nick": rs[0]["nickname"] or rs[0]["creator"],
                "handle": rs[0]["creator"], "tier": _tier(gmv), "gmv": round(gmv, 2),
                "pedidas": len(rs), "enviadas": len(sent), "canceladas": len(canc),
                "liquidas": len(sent), "skus": len({r["sku_id"] for r in rs}),
                "produtos": len({r["product_id"] for r in rs}),
                "d_ini": datas[0] if datas else "", "d_fim": datas[-1] if datas else "",
            })
        creators.sort(key=lambda x: (-x["enviadas"], -x["gmv"]))
        return sub, creators

    tgt_rows, tgt = build_month(mes)
    prv_rows, prv = build_month(PREV)
    _tdatas = sorted(r["convite_data"] for r in tgt_rows if r["convite_data"])
    TMIN = _tdatas[0] if _tdatas else "—"
    TMAX = _tdatas[-1] if _tdatas else "—"

    def agg(creators, raw):
        sent = [r for r in raw if r["status"] in SENT]
        env_creators = [c for c in creators if c["enviadas"] > 0]
        return {
            "pedidas": sum(c["pedidas"] for c in creators),
            "enviadas": sum(c["enviadas"] for c in creators),
            "canceladas": sum(c["canceladas"] for c in creators),
            "liquidas": sum(c["liquidas"] for c in creators),
            "creators_env": len(env_creators), "creators_tot": len(creators),
            "prod_env": len({r["product_id"] for r in sent}),
            "sku_env": len({r["sku_id"] for r in sent}),
            "gmv_med": (sum(c["gmv"] for c in env_creators) / len(env_creators)) if env_creators else 0,
        }

    AT, AP = agg(tgt, tgt_rows), agg(prv, prv_rows)

    # ── conversão / ROI pós-amostra ─────────────────────────────────────────
    vendas_csv = wh / f"raw_vendas_seeding_{mes}.csv"
    if not vendas_csv.exists() and mes == "2026-06":
        alt = wh / "raw_vendas_seeding_jun.csv"          # back-compat 1ª geração
        if alt.exists():
            vendas_csv = alt
    TEM_VENDAS = vendas_csv.exists()
    vindex = defaultdict(list)
    if TEM_VENDAS:
        for v in csv.DictReader(open(vendas_csv)):
            vindex[_ck(v["creator"])].append(
                (v["data"], float(v["gmv"] or 0), int(v["unidades"] or 0)))

    def roi_for(canon, d_ini):
        if not TEM_VENDAS:
            return {"ok": False}
        gmv, unid = 0.0, 0
        for data, g, u in vindex.get(canon, []):
            if d_ini and d_ini <= data <= ATTR_FIM:
                gmv += g; unid += u
        return {"ok": True, "gmv": round(gmv, 2), "unid": unid, "vendeu": unid > 0}

    ROI = {c["canon"]: roi_for(c["canon"], c["d_ini"]) for c in tgt}
    atingidas = [c for c in tgt if c["enviadas"] > 0]
    conv = {
        "atingidas": len(atingidas),
        "venderam": sum(1 for c in atingidas if ROI[c["canon"]].get("vendeu")),
        "gmv_gerado": round(sum(ROI[c["canon"]].get("gmv", 0) for c in atingidas), 2),
        "unid": sum(ROI[c["canon"]].get("unid", 0) for c in atingidas),
        "enviadas": sum(c["enviadas"] for c in atingidas),
    }
    conv["nao_venderam"] = conv["atingidas"] - conv["venderam"]
    conv["taxa"] = conv["venderam"] / conv["atingidas"] * 100 if conv["atingidas"] else 0
    conv["gmv_por_peca"] = conv["gmv_gerado"] / conv["enviadas"] if conv["enviadas"] else 0

    # ── segmentos ───────────────────────────────────────────────────────────
    def tier_seg(raw):
        d = Counter()
        for r in raw:
            if r["status"] in SENT:
                d[_tier(float(r["creator_gmv"] or 0))] += 1
        return d

    TJ, TM = tier_seg(tgt_rows), tier_seg(prv_rows)
    prod_env = Counter(); prod_title = {}; prod_creators = defaultdict(set)
    for r in tgt_rows:
        if r["status"] in SENT:
            prod_env[r["product_id"]] += 1
            prod_title[r["product_id"]] = r["title"]
            prod_creators[r["product_id"]].add(_ck(r["creator"]))
    prod_rows = sorted(prod_env.items(), key=lambda x: -x[1])

    # conversão por tier
    cbt = {t: {"ating": 0, "vend": 0, "env": 0, "gmv": 0.0} for t in TIER_ORDER}
    for c in atingidas:
        d = cbt[c["tier"]]; roi = ROI.get(c["canon"], {})
        d["ating"] += 1; d["env"] += c["enviadas"]
        if roi.get("ok"):
            d["vend"] += 1 if roi["vendeu"] else 0
            d["gmv"] += roi["gmv"]

    # ── fatos data-driven p/ narrativa ──────────────────────────────────────
    undated = [r for r in rows if not r["convite_data"]]
    n_pending = sum(1 for r in undated if r["status"] in ("PENDING", "AWAITING_SHIPMENT"))
    n_canc_pre = len(undated) - n_pending
    top_prod = (prod_rows[0] if prod_rows else (None, 0))
    top_prod_title = prod_title.get(top_prod[0], "sem dado") if top_prod[0] else "sem dado"
    top_prod_n = top_prod[1]
    tiers_by_env = sorted(TJ.items(), key=lambda x: -x[1])
    top2_tiers = tiers_by_env[:2]
    top2_tiers_n = sum(n for _, n in top2_tiers)
    top2_tiers_lbl = " + ".join(t for t, _ in top2_tiers) if top2_tiers else "—"
    gmv_ranked = sorted(
        [(c, ROI[c["canon"]].get("gmv", 0)) for c in atingidas if ROI[c["canon"]].get("gmv", 0) > 0],
        key=lambda x: -x[1])
    top2_gmv = gmv_ranked[:2]
    top2_gmv_share = (sum(g for _, g in top2_gmv) / conv["gmv_gerado"] * 100
                      if conv["gmv_gerado"] else 0)
    dim = calendar.monthrange(int(mes[:4]), int(mes[5:7]))[1]
    late_cut = dim - 6
    late_nao = sum(1 for c in atingidas
                   if not ROI[c["canon"]].get("vendeu")
                   and c["d_ini"] and int(c["d_ini"][8:10]) >= late_cut)
    top_recv = tgt[0] if tgt and tgt[0]["enviadas"] > 0 else None
    top_recv_conv = ROI[top_recv["canon"]].get("vendeu") if top_recv else None

    # ── KPIs (label, alvo, prev, kind, tipo_delta) ──────────────────────────
    KPIS = [
        ("Peças enviadas (líquidas)", AT["enviadas"], AP["enviadas"], "int", "pct"),
        ("Peças pedidas (solicitações)", AT["pedidas"], AP["pedidas"], "int", "pct"),
        ("Peças canceladas", AT["canceladas"], AP["canceladas"], "int", "pct"),
        ("Taxa de envio (enviadas / pedidas)", taxa_env(AT), taxa_env(AP), "pctv", "pp"),
        ("Taxa de cancelamento", taxa_canc(AT), taxa_canc(AP), "pctv", "pp"),
        ("Creators atingidas (com envio)", AT["creators_env"], AP["creators_env"], "int", "pct"),
        ("Creators solicitantes (total)", AT["creators_tot"], AP["creators_tot"], "int", "pct"),
        ("Produtos distintos amostrados", AT["prod_env"], AP["prod_env"], "int", "pct"),
        ("SKUs distintos amostrados", AT["sku_env"], AP["sku_env"], "int", "pct"),
        ("Peças enviadas por creator (média)", AT["enviadas"] / max(1, AT["creators_env"]),
         AP["enviadas"] / max(1, AP["creators_env"]), "num1", "pct"),
        ("GMV médio de afiliação das atingidas (R$)", AT["gmv_med"], AP["gmv_med"], "money", "pct"),
    ]

    # ── resumo executivo (data-driven, ≤5 linhas) ──────────────────────────
    dpe = _d_pct(AT["enviadas"], AP["enviadas"])
    if TEM_VENDAS and conv["atingidas"]:
        conv_line = (f"Conversão da amostra: {conv['venderam']} das {conv['atingidas']} creators "
                     f"atingidas já venderam peça nossa após o convite ({conv['taxa']:.1f}%), "
                     f"gerando R$ {conv['gmv_gerado']:,.2f} de GMV de afiliada pós-envio "
                     f"({conv['unid']} un.) — R$ {conv['gmv_por_peca']:,.2f} por peça enviada.")
    else:
        conv_line = "Conversão/ROI da amostra: sem dado (fonte de vendas não coletada para o mês)."
    resumo = [
        (f"{MESN} enviou {AT['enviadas']} peças-amostra para {AT['creators_env']} creators "
         f"(vs. {AP['enviadas']} peças / {AP['creators_env']} creators em {PREVN}): "
         f"{AT['enviadas'] - AP['enviadas']:+d} peças "
         f"({dpe:+.1f}%)." if dpe is not None else
         f"{MESN} enviou {AT['enviadas']} peças-amostra para {AT['creators_env']} creators "
         f"(mês anterior sem base comparável)."),
        (f"Qualidade do envio: taxa de cancelamento {taxa_canc(AP):.1f}% → {taxa_canc(AT):.1f}% "
         f"({AT['canceladas']} peça(s) cancelada(s) em {MESN} vs {AP['canceladas']} em {PREVN})."),
        (f"Concentração no produto '{top_prod_title}' ({top_prod_n} das {AT['enviadas']} peças) "
         f"e nos tiers {top2_tiers_lbl} ({top2_tiers_n} das {AT['enviadas']} peças enviadas)."),
        conv_line,
        (f"Ressalva: ROI (lucro÷custo) fica 'sem dado' — falta COGS/frete real da peça; "
         f"{len(undated)} solicitações do programa seguem sem data (não entram em nenhum mês "
         f"nem afetam a contagem de {MESN})."),
    ]

    return dict(
        mes=mes, PREV=PREV, MESN=MESN, ANO=ANO, PREVN=PREVN, PREVANO=PREVANO,
        MES_LBL=MES_LBL, PREV_LBL=PREV_LBL, ATTR_FIM=ATTR_FIM, exec_date=exec_date,
        rows=rows, tgt=tgt, prv=prv, tgt_rows=tgt_rows, prv_rows=prv_rows,
        TMIN=TMIN, TMAX=TMAX, AT=AT, AP=AP, ROI=ROI, atingidas=atingidas, conv=conv,
        TJ=TJ, TM=TM, prod_rows=prod_rows, prod_title=prod_title, prod_creators=prod_creators,
        cbt=cbt, undated=undated, n_pending=n_pending, n_canc_pre=n_canc_pre,
        top_prod=top_prod, top_prod_title=top_prod_title, top_prod_n=top_prod_n,
        top2_tiers=top2_tiers, top2_tiers_n=top2_tiers_n, top2_tiers_lbl=top2_tiers_lbl,
        gmv_ranked=gmv_ranked, top2_gmv=top2_gmv, top2_gmv_share=top2_gmv_share,
        dim=dim, late_nao=late_nao, top_recv=top_recv, top_recv_conv=top_recv_conv,
        KPIS=KPIS, dpe=dpe, conv_line=conv_line, resumo=resumo, TEM_VENDAS=TEM_VENDAS,
    )


def build_report(mes: str, exec_date: str | None = None, base: Path = BASE) -> dict:
    """Renderiza o relatório (xlsx + md) do mês a partir do modelo de compute_seeding()."""
    base = Path(base)
    M = compute_seeding(mes, exec_date, base)
    (mes, PREV, MESN, ANO, PREVN, PREVANO, MES_LBL, PREV_LBL, ATTR_FIM, exec_date,
     rows, tgt, prv, tgt_rows, prv_rows, TMIN, TMAX, AT, AP, ROI, atingidas, conv,
     TJ, TM, prod_rows, prod_title, prod_creators, cbt, undated, n_pending, n_canc_pre,
     top_prod, top_prod_title, top_prod_n, top2_tiers, top2_tiers_n, top2_tiers_lbl,
     gmv_ranked, top2_gmv, top2_gmv_share, dim, late_nao, top_recv, top_recv_conv,
     KPIS, dpe, conv_line, resumo, TEM_VENDAS) = (
        M["mes"], M["PREV"], M["MESN"], M["ANO"], M["PREVN"], M["PREVANO"], M["MES_LBL"],
        M["PREV_LBL"], M["ATTR_FIM"], M["exec_date"], M["rows"], M["tgt"], M["prv"],
        M["tgt_rows"], M["prv_rows"], M["TMIN"], M["TMAX"], M["AT"], M["AP"], M["ROI"],
        M["atingidas"], M["conv"], M["TJ"], M["TM"], M["prod_rows"], M["prod_title"],
        M["prod_creators"], M["cbt"], M["undated"], M["n_pending"], M["n_canc_pre"],
        M["top_prod"], M["top_prod_title"], M["top_prod_n"], M["top2_tiers"],
        M["top2_tiers_n"], M["top2_tiers_lbl"], M["gmv_ranked"], M["top2_gmv"],
        M["top2_gmv_share"], M["dim"], M["late_nao"], M["top_recv"], M["top_recv_conv"],
        M["KPIS"], M["dpe"], M["conv_line"], M["resumo"], M["TEM_VENDAS"])

    # ═══════════════════════════════════════════════════════════════════════
    # XLSX
    # ═══════════════════════════════════════════════════════════════════════
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor=NAVY)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Resumo & KPIs"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col, w in zip("BCDEF", [40, 16, 16, 14, 12]):
        ws.column_dimensions[col].width = w
    ws["B2"] = f"Relatório de Seeding — Envio de Amostras · {MES_LBL}"
    ws["B2"].font = Font(bold=True, size=16, color=NAVY)
    ws["B3"] = ("Free sample REAL da API TikTok Shop (sample_applications) · peça enviada = "
                f"COMPLETED/SHIPPED/CONTENT_PENDING · comparação vs. {PREV_LBL}")
    ws["B3"].font = Font(size=10, color="6B7280")

    ws["B5"] = "Recortes temporais"; ws["B5"].font = Font(bold=True, size=11, color=ACCENT)
    recortes = [
        ("Janela", "Mensal (mês fechado)"),
        ("Mês analisado", MES_LBL),
        ("Mês de comparação", f"{PREV_LBL} (período imediatamente anterior)"),
        (f"Data de/até ({MESN})", f"{TMIN} a {TMAX}"),
        ("Data de geração", exec_date),
        ("Janela de atribuição (conversão)", f"1º convite do mês por creator → {ATTR_FIM}"),
    ]
    for i, (k, v) in enumerate(recortes):
        r = 6 + i
        ws[f"B{r}"] = k; ws[f"B{r}"].font = Font(size=10, bold=True, color="374151")
        ws[f"C{r}"] = v; ws[f"C{r}"].font = Font(size=10, color="374151")
        ws.merge_cells(f"C{r}:F{r}")

    r0 = 6 + len(recortes) + 1
    ws[f"B{r0}"] = "Resumo executivo"; ws[f"B{r0}"].font = Font(bold=True, size=11, color=ACCENT)
    for i, line in enumerate(resumo):
        r = r0 + 1 + i
        ws[f"B{r}"] = "• " + line; ws[f"B{r}"].font = Font(size=10, color="374151")
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(f"B{r}:F{r}"); ws.row_dimensions[r].height = 30

    kr = r0 + 1 + len(resumo) + 1
    ws[f"B{kr}"] = f"KPIs — {MESN} vs. {PREV_LBL}"; ws[f"B{kr}"].font = Font(bold=True, size=11, color=ACCENT)
    kr += 1
    for j, h in enumerate(["KPI", MES_LBL, PREV_LBL, "Δ abs.", "Δ %"]):
        c = ws.cell(row=kr, column=2 + j, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, (lbl, vj, vm, kind, dt) in enumerate(KPIS):
        r = kr + 1 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(size=10)
        ws.cell(row=r, column=3, value=fmt_val(vj, kind))
        ws.cell(row=r, column=4, value=fmt_val(vm, kind))
        if dt == "pp":
            da = vj - vm
            ws.cell(row=r, column=5, value=f"{da:+.1f} p.p.")
            ws.cell(row=r, column=6, value="—")
            good = (da < 0) if "cancel" in lbl.lower() else (da > 0)
        else:
            da = vj - vm; dp = _d_pct(vj, vm)
            if kind in ("money", "num1"):
                ws.cell(row=r, column=5, value=f"{da:+,.1f}" if kind == "num1" else f"R$ {da:+,.2f}")
            else:
                ws.cell(row=r, column=5, value=f"{da:+.0f}")
            ws.cell(row=r, column=6, value=(f"{dp:+.1f}%" if dp is not None else "sem dado"))
            good = da >= 0
        dcolor = GREEN_TXT if good else RED_TXT
        for cc in (5, 6):
            cell = ws.cell(row=r, column=cc)
            cell.font = Font(size=10, bold=True, color=dcolor)
            cell.alignment = Alignment(horizontal="center")
        for cc in (2, 3, 4, 5, 6):
            cell = ws.cell(row=r, column=cc); cell.border = border
            if cc in (3, 4):
                cell.alignment = Alignment(horizontal="center"); cell.font = Font(size=10)
        if i % 2:
            for cc in (2, 3, 4):
                ws.cell(row=r, column=cc).fill = PatternFill("solid", fgColor="FAFAFA")

    nota_kr = kr + 1 + len(KPIS) + 1
    ws.cell(row=nota_kr, column=2,
            value="Nota: em cancelamento, Δ negativo é bom (verde). Taxas comparadas em pontos "
                  "percentuais (p.p.). 'Líquidas' = peças de fato despachadas (canceladas nunca saíram; "
                  "esta fonte não rastreia devolução de amostra).").font = Font(size=8, italic=True, color="9CA3AF")
    ws.merge_cells(f"B{nota_kr}:F{nota_kr}"); ws.row_dimensions[nota_kr].height = 26
    ws.cell(row=nota_kr, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    cvr = nota_kr + 2
    ws.cell(row=cvr, column=2, value=f"Conversão / ROI do seeding ({MESN})").font = Font(bold=True, size=11, color=ACCENT)
    ws.cell(row=cvr + 1, column=2,
            value=f"Vendas de afiliada da creator do 1º convite de {MESN.lower()} dela até {ATTR_FIM}. "
                  "Denominador = creators que RECEBERAM amostra.").font = Font(size=8, italic=True, color="9CA3AF")
    ws.merge_cells(f"B{cvr + 1}:F{cvr + 1}")
    _sd = "sem dado" if not (TEM_VENDAS and conv["atingidas"]) else None
    conv_rows_kpi = [
        ("Creators atingidas (receberam amostra)", f"{conv['atingidas']}", None),
        ("Já venderam pós-amostra (Sim)", _sd or f"{conv['venderam']}", GREEN),
        ("Ainda não venderam (Não)", _sd or f"{conv['nao_venderam']}", AMBER),
        ("Taxa de conversão do seeding", _sd or f"{conv['taxa']:.1f}%", GREEN),
        ("Peças vendidas atribuídas (un.)", _sd or f"{conv['unid']}", None),
        ("GMV gerado atribuído (R$)", _sd or f"R$ {conv['gmv_gerado']:,.2f}", GREEN),
        ("GMV gerado por peça enviada (R$)", _sd or f"R$ {conv['gmv_por_peca']:,.2f}", GREEN),
        ("ROI (lucro ÷ custo)", "sem dado — falta COGS/frete real", RED),
    ]
    for i, (lbl, val, fill) in enumerate(conv_rows_kpi):
        r = cvr + 2 + i
        ws.cell(row=r, column=2, value=lbl).font = Font(size=10)
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=GREY)
        ws.cell(row=r, column=2).border = border
        cc = ws.cell(row=r, column=3, value=val)
        cc.font = Font(bold=True, size=10); cc.border = border
        cc.alignment = Alignment(horizontal="right")
        if fill:
            cc.fill = PatternFill("solid", fgColor=fill)

    # ── aba detalhe por creator (helper) ────────────────────────────────────
    def write_creator_sheet(title, creators, agg_, roi_map=None):
        w = wb.create_sheet(title)
        w.sheet_view.showGridLines = False
        cols = [
            ("#", 4), ("Creator", 22), ("@ TikTok", 20), ("Tier (estim.)", 12),
            ("GMV afiliação (R$)", 16), ("Peças pedidas", 10), ("Enviadas", 9),
            ("Canceladas", 10), ("Líquidas", 9), ("SKUs distintos", 10),
            ("Produtos distintos", 11), ("1º convite", 11), ("Último convite", 12),
            ("Status envio", 12),
        ]
        if roi_map is not None:
            cols += [("Vendeu pós-amostra?", 15), ("Peças vendidas (un.)", 12),
                     ("GMV gerado (R$)", 15), ("GMV gerado / peça env. (R$)", 15),
                     ("ROI amostra", 20)]
        ncols = len(cols)
        C_VENDEU, C_PVEND, C_GGER, C_GPP, C_ROI = 15, 16, 17, 18, 19
        for j, (h, wd) in enumerate(cols, 1):
            c = w.cell(row=1, column=j, value=h)
            c.font = hfont; c.fill = hfill; c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            w.column_dimensions[get_column_letter(j)].width = wd
            if j == 5:
                cm = Comment(GMV_AFIL_NOTA, "Relatório Seeding"); cm.width, cm.height = 360, 230
                c.comment = cm
        w.row_dimensions[1].height = 30
        w.freeze_panes = "C2"
        center = {1, 4, 6, 7, 8, 9, 10, 11, 12, 13, 14, C_VENDEU, C_PVEND}
        for i, c in enumerate(creators):
            r = 2 + i
            status = "Enviado" if c["enviadas"] > 0 else "Nada enviado"
            vals = [i + 1, c["nick"], "@" + c["handle"], c["tier"], c["gmv"],
                    c["pedidas"], c["enviadas"], c["canceladas"], c["liquidas"],
                    c["skus"], c["produtos"], c["d_ini"] or "sem dado",
                    c["d_fim"] or "sem dado", status]
            if roi_map is not None:
                roi = roi_map.get(c["canon"], {})
                if c["enviadas"] == 0:
                    vals += ["n/a (não recebida)", "n/a", "n/a", "n/a", "n/a"]
                elif not roi.get("ok"):
                    vals += ["sem dado", "sem dado", "sem dado", "sem dado", "sem dado"]
                else:
                    gpp = roi["gmv"] / c["enviadas"] if c["enviadas"] else 0
                    vals += ["Sim" if roi["vendeu"] else "Não", roi["unid"], roi["gmv"],
                             round(gpp, 2), "sem dado (falta COGS/frete)"]
            for j, v in enumerate(vals, 1):
                cell = w.cell(row=r, column=j, value=v)
                cell.border = border; cell.font = Font(size=9)
                if j in center:
                    cell.alignment = Alignment(horizontal="center")
                if j == 5:
                    cell.number_format = money
                if j == 4:
                    cell.font = Font(size=9, bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor=TIER_FILL[c["tier"]])
                    cell.alignment = Alignment(horizontal="center")
                if j == 14:
                    cell.font = Font(size=9, bold=True,
                                     color=GREEN_TXT if status == "Enviado" else AMBER_TXT)
                if roi_map is not None and c["enviadas"] > 0 and roi_map.get(c["canon"], {}).get("ok"):
                    if j in (C_GGER, C_GPP):
                        cell.number_format = money
                    if j == C_VENDEU:
                        cell.font = Font(size=9, bold=True,
                                         color=GREEN_TXT if v == "Sim" else AMBER_TXT)
                    if j == C_ROI:
                        cell.font = Font(size=8, italic=True, color="9CA3AF")
            if c["enviadas"] == 0:
                for j in range(1, ncols + 1):
                    if j != 4:
                        w.cell(row=r, column=j).fill = PatternFill("solid", fgColor=RED)
            elif i % 2:
                for j in range(1, ncols + 1):
                    if j != 4:
                        w.cell(row=r, column=j).fill = PatternFill("solid", fgColor="FAFAFA")
        last = 1 + len(creators); tr = last + 1
        w.cell(row=tr, column=3, value="TOTAL").font = Font(bold=True)
        tot_cols = [(6, "pedidas"), (7, "enviadas"), (8, "canceladas"), (9, "liquidas")]
        if roi_map is not None:
            tot_cols += [(C_PVEND, "unid"), (C_GGER, "gmv")]
        for j, key in tot_cols:
            L = get_column_letter(j)
            cell = w.cell(row=tr, column=j, value=f"=SUBTOTAL(109,{L}2:{L}{last})")
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.alignment = Alignment(horizontal="center")
            if j == C_GGER:
                cell.number_format = money
        if roi_map is not None:
            cell = w.cell(row=tr, column=C_GPP,
                          value=f"=IF({get_column_letter(7)}{tr}>0,"
                                f"{get_column_letter(C_GGER)}{tr}/{get_column_letter(7)}{tr},0)")
            cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.number_format = money; cell.alignment = Alignment(horizontal="center")
        w.cell(row=tr, column=2, value=f"{agg_['creators_tot']} creators").font = Font(bold=True, size=9)
        if last >= 2:
            w.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last}"
        return w

    write_creator_sheet(f"Detalhe {MESN}", tgt, AT, roi_map=ROI)

    # ── aba itens linha a linha ─────────────────────────────────────────────
    wi = wb.create_sheet(f"Itens {MESN} (linha a linha)")
    wi.sheet_view.showGridLines = False
    icols = [("#", 4), ("Data convite", 11), ("Creator", 20), ("@ TikTok", 18),
             ("Tier (estim.)", 12), ("Produto (título)", 46), ("SKU / variação", 22),
             ("Status", 20), ("Enviada?", 9), ("Comissão %", 10), ("Order ID", 20)]
    for j, (h, wd) in enumerate(icols, 1):
        c = wi.cell(row=1, column=j, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wi.column_dimensions[get_column_letter(j)].width = wd
    wi.row_dimensions[1].height = 28; wi.freeze_panes = "C2"
    items = sorted(tgt_rows, key=lambda r: (r["status"] not in SENT, _ck(r["creator"]), r["convite_data"]))
    for i, r in enumerate(items):
        rr = 2 + i; env = r["status"] in SENT; gmv = float(r["creator_gmv"] or 0)
        try:
            rate = f"{float(r['commission_rate']) * 100:.0f}%"
        except (ValueError, TypeError):
            rate = "sem dado"
        vals = [i + 1, r["convite_data"] or "sem dado", r["nickname"] or r["creator"],
                "@" + r["creator"], _tier(gmv), r["title"], r["sku_name"], r["status"],
                "Sim" if env else "Não", rate,
                r["order_id"] if r["order_id"] not in ("0", "") else "sem dado"]
        for j, v in enumerate(vals, 1):
            cell = wi.cell(row=rr, column=j, value=v)
            cell.border = border; cell.font = Font(size=9)
            if j in (1, 2, 5, 9, 10):
                cell.alignment = Alignment(horizontal="center")
            if j == 5:
                cell.font = Font(size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=TIER_FILL[_tier(gmv)])
            if j == 9:
                cell.font = Font(size=9, bold=True, color=GREEN_TXT if env else RED_TXT)
        if not env:
            for j in range(1, len(icols) + 1):
                if j != 5:
                    wi.cell(row=rr, column=j).fill = PatternFill("solid", fgColor=RED)
        elif i % 2:
            for j in range(1, len(icols) + 1):
                if j != 5:
                    wi.cell(row=rr, column=j).fill = PatternFill("solid", fgColor="FAFAFA")
    if items:
        wi.auto_filter.ref = f"A1:{get_column_letter(len(icols))}{1 + len(items)}"

    write_creator_sheet(f"Comparativo {PREVN}", prv, AP)

    # ── aba segmentos ───────────────────────────────────────────────────────
    wsg = wb.create_sheet("Segmentos"); wsg.sheet_view.showGridLines = False
    wsg.column_dimensions["A"].width = 2
    wsg["B2"] = f"Segmentações — {MES_LBL} (peças enviadas)"
    wsg["B2"].font = Font(bold=True, size=14, color=NAVY)
    wsg["B4"] = "Por tier de creator (enviadas)"; wsg["B4"].font = Font(bold=True, size=11, color=ACCENT)
    for j, h in enumerate(["Tier", MESN, PREVN, "Δ abs.", "% do total"]):
        c = wsg.cell(row=5, column=2 + j, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center")
        wsg.column_dimensions[get_column_letter(2 + j)].width = 16
    totj = sum(TJ.values()) or 1
    for i, t in enumerate(TIER_ORDER):
        r = 6 + i; vj, vm = TJ.get(t, 0), TM.get(t, 0)
        cells = [t, vj, vm, f"{vj - vm:+d}", f"{vj / totj * 100:.1f}%"]
        for j, v in enumerate(cells):
            cell = wsg.cell(row=r, column=2 + j, value=v); cell.border = border; cell.font = Font(size=10)
            if j > 0:
                cell.alignment = Alignment(horizontal="center")
            if j == 0:
                cell.font = Font(size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=TIER_FILL[t])
    rtot = 6 + len(TIER_ORDER)
    wsg.cell(row=rtot, column=2, value="TOTAL").font = Font(bold=True)
    wsg.cell(row=rtot, column=3, value=sum(TJ.values())).font = Font(bold=True)
    wsg.cell(row=rtot, column=4, value=sum(TM.values())).font = Font(bold=True)
    for cc in (3, 4):
        wsg.cell(row=rtot, column=cc).alignment = Alignment(horizontal="center")
        wsg.cell(row=rtot, column=cc).fill = PatternFill("solid", fgColor=LIGHT)

    pr0 = rtot + 3
    wsg.cell(row=pr0, column=2, value=f"Por produto amostrado ({MESN} · enviadas)").font = Font(bold=True, size=11, color=ACCENT)
    for j, h in enumerate(["#", "Produto", "Peças enviadas", "Creators distintas"]):
        c = wsg.cell(row=pr0 + 1, column=2 + j, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center")
    wsg.column_dimensions["C"].width = 52
    for i, (pid, n) in enumerate(prod_rows):
        r = pr0 + 2 + i
        cells = [i + 1, prod_title[pid], n, len(prod_creators[pid])]
        for j, v in enumerate(cells):
            cell = wsg.cell(row=r, column=2 + j, value=v); cell.border = border; cell.font = Font(size=9)
            if j in (0, 2, 3):
                cell.alignment = Alignment(horizontal="center")
        if i % 2:
            for j in range(4):
                wsg.cell(row=r, column=2 + j).fill = PatternFill("solid", fgColor="FAFAFA")

    cv0 = pr0 + 2 + len(prod_rows) + 2
    wsg.cell(row=cv0, column=2, value=f"Conversão pós-amostra por tier ({MESN} · atingidas)").font = Font(bold=True, size=11, color=ACCENT)
    for j, h in enumerate(["Tier", "Atingidas", "Venderam", "Conversão", "Peças env.", "GMV gerado (R$)", "GMV / peça env."]):
        c = wsg.cell(row=cv0 + 1, column=2 + j, value=h)
        c.font = hfont; c.fill = hfill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ri = cv0 + 2
    for t in TIER_ORDER:
        d = cbt[t]
        if d["ating"] == 0:
            continue
        txa = f"{d['vend'] / d['ating'] * 100:.1f}%" if d["ating"] else "—"
        gpp = f"R$ {d['gmv'] / d['env']:,.2f}" if d["env"] else "—"
        cells = [t, d["ating"], d["vend"], txa, d["env"], f"R$ {d['gmv']:,.2f}", gpp]
        for j, v in enumerate(cells):
            cell = wsg.cell(row=ri, column=2 + j, value=v); cell.border = border; cell.font = Font(size=9)
            if j > 0:
                cell.alignment = Alignment(horizontal="center")
            if j == 0:
                cell.font = Font(size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=TIER_FILL[t])
        ri += 1
    cells = ["TOTAL", conv["atingidas"], conv["venderam"], f"{conv['taxa']:.1f}%",
             conv["enviadas"], f"R$ {conv['gmv_gerado']:,.2f}", f"R$ {conv['gmv_por_peca']:,.2f}"]
    for j, v in enumerate(cells):
        cell = wsg.cell(row=ri, column=2 + j, value=v)
        cell.font = Font(size=9, bold=True); cell.fill = PatternFill("solid", fgColor=LIGHT); cell.border = border
        if j > 0:
            cell.alignment = Alignment(horizontal="center")

    # ── aba metodologia ─────────────────────────────────────────────────────
    wm = wb.create_sheet("Metodologia & Fontes"); wm.sheet_view.showGridLines = False
    wm.column_dimensions["B"].width = 110
    meto = [
        ("Metodologia & Fontes", 14, True, NAVY),
        ("", 10, False, "000000"),
        ("Fonte das peças (100% API TikTok Shop)", 11, True, ACCENT),
        ("• Endpoint: POST /affiliate_seller/202409/sample_applications/search — pedidos de free sample.", 10, False, "374151"),
        ("• Persistido em warehouse/raw_sample_applications.csv (fonte de verdade do seeding). 1 linha = 1 creator × 1 SKU = 1 peça.", 10, False, "374151"),
        ("• Peça ENVIADA = status COMPLETED / SHIPPED / CONTENT_PENDING. CANCELADA = *_CANCELLED. PENDING/AWAITING = ainda não despachada.", 10, False, "374151"),
        ("", 10, False, "000000"),
        ("Recorte temporal (mês)", 11, True, ACCENT),
        ("• 'Mês' = mês de convite_data = create_time do pedido de amostra. Proxy de quando a peça saiu.", 10, False, "374151"),
        (f"• Alvo = {MES_LBL} (convite_data em {mes}); comparação = {PREV_LBL} ({PREV}). Sempre vs. mês imediatamente anterior.", 10, False, "374151"),
        ("• 'Líquidas' = peças de fato despachadas (= enviadas). Canceladas nunca saíram; esta fonte não rastreia devolução de amostra.", 10, False, "374151"),
        ("", 10, False, "000000"),
        ("Coluna 'GMV afiliação (R$)' — o que é e o que NÃO é", 11, True, ACCENT),
        ("• É o creator_gmv da API por creator = GMV BRUTO ACUMULADO dela como afiliada (histórico / 'cacife' all-time).", 10, False, "374151"),
        ("• Uso: SÓ proxy para estimar o Tier (por isso 'estim.'). Tier canônico (etl_v2.TIER_RULES):", 10, False, "374151"),
        ("   Black ≥ R$500k · Diamond ≥ R$150k · Gold ≥ R$80k · Silver ≥ R$50k · Bronze ≥ R$20k · Iniciante < R$20k.", 10, False, "374151"),
        ("• NÃO é o retorno da amostra (isso é 'GMV gerado'). NÃO é o GMV oficial do programa (GMV líquido acumulado da performance_periods).", 10, False, AMBER_TXT),
        ("", 10, False, "000000"),
        ("Conversão / ROI do seeding — atribuição de vendas pós-amostra", 11, True, ACCENT),
        ("• Fonte de vendas: Affiliate Orders API (POST /affiliate_seller/202410/orders/search), materializada em warehouse/raw_vendas_seeding_<mês>.csv.", 10, False, "374151"),
        (f"• Janela por creator: do 1º convite do mês DELA até {ATTR_FIM} (data de execução). Pull global: 1º dia do mês → {ATTR_FIM}.", 10, False, "374151"),
        ("• 'Vendeu pós-amostra?' = Sim se gerou ≥1 venda de afiliada (qualquer produto Rhode) após o convite. 'n/a' = não recebeu amostra.", 10, False, "374151"),
        ("• 'Peças vendidas' = soma de quantity. 'GMV gerado' = soma de estimated_commission_base (bruto). Case por @handle (aliases colapsados).", 10, False, "374151"),
        ("• Atribuição por PROXIMIDADE TEMPORAL (vendeu depois de receber), não causalidade — não isola o produto amostrado.", 10, False, AMBER_TXT),
        ("• 'ROI amostra' (lucro ÷ custo) = SEM DADO: custo real da peça (COGS) e frete não estão nas fontes. GMV gerado por peça = GMV ÷ peças enviadas.", 10, False, AMBER_TXT),
        ("", 10, False, "000000"),
        ("Aliases de handle (mesma creator)", 11, True, ACCENT),
        ("• @natmarquesss → @natmarquesvi · @tacianecreator + @tacirecomenda → @tacianetorress. Somados como 1 creator.", 10, False, "374151"),
        ("", 10, False, "000000"),
        ("Ressalvas / 'sem dado'", 11, True, AMBER_TXT),
        (f"• {len(undated)} solicitações do PROGRAMA estão sem data (order_id=0): {n_pending} PENDING + {n_canc_pre} canceladas pré-pedido. Não entram em nenhum mês nem são 'enviadas'.", 10, False, "374151"),
        ("• Conversão/ROI é calculada só para o mês-alvo (a coleta de vendas cobre o mês → hoje). O mês de comparação não tem essas colunas.", 10, False, "374151"),
    ]
    for i, (txt, sz, bold, color) in enumerate(meto, 1):
        c = wm.cell(row=i, column=2, value=txt)
        c.font = Font(size=sz, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    out_dir = base / "relatorios" / mes
    out_dir.mkdir(parents=True, exist_ok=True)
    outx = out_dir / f"Relatorio Seeding {MESN} {ANO}_{exec_date}.xlsx"
    outm = out_dir / f"Relatorio Seeding {MESN} {ANO}_{exec_date}.md"
    wb.save(outx)

    # ═══════════════════════════════════════════════════════════════════════
    # MARKDOWN
    # ═══════════════════════════════════════════════════════════════════════
    def md_delta(vj, vm, kind, dt):
        if dt == "pp":
            return f"{vj - vm:+.1f} p.p.", "—"
        da = vj - vm; dp = _d_pct(vj, vm)
        a = (f"R$ {da:+,.2f}" if kind == "money" else
             f"{da:+,.1f}" if kind == "num1" else f"{da:+.0f}")
        return a, (f"{dp:+.1f}%" if dp is not None else "sem dado")

    L = []
    L.append(f"# Relatório de Seeding — Envio de Amostras · {MES_LBL}\n")
    L.append(f"*Gerado em {exec_date} · Fonte: API TikTok Shop `sample_applications` + `affiliate/orders` · "
             f"Comparação vs. {PREV_LBL}*\n")
    L.append(f"**Recortes temporais** — Janela: Mensal (mês fechado) · Mês: {MES_LBL} · "
             f"Data de/até: {TMIN} a {TMAX} · Atribuição de vendas: 1º convite do mês por creator → {ATTR_FIM}\n")

    L.append("## 1. Resumo executivo\n")
    for line in resumo:
        L.append(f"- {line}")
    L.append("")

    L.append(f"## 2. KPIs — {MESN} vs. {PREV_LBL}\n")
    L.append(f"| KPI | {MES_LBL} | {PREV_LBL} | Δ abs. | Δ % |")
    L.append("|---|---:|---:|---:|---:|")
    for lbl, vj, vm, kind, dt in KPIS:
        da, dp = md_delta(vj, vm, kind, dt)
        L.append(f"| {lbl} | {fmt_val(vj, kind)} | {fmt_val(vm, kind)} | {da} | {dp} |")
    L.append("")
    L.append("> Em *cancelamento*, Δ negativo é positivo para o negócio. Taxas comparadas em pontos percentuais (p.p.). "
             "'Líquidas' = peças de fato despachadas.\n")

    L.append("## 3. Performance por segmento\n")
    L.append("### 3.1 Por tier de creator (peças enviadas)\n")
    L.append(f"| Tier | {MESN} | {PREVN} | Δ abs. | % do total |")
    L.append("|---|---:|---:|---:|---:|")
    for t in TIER_ORDER:
        vj, vm = TJ.get(t, 0), TM.get(t, 0)
        L.append(f"| {t} | {vj} | {vm} | {vj - vm:+d} | {vj / totj * 100:.1f}% |")
    L.append(f"| **TOTAL** | **{sum(TJ.values())}** | **{sum(TM.values())}** | "
             f"**{sum(TJ.values()) - sum(TM.values()):+d}** | 100.0% |")
    L.append("")
    L.append(f"### 3.2 Por produto amostrado ({MESN} · enviadas)\n")
    L.append("| # | Produto | Peças enviadas | Creators distintas |")
    L.append("|---:|---|---:|---:|")
    for i, (pid, n) in enumerate(prod_rows):
        L.append(f"| {i + 1} | {prod_title[pid]} | {n} | {len(prod_creators[pid])} |")
    L.append("")
    L.append(f"### 3.3 Conversão pós-amostra por tier ({MESN} · creators atingidas)\n")
    L.append("| Tier | Atingidas | Venderam | Conversão | Peças env. | GMV gerado (R$) | GMV / peça env. |")
    L.append("|---|---:|---:|---:|---:|---:|---:|")
    for t in TIER_ORDER:
        d = cbt[t]
        if d["ating"] == 0:
            continue
        gpp = f"R$ {d['gmv'] / d['env']:,.2f}" if d["env"] else "—"
        L.append(f"| {t} | {d['ating']} | {d['vend']} | {d['vend'] / d['ating'] * 100:.1f}% | "
                 f"{d['env']} | R$ {d['gmv']:,.2f} | {gpp} |")
    L.append(f"| **TOTAL** | **{conv['atingidas']}** | **{conv['venderam']}** | **{conv['taxa']:.1f}%** | "
             f"**{conv['enviadas']}** | **R$ {conv['gmv_gerado']:,.2f}** | **R$ {conv['gmv_por_peca']:,.2f}** |")
    L.append("")
    L.append(f"> Atribuição: vendas de afiliada da creator do 1º convite de {MESN.lower()} dela até {ATTR_FIM} "
             "(fonte: Affiliate Orders API). Proximidade temporal, não causalidade. ROI (lucro÷custo) = sem dado (falta COGS/frete real).\n")

    L.append(f"### 3.4 Detalhe por creator ({MESN} · ordenado por peças enviadas)\n")
    L.append("| # | Creator | @ TikTok | Tier (estim.) | GMV afiliação | Pedidas | Enviadas | Canc. | SKUs | Vendeu? | Peças vend. | GMV gerado | GMV/peça env. | 1º convite |")
    L.append("|---:|---|---|---|---:|---:|---:|---:|---:|:--:|---:|---:|---:|---|")
    for i, c in enumerate(tgt):
        roi = ROI.get(c["canon"], {})
        if c["enviadas"] == 0:
            vd, pv, gg, gpp = "n/a", "n/a", "n/a", "n/a"
        elif not roi.get("ok"):
            vd, pv, gg, gpp = "sem dado", "sem dado", "sem dado", "sem dado"
        else:
            vd = "Sim" if roi["vendeu"] else "Não"; pv = f"{roi['unid']}"
            gg = f"R$ {roi['gmv']:,.2f}"; gpp = f"R$ {roi['gmv'] / c['enviadas']:,.2f}"
        L.append(f"| {i + 1} | {c['nick']} | @{c['handle']} | {c['tier']} | R$ {c['gmv']:,.2f} | "
                 f"{c['pedidas']} | {c['enviadas']} | {c['canceladas']} | {c['skus']} | "
                 f"{vd} | {pv} | {gg} | {gpp} | {c['d_ini'] or 'sem dado'} |")
    L.append(f"| | **TOTAL ({AT['creators_tot']} creators)** | | | | **{AT['pedidas']}** | "
             f"**{AT['enviadas']}** | **{AT['canceladas']}** | | | **{conv['unid']}** | "
             f"**R$ {conv['gmv_gerado']:,.2f}** | **R$ {conv['gmv_por_peca']:,.2f}** | |")
    L.append("")
    L.append("> 'GMV afiliação' é o cacife histórico all-time (proxy de tier), NÃO o retorno da amostra. "
             "'GMV gerado' é a conversão (vendas de afiliada pós-convite). 'n/a' = não recebeu amostra. "
             "'ROI amostra' na planilha = sem dado (falta COGS/frete).\n")

    L.append("## 4. Destaques e alertas\n")
    L.append("**Destaques positivos**")
    if AT["canceladas"] <= AP["canceladas"]:
        L.append(f"- Cancelamento: {taxa_canc(AP):.1f}% → {taxa_canc(AT):.1f}% "
                 f"({AP['canceladas']} peça(s) perdida(s) em {PREVN} vs. {AT['canceladas']} em {MESN}).")
    if TEM_VENDAS and conv["atingidas"]:
        L.append(f"- Conversão em {conv['taxa']:.1f}%: {conv['venderam']} das {conv['atingidas']} atingidas "
                 f"venderam pós-convite, R$ {conv['gmv_gerado']:,.2f} de GMV gerado ({conv['unid']} un.), "
                 f"R$ {conv['gmv_por_peca']:,.2f} por peça enviada.")
        if top2_gmv:
            quem = " e ".join(f"@{c['handle']} ({c['tier']}, R$ {g:,.2f})" for c, g in top2_gmv)
            L.append(f"- Concentração de retorno: {quem} puxaram ~{top2_gmv_share:.1f}% do GMV gerado atribuído.")
    if top_prod[0]:
        L.append(f"- Produto herói: '{top_prod_title}' puxou {top_prod_n} das {AT['enviadas']} peças enviadas.")
    L.append("")
    L.append("**Alertas**")
    if dpe is not None and AT["enviadas"] < AP["enviadas"]:
        L.append(f"- Volume caiu {abs(dpe):.1f}% vs. {PREVN} ({AP['enviadas']} → {AT['enviadas']} peças); "
                 f"alcance {AP['creators_env']} → {AT['creators_env']} creators.")
    elif dpe is not None and AT["enviadas"] > AP["enviadas"]:
        L.append(f"- Volume subiu {dpe:.1f}% vs. {PREVN} — garantir que a régua de qualidade/cancelamento acompanhe a escala.")
    if top_recv is not None and top_recv_conv is False:
        L.append(f"- Maior recebedora sem conversão (ainda): @{top_recv['handle']} ({top_recv['tier']}) recebeu "
                 f"{top_recv['enviadas']} peças e tem R$ 0,00 de GMV gerado na janela — investigar conteúdo/link de afiliada.")
    if late_nao:
        L.append(f"- Efeito data de convite: {late_nao} das {conv['nao_venderam']} que não venderam foram convidadas "
                 f"nos últimos ~7 dias do mês (janela curta) — a conversão real ainda vai maturar; reavaliar no fechamento seguinte.")
    L.append(f"- {len(undated)} solicitações do programa seguem sem data ({n_pending} PENDING + {n_canc_pre} canceladas pré-pedido). "
             "As PENDING podem ser fila a despachar.")
    L.append("- ROI (lucro÷custo) segue 'sem dado': falta COGS/frete real. GMV gerado está medido, mas o lucro depende do custo real.")
    L.append("")

    prox = prev_month  # não usado; placeholder para clareza
    L.append("## 5. Recomendações para o próximo período\n")
    L.append("1. **Manter a régua de cancelamento** e escalar volume só sobre tiers que retornam "
             "(priorizar quem já converteu histórico).")
    L.append("2. **Limpar a fila PENDING**: revisar as solicitações sem data — despachar as válidas ou cancelar formalmente.")
    L.append("3. **Fechar o ROI real**: carregar COGS + frete por peça (hoje 'sem dado') para converter o GMV gerado já medido em lucro/ROI.")
    if conv["nao_venderam"]:
        alvo = f"@{top_recv['handle']}" if (top_recv and top_recv_conv is False) else "as atingidas com R$ 0"
        L.append(f"4. **Cobrar quem recebeu e não converteu**: {conv['nao_venderam']} atingidas com R$ 0 de GMV gerado "
                 f"(incl. {alvo}). Follow-up de conteúdo/link de afiliada.")
    L.append("5. **Diversificar além do produto herói** e definir critério de conteúdo antes de mandar amostra "
             "para tiers baixos, para não gastar peça sem retorno.")
    L.append("")
    L.append("---")
    L.append(f"*Metodologia: peça enviada = COMPLETED/SHIPPED/CONTENT_PENDING; mês = convite_data. Tier estimado pelo "
             f"GMV de afiliação da API. Conversão/ROI = vendas de afiliada (Affiliate Orders API) do 1º convite do mês até "
             f"{ATTR_FIM}, atribuição por proximidade temporal (não causal); ROI lucro÷custo = sem dado (falta COGS/frete). "
             f"Aliases somados como mesma creator. Total da API no acumulado do programa: {len(rows)} pedidos.*")

    outm.write_text("\n".join(L), encoding="utf-8")

    return {
        "mes": mes, "prev": PREV, "exec_date": exec_date,
        "xlsx": str(outx), "md": str(outm),
        "enviadas": AT["enviadas"], "canceladas": AT["canceladas"],
        "creators_atingidas": conv["atingidas"], "creators_env": AT["creators_env"],
        "conv_venderam": conv["venderam"], "conv_taxa": round(conv["taxa"], 1),
        "gmv_gerado": conv["gmv_gerado"], "delta_enviadas_pct": (round(dpe, 1) if dpe is not None else None),
        "tem_vendas": TEM_VENDAS,
    }
