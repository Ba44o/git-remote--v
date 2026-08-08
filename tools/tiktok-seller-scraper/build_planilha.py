#!/usr/bin/env python3
"""Monta o relatorio de criativos GMV Max a partir das 3 janelas raspadas.
Le dados/campanhas/tiktok/criativos/<janela>_p<N>.json e gera o .xlsx no Desktop + repo."""
import json, math, statistics, shutil, re, os
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

BRT = timezone(timedelta(hours=-3))
DATA = "/Users/user/Documents/VS Claude Teste/dados/campanhas/tiktok/criativos"
DESK = "/Users/user/Desktop/Relatorio Criativos GMV Max 3 Janelas_2026-07-23.xlsx"
REPO = "/Users/user/Documents/VS Claude Teste/relatorios/2026-07/Relatorio Criativos GMV Max 3 Janelas_2026-07-23.xlsx"
DECAY = 0.20

def br(s):
    s=str(s).replace('BRL','').strip()
    try: return float(s.replace('.','').replace(',','.'))
    except: return 0.0
def us(s):
    try: return float(str(s).replace(',','').strip())
    except: return 0.0
def pc(s):
    try: return float(str(s).replace('%','').strip())
    except: return 0.0
def roif(s):
    try: return float(str(s).split()[0].replace(',','.'))
    except: return 0.0

def load(win):
    recs={}; url=None
    for p in (1,2,3):
        path=f"{DATA}/{win}_p{p}.json"
        if not os.path.exists(path): continue
        d=json.load(open(path)); url=url or d['url']
        for r in d['rows']:
            if len(r)<21 or r[1] in recs: continue
            recs[str(r[1])]=dict(id=str(r[1]), nome=r[0], creator=r[2], custo=br(r[6]),
                pedidos=int(us(r[7])), cpp=br(r[8]), receita=br(r[9]), roi=roif(r[10]),
                impr=us(r[11]), ctr=pc(r[13]), conv=pc(r[14]), hook=pc(r[15]), s6=pc(r[16]),
                p25=pc(r[17]), p50=pc(r[18]), p75=pc(r[19]), p100=pc(r[20]))
    a=re.search(r'campaign_start_date=(\d+)',url); b=re.search(r'campaign_end_date=(\d+)',url)
    per="%s a %s"%(datetime.fromtimestamp(int(a.group(1))/1000,BRT).strftime('%d/%m/%y'),
                   datetime.fromtimestamp(int(b.group(1))/1000,BRT).strftime('%d/%m/%y'))
    return recs, per

WINS={'jun26':None,'jul26':None,'d90':None}
PER={}
for w in WINS: WINS[w],PER[w]=load(w)

# ---------- estilos ----------
RED='FE2C55'; GREY='888888'; DARK='1A1A1A'; LT='F5F5F5'; YEL='FFF2CC'
G='63BE7B'; Y='FFEB84'; RD='F8696B'; BLUE='0563C1'
MONEY='"R$" #,##0.00'; M0='"R$" #,##0'; ROI2='0.00'; P1='0.0"%"'; P2='0.00"%"'; PCT='0%'; MUL='0.0"x"'; TXT='@'; INT='#,##0'
wb=openpyxl.Workbook(); wb.remove(wb.active)
def idlink(ws,r,c,vid):
    cell=ws.cell(row=r,column=c,value=vid); cell.number_format=TXT
    cell.hyperlink="https://www.tiktok.com/@_/video/"+vid; cell.font=Font(color=BLUE,underline='single',size=9); return cell
def H(ws,row,cols,size=9):
    for i,h in enumerate(cols,1):
        c=ws.cell(row=row,column=i,value=h); c.font=Font(bold=True,color='FFFFFF',size=size)
        c.fill=PatternFill('solid',fgColor=RED); c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
def zebra(ws,r,ncol):
    for c in range(1,ncol+1): ws.cell(row=r,column=c).fill=PatternFill('solid',fgColor=LT)
def cscale(ws,rng,rev=False):
    a,b=(RD,G) if not rev else (G,RD)
    ws.conditional_formatting.add(rng, ColorScaleRule(start_type='min',start_color=a,mid_type='percentile',mid_value=50,mid_color=Y,end_type='max',end_color=b))
def title(ws,t,sub):
    ws['A1']=t; ws['A1'].font=Font(bold=True,size=14,color=RED)
    ws['A2']=sub; ws['A2'].font=Font(size=9,italic=True,color=GREY); ws['A2'].alignment=Alignment(wrap_text=True); ws.row_dimensions[2].height=28

# ================= ABA 1: RESUMO =================
ws=wb.create_sheet('Resumo Janelas'); ws.sheet_view.showGridLines=False
title(ws,'Criativos GMV Max — 3 Janelas de Data','Mesmo filtro (Veiculado) nas 3 janelas: a data muda as MÉTRICAS, não a lista. Raspado do Seller via Chrome, período lido da URL. Campanha 1837899752347697.')
H(ws,4,['Janela','Período','Pra quê','Criativos c/ gasto','Custo','Receita','ROI','Pedidos','Flagship (custo · ROI)'])
purpose={'jun26':'Mês fechado (base estável)','jul26':'Mês corrente (estado atual)','d90':'90 dias (inclui pausados/histórico)'}
r=5
for w in ['jun26','jul26','d90']:
    com=[x for x in WINS[w].values() if x['custo']>0]
    tc=sum(x['custo'] for x in com); tr=sum(x['receita'] for x in com)
    fl=WINS[w].get('7610417484289690901')
    ws.cell(row=r,column=1,value=w).font=Font(bold=True)
    ws.cell(row=r,column=2,value=PER[w]); ws.cell(row=r,column=3,value=purpose[w])
    ws.cell(row=r,column=4,value=len(com))
    ws.cell(row=r,column=5,value=round(tc)).number_format=M0
    ws.cell(row=r,column=6,value=round(tr)).number_format=M0
    ws.cell(row=r,column=7,value=round(tr/tc,2)).number_format=ROI2
    ws.cell(row=r,column=8,value=sum(x['pedidos'] for x in com))
    ws.cell(row=r,column=9,value='R$%.0f · %.2f'%(fl['custo'],fl['roi']) if fl else '-')
    r+=1
for col,wd in zip('ABCDEFGHI',[9,20,30,17,13,13,7,9,20]): ws.column_dimensions[col].width=wd

# ================= ABAS 2-4: DETALHE POR JANELA =================
def aba_detalhe(win):
    ws=wb.create_sheet('Criativos %s'%win); ws.sheet_view.showGridLines=False
    com=sorted([x for x in WINS[win].values() if x['custo']>0], key=lambda z:-z['roi'])
    tc=sum(x['custo'] for x in com); tr=sum(x['receita'] for x in com)
    title(ws,'Criativos %s — %s'%(win,PER[win]),'%d criativos com gasto · Custo R$%s · Receita R$%s · ROI %.2f. Clique no ID p/ abrir o vídeo.'%(len(com),format(tc,',.0f'),format(tr,',.0f'),tr/tc))
    HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','Custo','Pedidos','ROI','Custo/pedido','Conversão','Hook 2s','Ret. 6s','Receita'])
    f=HR+1
    for k,x in enumerate(com):
        rr=f+k
        idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'])
        ws.cell(row=rr,column=4,value=x['custo']).number_format=MONEY
        ws.cell(row=rr,column=5,value=x['pedidos'])
        ws.cell(row=rr,column=6,value=round(x['roi'],2)).number_format=ROI2
        ws.cell(row=rr,column=7,value=x['cpp']).number_format=MONEY
        ws.cell(row=rr,column=8,value=x['conv']).number_format=P2
        ws.cell(row=rr,column=9,value=x['hook']).number_format=P1
        ws.cell(row=rr,column=10,value=x['s6']).number_format=P1
        ws.cell(row=rr,column=11,value=x['receita']).number_format=MONEY
        if k%2: zebra(ws,rr,11)
    last=f+len(com)-1
    cscale(ws,f'F{f}:F{last}'); cscale(ws,f'H{f}:H{last}'); cscale(ws,f'I{f}:I{last}'); cscale(ws,f'G{f}:G{last}',rev=True)
    for col,wd in zip('ABCDEFGHIJK',[21,40,16,12,8,7,12,11,9,9,12]): ws.column_dimensions[col].width=wd
    ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:K{last}'
for w in ['jun26','jul26','d90']: aba_detalhe(w)

# ================= ABA 5: FADIGA DE HOOK (jun -> jul) =================
ws=wb.create_sheet('Fadiga de Hook'); ws.sheet_view.showGridLines=False
jun,jul=WINS['jun26'],WINS['jul26']
com=[k for k in jun if k in jul and jun[k]['impr']>=2000 and jul[k]['impr']>=2000]
fad=[]
for k in com:
    a,b=jun[k],jul[k]; dh=b['hook']-a['hook']; dro=b['roi']-a['roi']
    if dh<=-3 and dro<0: ver='Fadiga real'
    elif dh<=-3: ver='Hook caindo'
    elif dh>=3: ver='Subindo'
    else: ver='Estável'
    fad.append(dict(id=k,nome=a['nome'],creator=a['creator'],hj=a['hook'],hl=b['hook'],dh=dh,rj=a['roi'],rl=b['roi'],ver=ver))
fad.sort(key=lambda z:z['dh'])
title(ws,'Fadiga de Hook — Junho → Julho/26','Mesmo criativo nas 2 janelas (reach ≥2.000 em ambas). Hook caindo + ROI caindo = fadiga real. Só hook caindo ≠ problema. Δ em pontos percentuais (pp).')
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','Hook jun','Hook jul','Δ hook (pp)','ROI jun','ROI jul','Veredito'])
f=HR+1
cor={'Fadiga real':'C62828','Hook caindo':'B8860B','Subindo':'2E7D32','Estável':'888888'}
for k,x in enumerate(fad):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=x['hj']).number_format=P1
    ws.cell(row=rr,column=5,value=x['hl']).number_format=P1
    ws.cell(row=rr,column=6,value=round(x['dh'],1)).number_format='+0.0;-0.0'
    ws.cell(row=rr,column=7,value=round(x['rj'],2)).number_format=ROI2
    ws.cell(row=rr,column=8,value=round(x['rl'],2)).number_format=ROI2
    v=ws.cell(row=rr,column=9,value=x['ver']); v.font=Font(bold=True,color=cor[x['ver']])
    if k%2: zebra(ws,rr,9)
last=f+len(fad)-1
cscale(ws,f'F{f}:F{last}')
for col,wd in zip('ABCDEFGHI',[21,40,16,10,10,13,9,9,14]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:I{last}'

# ===== base dos mapas "Melhores" (d90, reach>=2000) =====
mpop=[x for x in WINS['d90'].values() if x['impr']>=2000]
for x in mpop:
    x['body']=(x['p50']/x['hook']*100) if x['hook']>0 else 0   # de quem passou do hook, % que chega à metade
    x['fim']=(x['p100']/x['hook']*100) if x['hook']>0 else 0
def _pr(vals):
    s=sorted(vals); n=len(s) or 1
    return lambda v:(sum(1 for y in s if y<=v)/n)
_rroi=_pr([x['roi'] for x in mpop]); _rhk=_pr([x['hook'] for x in mpop]); _rbd=_pr([x['body'] for x in mpop]); _rct=_pr([x['ctr'] for x in mpop])
for x in mpop: x['score']=round((_rroi(x['roi'])+_rhk(x['hook'])+_rbd(x['body'])+_rct(x['ctr']))/4*100)

# ================= ABA: MELHORES VÍDEOS (pacote completo, d90) =================
ws=wb.create_sheet('Melhores Vídeos'); ws.sheet_view.showGridLines=False
title(ws,'Melhores Vídeos — o pacote completo (90d)','Score = média dos percentis em ROI + Hook + Body + CTA. Surge quem é forte em TUDO, não só num eixo. Reach ≥2.000. Clique no ID.')
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','Score','ROI','Hook 2s','Body (meio)','CTR (CTA)','CVR','GMV','Custo'])
data=sorted(mpop,key=lambda z:-z['score']); f=HR+1
for k,x in enumerate(data):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=x['score'])
    ws.cell(row=rr,column=5,value=round(x['roi'],2)).number_format=ROI2
    ws.cell(row=rr,column=6,value=x['hook']).number_format=P1
    ws.cell(row=rr,column=7,value=round(x['body'])).number_format='0"%"'
    ws.cell(row=rr,column=8,value=x['ctr']).number_format=P2
    ws.cell(row=rr,column=9,value=x['conv']).number_format=P2
    ws.cell(row=rr,column=10,value=round(x['receita'])).number_format=M0
    ws.cell(row=rr,column=11,value=round(x['custo'])).number_format=M0
    if k%2: zebra(ws,rr,11)
last=f+len(data)-1
cscale(ws,f'D{f}:D{last}'); cscale(ws,f'E{f}:E{last}'); cscale(ws,f'F{f}:F{last}'); cscale(ws,f'H{f}:H{last}')
for col,wd in zip('ABCDEFGHIJK',[21,40,16,7,7,10,11,10,9,11,11]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:K{last}'

# ================= ABA 6: MELHORES HOOKS (d90) =================
ws=wb.create_sheet('Melhores Hooks'); ws.sheet_view.showGridLines=False
pop=[x for x in WINS['d90'].values() if x['impr']>=2000]
for x in pop: x['ret']=(x['s6']/x['hook']*100) if x['hook']>0 else 0
pop.sort(key=lambda z:-z['hook'])
mh=statistics.median([x['hook'] for x in pop]); m6=statistics.median([x['s6'] for x in pop])
title(ws,'Melhores Hooks — 90 dias (%s)'%PER['d90'],'Ordenado por Hook 2s. Reach ≥2.000. Retenção 6s/2s = de quem foi fisgado, quantos %% seguem aos 6s. Medianas: hook %.1f%% · 6s %.1f%%. Clique no ID.'%(mh,m6))
HR=4; H(ws,HR,['ID do vídeo','Criativo (o gancho)','Creator','Reach','Hook 2s','Hold 6s','Retenção 6s/2s','Custo','ROI'])
f=HR+1
for k,x in enumerate(pop):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=round(x['impr'])).number_format=INT
    ws.cell(row=rr,column=5,value=x['hook']).number_format=P1
    ws.cell(row=rr,column=6,value=x['s6']).number_format=P1
    ws.cell(row=rr,column=7,value=round(x['ret'])).number_format='0"%"'
    ws.cell(row=rr,column=8,value=round(x['custo'])).number_format=M0
    ws.cell(row=rr,column=9,value=round(x['roi'],2)).number_format=ROI2
    if k%2: zebra(ws,rr,9)
last=f+len(pop)-1
cscale(ws,f'E{f}:E{last}'); cscale(ws,f'F{f}:F{last}'); cscale(ws,f'G{f}:G{last}')
for col,wd in zip('ABCDEFGHI',[21,46,16,12,10,10,14,10,7]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:I{last}'

# ================= ABA: MELHORES BODIES (d90) =================
ws=wb.create_sheet('Melhores Bodies'); ws.sheet_view.showGridLines=False
title(ws,'Melhores Bodies — o corpo/demonstração que SEGURA (90d)','Body-hold = de quem passou do hook, % que chega à METADE do vídeo. Isola o meio do hook. Reach ≥2.000. Clique no ID.')
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','Hook 2s','6s','25%','50%','75%','Body-hold','ROI'])
data=sorted(mpop,key=lambda z:-z['body']); f=HR+1
for k,x in enumerate(data):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=x['hook']).number_format=P1
    ws.cell(row=rr,column=5,value=x['s6']).number_format=P1
    ws.cell(row=rr,column=6,value=x['p25']).number_format=P1
    ws.cell(row=rr,column=7,value=x['p50']).number_format=P1
    ws.cell(row=rr,column=8,value=x['p75']).number_format=P1
    ws.cell(row=rr,column=9,value=round(x['body'])).number_format='0"%"'
    ws.cell(row=rr,column=10,value=round(x['roi'],2)).number_format=ROI2
    if k%2: zebra(ws,rr,10)
last=f+len(data)-1
cscale(ws,f'I{f}:I{last}'); cscale(ws,f'G{f}:G{last}')
for col,wd in zip('ABCDEFGHIJ',[21,42,16,9,8,8,8,8,11,8]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:J{last}'

# ================= ABA: MELHORES CTAs (d90) =================
ws=wb.create_sheet('Melhores CTAs'); ws.sheet_view.showGridLines=False
title(ws,'Melhores CTAs — o que gera o clique (90d)','CTR = força do CTA (o clique). CVR = a oferta fecha. Completude = % dos hooked que chegam ao fim. CTR alto + CVR baixa = vídeo funciona, oferta trava → re-testar. Reach ≥2.000.')
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','CTR (CTA)','CVR (fecha)','Completude fim','CPA','ROI'])
data=sorted(mpop,key=lambda z:-z['ctr']); f=HR+1
for k,x in enumerate(data):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=x['ctr']).number_format=P2
    ws.cell(row=rr,column=5,value=x['conv']).number_format=P2
    ws.cell(row=rr,column=6,value=round(x['fim'])).number_format='0"%"'
    ws.cell(row=rr,column=7,value=x['cpp']).number_format=MONEY
    ws.cell(row=rr,column=8,value=round(x['roi'],2)).number_format=ROI2
    if k%2: zebra(ws,rr,8)
last=f+len(data)-1
cscale(ws,f'D{f}:D{last}'); cscale(ws,f'E{f}:E{last}'); cscale(ws,f'G{f}:G{last}',rev=True)
for col,wd in zip('ABCDEFGH',[21,42,16,11,11,13,11,8]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:H{last}'

# ================= ABA 7: RADAR POTENCIAL (jul26) =================
ws=wb.create_sheet('Radar Potencial'); ws.sheet_view.showGridLines=False
pj=[x for x in WINS['jul26'].values() if x['impr']>=2000]
mhj=statistics.median([x['hook'] for x in pj]); mcj=statistics.median([x['ctr'] for x in pj]); mvj=statistics.median([x['conv'] for x in pj])
radar=[x for x in pj if x['custo']<150 and x['hook']>=mhj and x['ctr']>=mcj]
for x in radar:
    x['score']=round(x['hook']/mhj+x['ctr']/mcj+x['conv']/mvj,2)
    x['ver']='Escalar (converte)' if x['conv']>=mvj else 'Testar oferta/link'
radar.sort(key=lambda z:-z['score'])
title(ws,'Radar de Potencial — Julho/26 (%s)'%PER['jul26'],'Hook ≥ mediana %.1f%% + CTR ≥ %.2f%% (CTA forte) + reach ≥2.000 + gasto <R$150. O vídeo já funciona, falta budget. Clique no ID.'%(mhj,mcj))
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','Custo','Reach','Hook 2s','CTR (CTA)','Conversão','Pedidos','Score','Veredito'])
f=HR+1
for k,x in enumerate(radar):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome']); ws.cell(row=rr,column=3,value=x['creator'][:16])
    ws.cell(row=rr,column=4,value=round(x['custo'])).number_format=M0
    ws.cell(row=rr,column=5,value=round(x['impr'])).number_format=INT
    ws.cell(row=rr,column=6,value=x['hook']).number_format=P1
    ws.cell(row=rr,column=7,value=x['ctr']).number_format=P2
    ws.cell(row=rr,column=8,value=x['conv']).number_format=P2
    ws.cell(row=rr,column=9,value=x['pedidos'])
    ws.cell(row=rr,column=10,value=x['score']).number_format=ROI2
    v=ws.cell(row=rr,column=11,value=x['ver']); v.font=Font(bold=True,color='2E7D32' if 'Escalar' in x['ver'] else 'B8860B')
    if k%2: zebra(ws,rr,11)
last=f+len(radar)-1
if radar:
    cscale(ws,f'G{f}:G{last}'); cscale(ws,f'H{f}:H{last}'); cscale(ws,f'J{f}:J{last}')
for col,wd in zip('ABCDEFGHIJK',[21,40,16,11,12,10,11,11,8,8,18]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'; ws.auto_filter.ref=f'A{HR}:K{last}'

# ================= ABAS 8-10: LASTRO / ONDAS / REALOCAÇÃO (jul26) =================
def b6(B0,ROI0): return B0*2**(math.log(6.0/ROI0)/math.log(1-DECAY)) if ROI0>6 else B0
def lastro(B0,ROI0,T): return 0.0 if ROI0<=T else B0*2**(math.log(T/ROI0)/math.log(1-DECAY))-B0
win=[x for x in WINS['jul26'].values() if x['roi']>=6 and x['custo']>=100 and x['pedidos']>=8]
win.sort(key=lambda z:-lastro(z['custo'],z['roi'],6.0))
flag=WINS['jul26']['7610417484289690901']
TCj=sum(x['custo'] for x in WINS['jul26'].values() if x['custo']>0)
TRj=sum(x['receita'] for x in WINS['jul26'].values() if x['custo']>0)

# ---- Lastro Crescimento ----
ws=wb.create_sheet('Lastro Crescimento'); ws.sheet_view.showGridLines=False
title(ws,'Lastro de Crescimento — Julho/26','Quanto cada criativo aguenta segurando o ROI. Base confiável (gasto ≥R$100, ≥8 pedidos). Decaimento %d%% por dobra (edite a premissa). Clique no ID.'%(int(DECAY*100)))
ws['A4']='PREMISSAS (amarelas)'; ws['A4'].font=Font(bold=True,size=11,color=RED)
for cc,val,fm,lab in [('B5',DECAY,PCT,'Decaimento de ROI por dobra'),('B6',7,ROI2,'ROI alvo A'),('B7',6,ROI2,'ROI alvo B'),('B8',3.0,MUL,'Teto de escala por onda')]:
    ws[cc]=val; ws[cc].fill=PatternFill('solid',fgColor=YEL); ws[cc].font=Font(bold=True); ws[cc].number_format=fm
    ws['A'+cc[1:]]=lab
HR=10; H(ws,HR,['ID do vídeo','Criativo','Creator','ROI atual','Custo atual','Reach','Lastro @Alvo B','Lastro @Alvo A','Passo (onda-1)','Novo budget'])
f=HR+1
for k,x in enumerate(win):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome'][:40]); ws.cell(row=rr,column=3,value=x['creator'][:15])
    ws.cell(row=rr,column=4,value=round(x['roi'],2)).number_format=ROI2
    ws.cell(row=rr,column=5,value=round(x['custo'])).number_format=M0
    ws.cell(row=rr,column=6,value=round(x['impr'])).number_format=INT
    ws.cell(row=rr,column=7,value=f'=IF(D{rr}<=$B$7,0,E{rr}*2^(LN($B$7/D{rr})/LN(1-$B$5))-E{rr})').number_format=M0
    ws.cell(row=rr,column=8,value=f'=IF(D{rr}<=$B$6,0,E{rr}*2^(LN($B$6/D{rr})/LN(1-$B$5))-E{rr})').number_format=M0
    ws.cell(row=rr,column=9,value=f'=MIN(G{rr},E{rr}*($B$8-1))').number_format=M0
    ws.cell(row=rr,column=10,value=f'=E{rr}+I{rr}').number_format=M0
    if k%2: zebra(ws,rr,10)
last=f+len(win)-1; tr=last+1
ws.cell(row=tr,column=2,value='TOTAL').font=Font(bold=True)
for let in 'GHI':
    c=ws.cell(row=tr,column=ord(let)-64,value=f'=SUM({let}{f}:{let}{last})'); c.font=Font(bold=True,color=RED); c.number_format=M0
fr=tr+2
ws.cell(row=fr,column=1,value='FLAGSHIP (contraste)').font=Font(bold=True,color=RED)
idlink(ws,fr+1,1,flag['id']); ws.cell(row=fr+1,column=2,value=flag['nome']); ws.cell(row=fr+1,column=4,value=round(flag['roi'],2)).number_format=ROI2; ws.cell(row=fr+1,column=5,value=round(flag['custo'])).number_format=M0
ws.cell(row=fr+2,column=2,value='Budget eficiente p/ ROI ≥ Alvo B:')
c=ws.cell(row=fr+2,column=5,value=f'=IF(D{fr+1}<=$B$7,E{fr+1},E{fr+1}*2^(LN($B$7/D{fr+1})/LN(1-$B$5)))'); c.number_format=M0; c.font=Font(bold=True,color=RED)
for col,wd in zip('ABCDEFGHIJ',[21,40,15,10,12,12,13,13,14,13]): ws.column_dimensions[col].width=wd
ws.freeze_panes='B11'

# ---- Plano de Ondas ----
ws=wb.create_sheet('Plano de Ondas'); ws.sheet_view.showGridLines=False
title(ws,'Plano de Ondas — Julho/26','Cada onda dobra o budget, com teto no budget eficiente de ROI 6. GATE: só avança se o ROI segurar (3–4 dias). Premissas puxadas de "Lastro Crescimento".')
HR=4; H(ws,HR,['ID do vídeo','Criativo','Creator','ROI','Budget hoje','Teto (ROI 6)','Onda 1','Onda 2','Onda 3'])
TRr="'Lastro Crescimento'!$B$7"; DRr="'Lastro Crescimento'!$B$5"; f=HR+1
for k,x in enumerate(win):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome'][:40]); ws.cell(row=rr,column=3,value=x['creator'][:15])
    ws.cell(row=rr,column=4,value=round(x['roi'],2)).number_format=ROI2
    ws.cell(row=rr,column=5,value=round(x['custo'])).number_format=M0
    ws.cell(row=rr,column=6,value=f'=IF(D{rr}<={TRr},E{rr},E{rr}*2^(LN({TRr}/D{rr})/LN(1-{DRr})))').number_format=M0
    ws.cell(row=rr,column=7,value=f'=MIN(2*E{rr},F{rr})').number_format=M0
    ws.cell(row=rr,column=8,value=f'=MIN(2*G{rr},F{rr})').number_format=M0
    ws.cell(row=rr,column=9,value=f'=MIN(2*H{rr},F{rr})').number_format=M0
    if k%2: zebra(ws,rr,9)
last=f+len(win)-1; tt=last+1
ws.cell(row=tt,column=2,value='TOTAL').font=Font(bold=True)
for let in 'EFGHI':
    c=ws.cell(row=tt,column=ord(let)-64,value=f'=SUM({let}{f}:{let}{last})'); c.font=Font(bold=True,color=RED); c.number_format=M0
for col,wd in zip('ABCDEFGHI',[21,40,15,9,12,12,11,11,11]): ws.column_dimensions[col].width=wd
ws.freeze_panes=f'B{f}'

# ---- Realocação Flagship ----
ws=wb.create_sheet('Realocação Flagship'); ws.sheet_view.showGridLines=False
title(ws,'Realocação de Budget — Flagship → Winners (Julho/26)','Tira verba do flagship saturado e alimenta os de alto ROI. Budget total constante. Edite as amarelas.')
ws['A4']='% do budget do flagship a realocar'; ws['B4']=0.40; ws['B4'].fill=PatternFill('solid',fgColor=YEL); ws['B4'].font=Font(bold=True); ws['B4'].number_format=PCT
ws['A5']='ROI mantido pelos receptores ao escalar'; ws['B5']=0.65; ws['B5'].fill=PatternFill('solid',fgColor=YEL); ws['B5'].font=Font(bold=True); ws['B5'].number_format=PCT
ws['A7']='FLAGSHIP'; ws['A7'].font=Font(bold=True,color=RED)
H(ws,8,['ID do vídeo','Criativo','Custo atual','ROI','Pedidos','Receita'])
idlink(ws,9,1,flag['id']); ws['B9']=flag['nome']; ws['C9']=flag['custo']; ws['D9']=flag['roi']; ws['E9']=flag['pedidos']; ws['F9']=flag['receita']
for cc,fm in [('C9',MONEY),('D9',ROI2),('F9',MONEY)]: ws[cc].number_format=fm
ws['A11']='Verba realocada'; ws['C11']='=C9*B4'; ws['C11'].number_format=MONEY; ws['C11'].font=Font(bold=True,color=RED)
ws['A12']='Receita perdida ao cortar'; ws['C12']='=C11*D9'; ws['C12'].number_format=MONEY; ws['C12'].font=Font(color=RED)
ws['A14']='RECEPTORES (peso = ROI, editável)'; ws['A14'].font=Font(bold=True,color=RED)
H(ws,15,['ID do vídeo','Criativo','Creator','ROI','Custo atual','Peso','Verba extra','Nova verba','Multipl.','Receita extra'])
f=16
for k,x in enumerate(win):
    rr=f+k
    idlink(ws,rr,1,x['id']); ws.cell(row=rr,column=2,value=x['nome'][:38]); ws.cell(row=rr,column=3,value=x['creator'][:15])
    ws.cell(row=rr,column=4,value=round(x['roi'],2)).number_format=ROI2
    ws.cell(row=rr,column=5,value=round(x['custo'],2)).number_format=MONEY
    pc_=ws.cell(row=rr,column=6,value=round(x['roi'],2)); pc_.number_format=ROI2; pc_.fill=PatternFill('solid',fgColor=YEL)
    ws.cell(row=rr,column=7,value=f'=$C$11*F{rr}/SUM($F$16:$F${f+len(win)-1})').number_format=MONEY
    ws.cell(row=rr,column=8,value=f'=E{rr}+G{rr}').number_format=MONEY
    ws.cell(row=rr,column=9,value=f'=H{rr}/E{rr}').number_format=MUL
    ws.cell(row=rr,column=10,value=f'=G{rr}*D{rr}*$B$5').number_format=MONEY
    if k%2: zebra(ws,rr,10)
tot=f+len(win)
ws.cell(row=tot,column=2,value='TOTAL').font=Font(bold=True)
ws.cell(row=tot,column=7,value=f'=SUM(G{f}:G{tot-1})').number_format=MONEY; ws.cell(row=tot,column=7).font=Font(bold=True)
ws.cell(row=tot,column=10,value=f'=SUM(J{f}:J{tot-1})').number_format=MONEY; ws.cell(row=tot,column=10).font=Font(bold=True)
P=tot+2
ws.cell(row=P,column=1,value=f'PROJEÇÃO (budget constante = R$ {TCj:,.0f})').font=Font(bold=True,color=RED)
ws.cell(row=P+1,column=1,value='Ganho líquido'); c=ws.cell(row=P+1,column=3,value=f'=J{tot}-C12'); c.number_format=MONEY; c.font=Font(bold=True,size=12,color=RED)
ws.cell(row=P+2,column=1,value='Receita ANTES'); ws.cell(row=P+2,column=3,value=round(TRj)).number_format=M0
ws.cell(row=P+3,column=1,value='Receita DEPOIS'); c=ws.cell(row=P+3,column=3,value=f'=C{P+2}+C{P+1}'); c.number_format=M0; c.font=Font(bold=True,color=RED)
ws.cell(row=P+4,column=1,value='ROI campanha ANTES→DEPOIS'); ws.cell(row=P+4,column=3,value=f'=C{P+2}/{TCj:.2f}').number_format=ROI2; ws.cell(row=P+4,column=4,value=f'=C{P+3}/{TCj:.2f}').number_format=ROI2
for col,wd in zip('ABCDEFGHIJ',[30,34,15,9,12,10,12,12,10,14]): ws.column_dimensions[col].width=wd

wb.save(DESK); shutil.copy(DESK,REPO)
print('OK. Abas:', wb.sheetnames)
print('Winners jul26:', len(win), '| Radar:', len(radar), '| Fadiga:', len(fad), '| Melhores hooks:', len(pop))
