#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SKU × CANAL — Agosto: uma aba por canal (Live afiliada · Vídeo afiliada · Live loja · Vídeo loja · Card),
com SKU, base julho, meta agosto (crescimento editável), faturamento, alavanca e PACE (realizado MTD).
Re-rodável DIÁRIO → o pace atualiza (fonte: pedidos_sku × extrato, coletor diário). Hero/esteira EXATO (REF).
Canais afiliados = exatos (extrato content_type). Canais vendedor = split calibrado pelo export do Seller Center.
Uso: python3 gerar_sku_por_canal.py [AAAA-MM-DD versão]
"""
import os, sys, json, urllib.request, datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
ROOT=os.path.dirname(os.path.abspath(__file__))   # script mora no ROOT do repo (committado, roda no cron)
load_dotenv(os.path.join(ROOT,".env"))
BASE_PER="2026-07"; BASE_CUT="2026-07-31"   # mês-base do plano (editar quando virar o plano)
SB=os.environ["SUPABASE_URL"];SK=os.environ["SUPABASE_SERVICE_KEY"];H={"apikey":SK,"Authorization":f"Bearer {SK}"}
DATE=sys.argv[1] if len(sys.argv)>1 else datetime.date.today().isoformat()
HOJE=datetime.date.today().isoformat()
def pg(b,order="id"):
    out=[];off=0
    while True:
        p=json.load(urllib.request.urlopen(urllib.request.Request(f"{SB}/rest/v1/{b}&order={order}&limit=1000&offset={off}",headers=H)))
        out+=p
        if len(p)<1000:break
        off+=1000
    return out
def canc(s): return bool(s) and 'CANCEL' in str(s).upper()
# proporções vendedor (do export TikTok "Análise de produtos", jul) — calibra o não-afiliado
LV,VV,CARD=177660.0,5090.0,19216.0; S=LV+VV+CARD
PV={"live_loja":LV/S,"video_loja":VV/S,"card":CARD/S}
def matrix(per,cut):
    raw=pg(f"pedidos_sku?periodo=eq.{per}&select=id,order_id,qty,gmv,seller_sku,data,status")
    seen={};[seen.setdefault(r['id'],r) for r in raw]
    psku=[r for r in seen.values() if (r.get('data') or '')[:10]<=cut and not canc(r.get('status'))]
    ex=pg(f"extrato_pedidos?periodo=eq.{per}&select=id,order_id,content_type,gmv,data,status")
    exs={};[exs.setdefault(r['id'],r) for r in ex]
    octg=defaultdict(lambda:defaultdict(float))
    for r in exs.values():
        if not canc(r.get('status')) and (r.get('data') or '')[:10]<=cut and r.get('content_type'):
            octg[str(r.get('order_id'))][r['content_type']]+=float(r.get('gmv') or 0)
    oct_={o:max(c,key=c.get) for o,c in octg.items()}
    q=defaultdict(lambda:defaultdict(float));g=defaultdict(lambda:defaultdict(float))
    for r in psku:
        ref=str(r.get('seller_sku') or '')[:6];o=str(r.get('order_id'));qty=float(r.get('qty') or 0);gmv=float(r.get('gmv') or 0)
        ct=oct_.get(o)
        if ct=='VIDEO': q[ref]["video_afil"]+=qty;g[ref]["video_afil"]+=gmv
        elif ct=='LIVE': q[ref]["live_afil"]+=qty;g[ref]["live_afil"]+=gmv
        else:
            for ch,p in PV.items(): q[ref][ch]+=qty*p;g[ref][ch]+=gmv*p
    return q,g
qJ,gJ=matrix(BASE_PER,BASE_CUT)                           # base (mês-plano)
mtdPer=datetime.date.today().strftime("%Y-%m"); qM,gM=matrix(mtdPer,HOJE)   # realizado MTD (mês corrente, dinâmico)
DIA=int(HOJE[-2:]); DIAS_MES=31

CH=[("live_afil","Live afiliada",0.10,"Rhode Squad — squad estratégico de creators de live"),
    ("video_afil","Vídeo afiliada",0.30,"Copa Rhode Creators (gamificação) + Academia — puxa volume de vídeo"),
    ("live_loja","Live loja (própria)",0.15,"Escalar sessões próprias + restaurar mídia de live"),
    ("video_loja","Vídeo loja (vendedor)",0.50,"Residual — vídeo próprio do seller (quase zero hoje)"),
    ("card","Card de produto",0.20,"Holdout do card → então reabrir mídia")]
HERO={"REF516","REF525","REF527"}
NM={"REF516":"Marmorizada","REF525":"Stone Used","REF527":"Sky Used","REF588":"Mom Marmorizada","REF587":"Baggy Marmorizada","REF551":"Azul Médio","REF562":"Chocolate","REF549":"Rosa Bebê","REF529":"Wide 529","REF528":"Wide 528","REF550":"Amarela","REF559":"Mom Rosa","REF566":"Mom Choco"}
def refs_do_canal(k,src=qJ,mn=8): return sorted([r for r in src if src[r][k]>=mn],key=lambda r:-src[r][k])

RED="C0392B";DARK="1A1A1A";GREY="7A7A7A";GREEN="1E7E34";GREEN_BG="E8F5E9";REDBG="F8D7DA";HEROBG="FDECEA";ESTBG="EAF2FB";YEL="FFF7CC";LIGHT="F2F2F2";BLUE="0000FF"
def Fn(s=10,b=False,c=DARK): return Font(name="Arial",size=s,bold=b,color=c)
def fill(c): return PatternFill("solid",fgColor=c)
CT=Alignment("center",vertical="center");RT=Alignment("right",vertical="center");LT=Alignment("left",vertical="center",wrap_text=True)
thin=Side(style="thin",color="D9D9D9");BOX=Border(left=thin,right=thin,top=thin,bottom=thin);NUM='#,##0';BRL='"R$" #,##0';PCT='0.0%';RATE='#,##0"/d"'
wb=Workbook()

# ===== RESUMO =====
w=wb.active;w.title="Resumo";w.sheet_view.showGridLines=False
for i,x in enumerate([22,10,11,10,10,11,13],1): w.column_dimensions[get_column_letter(i)].width=x
w["A1"]="SKU × Canal — Agosto/2026 (metas + pace diário)";w["A1"].font=Fn(15,True,DARK);w.row_dimensions[1].height=22
w["A2"]=f"Base = julho. Meta = base × crescimento. Pace = realizado agosto MTD (≤{DIA:02d}/08, atualiza a cada rodada). Faturamento = GMV. Hero/esteira exato (REF).";w["A2"].font=Fn(9,False,GREY)
r=4
for j,h in enumerate(["Canal","Base pç","Meta pç","Meta R$","Real MTD","% meta","Falta"],1):
    c=w.cell(row=r,column=j,value=h);c.font=Fn(9,True,"FFFFFF");c.fill=fill(DARK);c.border=BOX;c.alignment=CT
r+=1;first=r
for k,nome,gr,_ in CH:
    base=sum(qJ[x][k] for x in qJ);metap=round(base*(1+gr));metar=round(sum(gJ[x][k] for x in gJ)*(1+gr));real=round(sum(qM[x][k] for x in qM))
    w.cell(row=r,column=1,value=nome).font=Fn(9,True);w.cell(row=r,column=1).border=BOX
    for col,val,fmt,fnt in [(2,round(base),NUM,Fn(9)),(3,metap,NUM,Fn(10,True)),(4,metar,BRL,Fn(9,False,GREEN)),(5,real,NUM,Fn(9,True,"008000"))]:
        c=w.cell(row=r,column=col,value=val);c.number_format=fmt;c.font=fnt;c.border=BOX;c.alignment=RT
    c=w.cell(row=r,column=6,value=f"=E{r}/C{r}");c.number_format=PCT;c.border=BOX;c.alignment=RT;c.font=Fn(9);c.fill=fill(GREEN_BG if (real/metap if metap else 0)>=DIA/DIAS_MES else REDBG)
    c=w.cell(row=r,column=7,value=f"=MAX(0,C{r}-E{r})");c.number_format=NUM;c.border=BOX;c.alignment=RT;c.font=Fn(9,True)
    r+=1
last=r-1
w.cell(row=r,column=1,value="NET").font=Fn(10,True);w.cell(row=r,column=1).border=BOX;w.cell(row=r,column=1).fill=fill(LIGHT)
for col in (2,3,4,5,7):
    L=get_column_letter(col);c=w.cell(row=r,column=col,value=f"=SUM({L}{first}:{L}{last})");c.number_format=(BRL if col==4 else NUM);c.font=Fn(10,True);c.border=BOX;c.alignment=RT;c.fill=fill(LIGHT)
r+=2
w.cell(row=r,column=1,value="Afiliada (live/vídeo) = exato (extrato). Vendedor (live loja/vídeo loja/card) = split calibrado pelo export TikTok 'Análise de produtos' (jul). Pace = pedidos_sku × extrato (coletor diário) — re-rode pra atualizar.").font=Fn(8,False,"555555");w.merge_cells(start_row=r,start_column=1,end_row=r+2,end_column=7);w.cell(row=r,column=1).alignment=LT;w.row_dimensions[r].height=40

# ===== ABA POR CANAL =====
def aba(k,nome,gr,alav):
    w=wb.create_sheet(nome[:31]);w.sheet_view.showGridLines=False
    for i,x in enumerate([9,17,7,9,9,9,12,9,8,8],1): w.column_dimensions[get_column_letter(i)].width=x
    w["A1"]=f"{nome} — SKU × meta × pace ({DATE})";w["A1"].font=Fn(14,True,DARK);w.row_dimensions[1].height=20
    w["A2"]=f"ALAVANCA: {alav}";w["A2"].font=Fn(9,True,RED)
    r=4;w.cell(row=r,column=1,value="Crescimento (editável):").font=Fn(9,True);gc=w.cell(row=r,column=2,value=gr);gc.font=Fn(10,False,BLUE);gc.number_format=PCT;gc.fill=fill(YEL);gc.alignment=CT;GC=f"B{r}";r+=2
    for j,h in enumerate(["SKU","Cor/modelo","Cl","Base pç","Meta pç","Meta R$","Preço","Real MTD","% meta","Falta"],1):
        c=w.cell(row=r,column=j,value=h);c.font=Fn(9,True,"FFFFFF");c.fill=fill(DARK);c.border=BOX;c.alignment=CT
    r+=1;first=r
    for ref in refs_do_canal(k):
        cl="H" if ref in HERO else "E";base=qJ[ref][k];preco=gJ[ref][k]/base if base else 0;real=qM[ref][k]
        w.cell(row=r,column=1,value=ref).font=Fn(9,True);w.cell(row=r,column=1).border=BOX
        c=w.cell(row=r,column=2,value=NM.get(ref,ref));c.font=Fn(9);c.border=BOX
        c=w.cell(row=r,column=3,value=cl);c.font=Fn(8,True);c.border=BOX;c.alignment=CT;c.fill=fill(HEROBG if cl=="H" else ESTBG)
        c=w.cell(row=r,column=4,value=round(base));c.number_format=NUM;c.font=Fn(9);c.border=BOX;c.alignment=RT
        c=w.cell(row=r,column=5,value=f"=ROUND(D{r}*(1+${GC}),0)");c.number_format=NUM;c.font=Fn(9,True,GREEN);c.border=BOX;c.alignment=RT;c.fill=fill(GREEN_BG)
        c=w.cell(row=r,column=6,value=f"=E{r}*G{r}");c.number_format=BRL;c.font=Fn(9);c.border=BOX;c.alignment=RT
        c=w.cell(row=r,column=7,value=round(preco,2));c.number_format=BRL;c.font=Fn(9,False,GREY);c.border=BOX;c.alignment=RT
        c=w.cell(row=r,column=8,value=round(real));c.number_format=NUM;c.font=Fn(9,True,"008000");c.border=BOX;c.alignment=RT
        c=w.cell(row=r,column=9,value=f"=IF(E{r}=0,0,H{r}/E{r})");c.number_format=PCT;c.font=Fn(9);c.border=BOX;c.alignment=RT
        c=w.cell(row=r,column=10,value=f"=MAX(0,E{r}-H{r})");c.number_format=NUM;c.font=Fn(9,True);c.border=BOX;c.alignment=RT
        r+=1
    last=r-1
    w.cell(row=r,column=1,value="TOTAL").font=Fn(10,True);w.cell(row=r,column=1).border=BOX;w.cell(row=r,column=1).fill=fill(LIGHT)
    for col in (4,5,6,8,10):
        L=get_column_letter(col);c=w.cell(row=r,column=col,value=f"=SUM({L}{first}:{L}{last})");c.number_format=(BRL if col==6 else NUM);c.font=Fn(10,True);c.border=BOX;c.alignment=RT;c.fill=fill(LIGHT)
    # hero/esteira subtotais
    r+=1
    for lbl,flt,col in [("Σ Hero","H",RED),("Σ Esteira","E",BLUE)]:
        rr=[i for i,ref in enumerate(refs_do_canal(k)) if (ref in HERO)==(flt=="H")]
        base_=sum(qJ[refs_do_canal(k)[i]][k] for i in rr)
        w.cell(row=r,column=1,value=lbl).font=Fn(9,True,col);w.cell(row=r,column=1).border=BOX
        c=w.cell(row=r,column=4,value=round(base_));c.number_format=NUM;c.font=Fn(9,True,col);c.border=BOX;c.alignment=RT
        r+=1
for k,nome,gr,alav in CH: aba(k,nome,gr,alav)

order=["Resumo"]+[n[:31] for _,n,_,_ in CH]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
OUTDIR=os.path.join(ROOT,"relatorios",mtdPer,"Projecao Agosto"); os.makedirs(OUTDIR,exist_ok=True)
wb.save(os.path.join(OUTDIR,f"SKU por Canal Agosto_{DATE}.xlsx"))         # cópia datada
wb.save(os.path.join(OUTDIR,"SKU por Canal Agosto_ATUAL.xlsx"))            # arquivo estável (pace do dia)
print("OK ->",os.path.join(OUTDIR,"SKU por Canal Agosto_ATUAL.xlsx"))
print("base jul:",{n:round(sum(qJ[x][k] for x in qJ)) for k,n,_,_ in CH})
print(f"realizado ago MTD (≤{HOJE}):",{n:round(sum(qM[x][k] for x in qM)) for k,n,_,_ in CH})
