#!/usr/bin/env python3
"""
Projeção do PRÓXIMO mês (rodar dia 25). Base = mês atual extrapolado (run-rate MTD)
+ crescimento por canal (editável na planilha). Escreve xlsx datado em
relatorios/<próximo-mês>/ + snapshot projecao_<próximo-mês>.json (pro relatório
geral do dia 05 comparar projetado vs realizado). Fonte: tracker_canais. ROOT.
Uso: python3 gerar_projecao_mensal.py [--mes AAAA-MM]  (default = mês atual → projeta o seguinte)
"""
import os, sys, json, urllib.request, datetime, calendar
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE=os.path.dirname(os.path.abspath(__file__)); load_dotenv(os.path.join(BASE,".env"))
SB=os.environ["SUPABASE_URL"]; SK=os.environ["SUPABASE_SERVICE_KEY"]
H={"apikey":SK,"Authorization":f"Bearer {SK}"}
def q(b): return json.load(urllib.request.urlopen(urllib.request.Request(f"{SB}/rest/v1/{b}",headers=H)))

META={"video_afiliada":3600,"live_afiliada":2800,"loja_propria":3600}; META_NET=10000
GROWTH={"video_afiliada":0.20,"live_afiliada":0.15,"loja_propria":0.05}  # default editável
RED="FE2C55"; INK="17151B"; WASH="FFECF0"
hf=PatternFill("solid",fgColor=RED); hfont=Font(bold=True,color="FFFFFF",size=11)
thin=Side(style="thin",color="D9D4CE"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

def cur_mes():
    for i,a in enumerate(sys.argv):
        if a=="--mes" and i+1<len(sys.argv): return sys.argv[i+1]
    return datetime.date.today().strftime("%Y-%m")

def main():
    cur=cur_mes(); y,m=int(cur[:4]),int(cur[5:7])
    dias_cur=calendar.monthrange(y,m)[1]
    hoje=datetime.date.today()
    dia=hoje.day if hoje.strftime("%Y-%m")==cur else dias_cur  # se rodar depois, usa mês fechado
    nxt=(datetime.date(y,m,1)+datetime.timedelta(days=32)).replace(day=1); nxtm=nxt.strftime("%Y-%m")

    tc=q(f"tracker_canais?mes=eq.{cur}&select=canal,pecas")
    by={r['canal']:r['pecas'] for r in tc}
    def full(c): return round(by.get(c,0)/dia*dias_cur) if dia else 0

    wb=openpyxl.Workbook(); ws=wb.active; ws.title=f"Projecao {nxtm}"
    ws.column_dimensions['A'].width=22
    for col in "BCDEFG": ws.column_dimensions[col].width=15
    ws['A1']=f"PROJEÇÃO {nxtm} — base run-rate de {cur}"; ws['A1'].font=Font(bold=True,size=14,color=RED); ws.merge_cells('A1:G1')
    ws['A2']=f"Base: {cur} extrapolado (MTD dia {dia}/{dias_cur}). Crescimento por canal EDITÁVEL (col D). Meta net {META_NET}. Gerado {hoje.isoformat()}."
    ws['A2'].font=Font(size=10,italic=True,color="6E6670"); ws.merge_cells('A2:G2')

    hdr=["Canal",f"{cur} realizado→full","Peças MTD","Crescimento (edit)",f"{nxtm} projetado","Meta",f"Gap vs meta"]
    for i,cn in enumerate(hdr):
        c=ws.cell(4,1+i,cn); c.fill=hf; c.font=hfont; c.border=bd; c.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[4].height=30
    canais=["video_afiliada","live_afiliada","loja_propria"]
    r=5; snap={"mes":nxtm,"canais":{},"net":0}
    for c in canais:
        f=full(c); g=GROWTH.get(c,0.1); proj=round(f*(1+g)); snap["canais"][c]=proj; snap["net"]+=proj
        ws.cell(r,1,c).border=bd
        ws.cell(r,2,f).border=bd; ws.cell(r,2).number_format="0"
        ws.cell(r,3,by.get(c,0)).border=bd
        gc=ws.cell(r,4,g); gc.number_format="0%"; gc.border=bd; gc.fill=PatternFill("solid",fgColor=WASH)
        ws.cell(r,5,f"=ROUND(B{r}*(1+D{r}),0)").border=bd
        ws.cell(r,6,META.get(c,0)).border=bd
        ws.cell(r,7,f"=E{r}-F{r}").border=bd; ws.cell(r,7).number_format="0;[Red]-0"
        r+=1
    ws.cell(r,1,"TOTAL").font=Font(bold=True); ws.cell(r,1).border=bd
    ws.cell(r,2,f"=SUM(B5:B{r-1})").border=bd
    ws.cell(r,3,f"=SUM(C5:C{r-1})").border=bd
    ws.cell(r,5,f"=SUM(E5:E{r-1})").font=Font(bold=True); ws.cell(r,5).border=bd; ws.cell(r,5).fill=PatternFill("solid",fgColor=WASH)
    ws.cell(r,6,META_NET).border=bd
    ws.cell(r,7,f"=E{r}-F{r}").border=bd; ws.cell(r,7).number_format="0;[Red]-0"
    r+=2
    ws.cell(r,1,"Editar a col D (crescimento) recalcula o projetado e o gap. Default: vídeo +20%, live afiliada +15%, loja +5% (empurrar afiliada/esteira).")
    ws.cell(r,1).font=Font(size=10,italic=True,color="6E6670"); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=7)

    outdir=os.path.join(BASE,"relatorios",nxtm); os.makedirs(outdir,exist_ok=True)
    out=os.path.join(outdir,f"Relatorio Projecao {nxtm}_{hoje.isoformat()}.xlsx")
    wb.save(out)
    json.dump(snap,open(os.path.join(outdir,f"projecao_{nxtm}.json"),"w"),ensure_ascii=False,indent=1)
    print(f"✓ projeção {nxtm}: {out}")
    print(f"  net projetado {snap['net']} vs meta {META_NET}")
    return out

if __name__=="__main__": main()
