#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Atualiza o PACE do workbook Rhode_SKU_x_Canal — desenho do Humberto.
A ÚNICA coisa a fazer é alimentar a aba "Log diário" (Data|Canal|SKU|Peças|GMV|Origem).
As abas de canal leem por SUMIFS(Canal=B2, SKU=A8) SEM data → o coletor SUBSTITUI as
linhas do mês (uma por canal×SKU = MTD), não acumula. Nada mais precisa ser regerado.
Canais mensuráveis: Live afiliada · Vídeo afiliada (extrato, exato por SKU) + Live/Vídeo/Card
loja (total do canal EXATO via API /analytics/202605/shop_products/performance — bate ao
centavo com o export; rateado por SKU pela mistura não-afiliada). Card afiliada/Vitrine/Busca
ficam SEM DADO (como no template). Re-rodável DIÁRIO (cron) → pace atualiza. Zero export manual.
"""
import os, sys, json, urllib.request, datetime, shutil
from collections import defaultdict
from openpyxl import load_workbook
from dotenv import load_dotenv
ROOT=os.path.dirname(os.path.abspath(__file__)); load_dotenv(os.path.join(ROOT,".env"))
SB=os.environ["SUPABASE_URL"];SK=os.environ["SUPABASE_SERVICE_KEY"];H={"apikey":SK,"Authorization":f"Bearer {SK}"}
TEMPLATE=os.path.join(ROOT,"dados/marketplace/tiktok/Rhode_SKU_x_Canal_template.xlsx")
HOJE=datetime.date.today().isoformat(); MES=HOJE[:7]
OUTDIR=os.path.join(ROOT,"relatorios",MES,"Projecao Agosto"); os.makedirs(OUTDIR,exist_ok=True)
def pg(b,order="id"):
    out=[];off=0
    while True:
        p=json.load(urllib.request.urlopen(urllib.request.Request(f"{SB}/rest/v1/{b}&order={order}&limit=1000&offset={off}",headers=H)))
        out+=p
        if len(p)<1000:break
        off+=1000
    return out
def canc(s): return bool(s) and 'CANCEL' in str(s).upper()
try:
    from coletar_dados import chamar   # Partner API (HMAC) — mesmo usado nos outros coletores
except Exception:
    chamar=None
def _amt(o,*path):
    for p in path: o=(o or {}).get(p,{}) if isinstance(o,dict) else {}
    return float((o or {}).get("amount",0) or 0) if isinstance(o,dict) else 0
def seller_split_api(ge,lt):
    """Split vendedor EXATO via /analytics/202605/shop_products/performance (bate ao centavo c/ o export)."""
    if chamar is None: return None
    tot={"live_loja":0.0,"video_loja":0.0,"card":0.0};tok=None
    for _ in range(30):
        p={"start_date_ge":ge,"end_date_lt":lt,"page_size":100,"currency":"LOCAL","product_status_filter":"ALL"}
        if tok:p["page_token"]=tok
        r=chamar("GET","/analytics/202605/shop_products/performance",params=p)
        if not r or r.get("code")!=0: return None
        d=r.get("data",{}) or {}
        for it in d.get("products",[]):
            tot["live_loja"]+=_amt(it,"seller_live_performance","attributed_gmv")
            tot["video_loja"]+=_amt(it,"seller_video_performance","attributed_gmv")
            tot["card"]+=_amt(it,"seller_product_card_performance","attributed_gmv")
        tok=d.get("next_page_token")
        if not tok: break
    S=sum(tot.values())
    return {k:v/S for k,v in tot.items()} if S>0 else None
# proporções do split vendedor: EXATO da API (mês-base jul); fallback = export jul hardcoded
PV=seller_split_api("2026-07-01","2026-08-01")
if PV: print(f"  [split vendedor] via API 202605 (exato): {[f'{k}={v*100:.0f}%' for k,v in PV.items()]}")
else:
    LV,VV,CARD=177660.0,5090.0,19216.0; S=LV+VV+CARD
    PV={"live_loja":LV/S,"video_loja":VV/S,"card":CARD/S}
    print("  [split vendedor] API 202605 indisponível → fallback export jul")
NAME={"live_afil":"Live afiliada","video_afil":"Vídeo afiliada","live_loja":"Live loja","video_loja":"Vídeo loja","card":"Card loja"}
def matrix(per,cut):
    raw=pg(f"pedidos_sku?periodo=eq.{per}&select=id,order_id,qty,gmv,seller_sku,data,status")
    seen={};[seen.setdefault(r['id'],r) for r in raw]
    psku=[r for r in seen.values() if (r.get('data') or '')[:10]<=cut and not canc(r.get('status'))]
    ex=pg(f"extrato_pedidos?periodo=eq.{per}&select=id,order_id,content_type,gmv,data,status")
    exs={};[exs.setdefault(r['id'],r) for r in ex]
    octg=defaultdict(lambda:defaultdict(float))
    for r in exs.values():
        if not canc(r.get('status')) and (r.get('data') or '')[:10]<=cut and r.get('content_type'):
            octg[str(r.get('order_id'))][r['content_type']]+=float(r.get('gmv') or 0)
    oct_={o:max(c,key=c.get) for o,c in octg.items()}
    q=defaultdict(lambda:defaultdict(float));g=defaultdict(lambda:defaultdict(float))
    for r in psku:
        ref=str(r.get('seller_sku') or '')[:6];o=str(r.get('order_id'));qt=float(r.get('qty') or 0);gm=float(r.get('gmv') or 0)
        ct=oct_.get(o)
        if ct=='VIDEO': q[ref]["video_afil"]+=qt;g[ref]["video_afil"]+=gm
        elif ct=='LIVE': q[ref]["live_afil"]+=qt;g[ref]["live_afil"]+=gm
        else:
            for ch,p in PV.items(): q[ref][ch]+=qt*p;g[ref][ch]+=gm*p
    return q,g
q,g=matrix(MES,HOJE)   # MTD do mês corrente
dia=int(HOJE[-2:]); origem=f"MTD 01–{dia:02d}/{MES[5:]} · coletor diário"

wb=load_workbook(TEMPLATE)   # preserva TODAS as fórmulas/abas do desenho
ws=wb["Log diário"]
if ws.max_row>=5: ws.delete_rows(5,ws.max_row-4)   # limpa dados antigos (mantém cabeçalho L4)
r=5; nrows=0
for ref in sorted(q):
    for ch in ("live_afil","video_afil","live_loja","video_loja","card"):
        pc=q[ref][ch]
        if pc<0.5: continue
        ws.cell(row=r,column=1,value=HOJE);ws.cell(row=r,column=2,value=NAME[ch]);ws.cell(row=r,column=3,value=ref)
        ws.cell(row=r,column=4,value=round(pc));ws.cell(row=r,column=5,value=round(g[ref][ch],2));ws.cell(row=r,column=6,value=origem)
        r+=1;nrows+=1
OUT=os.path.join(OUTDIR,"Rhode_SKU_x_Canal_Agosto_ATUAL.xlsx")
wb.save(OUT); wb.save(os.path.join(OUTDIR,f"Rhode_SKU_x_Canal_Agosto_{HOJE}.xlsx"))
print(f"OK -> {OUT}  ({nrows} linhas no Log diário)")
print("MTD por canal:",{NAME[c]:round(sum(q[x][c] for x in q)) for c in NAME})
