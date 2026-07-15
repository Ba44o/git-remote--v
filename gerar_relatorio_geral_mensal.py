#!/usr/bin/env python3
"""
Relatório GERAL mensal (rodar dia 05) — fechamento do mês anterior.
KPIs: meta 10k · realizado (tracker_canais) · % meta · por canal realizado vs meta
      · projetado vs realizado (lê relatorios/<mes>/projecao_<mes>.json se existir)
      · top SKUs por canal. Salva xlsx d-atado em relatorios/<mes-fechado>/.
Fonte: tracker_canais (peças reais atribuídas por pedido). ROOT (não dispara etl_sync).
Uso: python3 gerar_relatorio_geral_mensal.py [--mes AAAA-MM]  (default = mês anterior)
"""
import os, sys, json, urllib.request, datetime, calendar
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

BASE=os.path.dirname(os.path.abspath(__file__)); load_dotenv(os.path.join(BASE,".env"))
SB=os.environ["SUPABASE_URL"]; SK=os.environ["SUPABASE_SERVICE_KEY"]
H={"apikey":SK,"Authorization":f"Bearer {SK}"}
def q(b): return json.load(urllib.request.urlopen(urllib.request.Request(f"{SB}/rest/v1/{b}",headers=H)))

META={"video_afiliada":3600,"live_afiliada":2800,"loja_propria":3600}; META_NET=10000
RED="FE2C55"; INK="17151B"; WASH="FFECF0"
hf=PatternFill("solid",fgColor=RED); hfont=Font(bold=True,color="FFFFFF",size=11)
thin=Side(style="thin",color="D9D4CE"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)

def target_mes():
    for i,a in enumerate(sys.argv):
        if a=="--mes" and i+1<len(sys.argv): return sys.argv[i+1]
    t=datetime.date.today().replace(day=1)-datetime.timedelta(days=1)
    return t.strftime("%Y-%m")

def main():
    mes=target_mes()
    tc=q(f"tracker_canais?mes=eq.{mes}&select=canal,pecas,meta_pecas,top_skus")
    by={r['canal']:r for r in tc}
    net=by.get('net',{}).get('pecas',0)
    proj=None
    pj=os.path.join(BASE,"relatorios",mes,f"projecao_{mes}.json")
    if os.path.exists(pj):
        try: proj=json.load(open(pj))
        except: proj=None

    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Fechamento"
    ws.column_dimensions['A'].width=26
    for col in "BCDEF": ws.column_dimensions[col].width=15
    ws['A1']=f"RELATÓRIO GERAL — {mes} (fechamento)"; ws['A1'].font=Font(bold=True,size=14,color=RED); ws.merge_cells('A1:F1')
    ws['A2']=f"Meta {META_NET} peças. Realizado (tracker_canais, peças reais atribuídas por pedido). Gerado {datetime.date.today().isoformat()}."
    ws['A2'].font=Font(size=10,italic=True,color="6E6670"); ws.merge_cells('A2:F2')

    # KPIs
    ws['A4']="KPI"; ws['B4']="Valor"
    for c in "AB": ws[f"{c}4"].fill=hf; ws[f"{c}4"].font=hfont; ws[f"{c}4"].border=bd
    kpis=[("Meta (peças)",META_NET,"0"),("Realizado (net)",net,"0"),
          ("% da meta",(net/META_NET if META_NET else 0),"0%"),
          ("Gap vs meta",net-META_NET,"0;[Red]-0")]
    r=5
    for lab,val,fmt in kpis:
        ws[f"A{r}"]=lab; ws[f"A{r}"].border=bd
        ws[f"B{r}"]=val; ws[f"B{r}"].number_format=fmt; ws[f"B{r}"].font=Font(bold=True); ws[f"B{r}"].border=bd
        r+=1
    ws.conditional_formatting.add(f"B7",CellIsRule(operator="lessThan",formula=["0.9"],fill=PatternFill("solid",fgColor="FFE1E7")))
    ws[f"B5"].fill=PatternFill("solid",fgColor=WASH)

    # Por canal: realizado vs meta (+ projetado se houver)
    hr=r+1
    cols=["Canal","Realizado","Meta","% meta"]+(["Projetado","Real−Proj"] if proj else [])
    for i,cn in enumerate(cols):
        cell=ws.cell(hr,1+i,cn); cell.fill=hf; cell.font=hfont; cell.border=bd
    r=hr+1
    for c in ["video_afiliada","live_afiliada","loja_propria"]:
        rz=by.get(c,{}).get('pecas',0); mt=META.get(c,0)
        ws.cell(r,1,c).border=bd; ws.cell(r,2,rz).border=bd
        ws.cell(r,3,mt).border=bd
        cell=ws.cell(r,4,(rz/mt if mt else 0)); cell.number_format="0%"; cell.border=bd
        if proj:
            pv=proj.get('canais',{}).get(c,0)
            ws.cell(r,5,pv).border=bd; d=ws.cell(r,6,rz-pv); d.number_format="0;[Red]-0"; d.border=bd
        r+=1
    ws.cell(r,1,"TOTAL (net)").font=Font(bold=True); ws.cell(r,1).border=bd
    ws.cell(r,2,net).font=Font(bold=True); ws.cell(r,2).border=bd
    ws.cell(r,3,META_NET).border=bd
    cell=ws.cell(r,4,(net/META_NET if META_NET else 0)); cell.number_format="0%"; cell.border=bd
    if proj:
        pn=proj.get('net',0); ws.cell(r,5,pn).border=bd; ws.cell(r,6,net-pn).number_format="0;[Red]-0"; ws.cell(r,6).border=bd

    # Top SKUs por canal
    r+=2; ws.cell(r,1,"Top SKUs por canal").font=Font(bold=True,color=RED); r+=1
    for c in ["video_afiliada","live_afiliada","loja_propria"]:
        ts=by.get(c,{}).get('top_skus') or []
        skus=", ".join(f"{t.get('sku')} ({t.get('pecas')})" for t in ts[:5]) if isinstance(ts,list) else ""
        ws.cell(r,1,c).border=bd; cc=ws.cell(r,2,skus); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); cc.border=bd
        r+=1

    outdir=os.path.join(BASE,"relatorios",mes); os.makedirs(outdir,exist_ok=True)
    out=os.path.join(outdir,f"Relatorio Geral Meta 10k {mes}_{datetime.date.today().isoformat()}.xlsx")
    wb.save(out)
    print(f"✓ relatório geral: {out}")
    print(f"  {mes}: realizado {net}/{META_NET} ({net/META_NET*100:.0f}%)")
    return out

if __name__=="__main__": main()
